"""
cli.py — точка входа командной строки (Click).

Все команды проекта:
- init-db
- load-data
- calculate
- process (минимальная версия)
- history

Запуск после `pip install -e .`:
    request-processor init-db

Или без установки:
    PYTHONPATH=src python -m cli init-db
"""

from __future__ import annotations
import re
import json
from pathlib import Path
from typing import Any, Optional

import click
from openpyxl import load_workbook

from request_processor.models import TestItemCreate
from request_processor.logging_setup import get_logger, setup_logging

from .models import ClimaticTestSettings

from .persistence.sqlite_repo import (
    init_db,
    load_price_list_from_xlsx,
    save_calculation,
    get_recent_calculations,
    list_test_items,
    add_test_item,
    bulk_upsert_test_items,
    build_default_hours_map,
    get_climatic_settings,
    save_climatic_settings,
    save_cable_marks_from_matches,
    list_cable_marks,
    list_organizations,
    migrate_db,
    save_document_extraction,
    save_organizations_from_extraction,
    create_order_from_kp,
    list_orders,
    get_order_details,
    get_last_document_extraction,
    list_test_applications,
    list_test_mappings,
    delete_test_mapping,
    update_test_mapping,
    add_test_mapping,
    list_generated_documents,
)
from .calculation.cost_calculator import calculate_cost, print_breakdown
from .generation.kp_generator import generate_kp_from_db
from .generation.application_generator import generate_application_from_order
from .extraction.pdf_extractor import extract_from_document
from .validation.extraction_validator import format_validation_report, validate_extraction
from .mapping.requirement_mapper import map_requirements_to_tests, suggest_tests_for_mark
from request_processor import __version__

def _slugify(text: str) -> str:
    """Делает безопасный код из названия (простая версия)."""
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "_", text)
    return text[:60]

@click.group()
@click.version_option(version=__version__)
@click.option(
    "--log-level",
    default="INFO",
    show_default=True,
    type=click.Choice(["DEBUG", "INFO", "WARNING", "ERROR"], case_sensitive=False),
    help="Уровень логов в консоли (в файл всегда DEBUG)",
)
@click.pass_context
def cli(ctx: click.Context, log_level: str) -> None:
    """request_processor — расчёт заявок на испытания кабельной продукции."""
    setup_logging(level=log_level)
    ctx.ensure_object(dict)
    ctx.obj["log"] = get_logger("cli")


@cli.command("init-db")
@click.option("--db", default="data/app.db", show_default=True, help="Путь к файлу SQLite")
def init_db_cmd(db: str) -> None:
    """Инициализирует базу данных и добавляет демо-тесты."""
    init_db(db)
    click.echo(click.style("✓ База данных инициализирована.", fg="green"))


@cli.command("migrate-db")
@click.option("--db", default="data/app.db", show_default=True)
def migrate_db_cmd(db: str) -> None:
    """Обновляет схему существующей БД (таблицы марок и настроек)."""
    migrate_db(db)
    click.echo(click.style("✓ Миграция выполнена.", fg="green"))


@cli.command("db-info")
@click.option("--db", default="data/app.db", show_default=True)
def db_info_cmd(db: str) -> None:
    """Показывает роль БД: dev / work_copy / work (см. docs/db_profile.example.yaml)."""
    from .persistence.db_profile import format_db_info, load_db_profile

    prof = load_db_profile(db)
    click.echo(format_db_info(db, profile=prof))
    if prof.is_dev_scratch:
        click.echo(
            click.style(
                "\n→ Это тестовая/неразмеченная БД. Не считайте org/заказы источником истины.",
                fg="yellow",
            )
        )
    elif prof.role == "work_copy":
        click.echo(
            click.style(
                "\n→ Копия с рабочего ПК: для данных — источник истины на этом компьютере.",
                fg="green",
            )
        )
    elif prof.role == "work":
        click.echo(click.style("\n→ Боевая БД рабочего ПК.", fg="green"))


@cli.command("db-role")
@click.option(
    "--set",
    "role",
    required=True,
    type=click.Choice(["dev", "work_copy", "work"], case_sensitive=False),
    help="Роль: dev | work_copy | work",
)
@click.option("--source", default="", help="Откуда БД, напр. «рабочий ПК USB 2026-07-28»")
@click.option("--notes", default="", help="Заметка для себя/агента")
@click.option("--label", default="", help="Короткая подпись в GUI (иначе по роли)")
@click.option("--db", default="data/app.db", show_default=True)
def db_role_cmd(role: str, source: str, notes: str, label: str, db: str) -> None:
    """Помечает активную data/app.db: тестовая или копия/рабочая.

    После копирования app.db с рабочего ПК на dev::

        request-processor db-role --set work_copy --source "рабочий ПК 2026-07-28"

    Вернуть scratch на dev::

        request-processor db-role --set dev --notes "локальные эксперименты"
    """
    from .persistence.db_profile import format_db_info, set_db_role

    prof = set_db_role(
        role.lower(),  # type: ignore[arg-type]
        db_path=db,
        source=source,
        notes=notes,
        label=label,
    )
    click.echo(click.style(f"✓ Роль БД: {prof.role} ({prof.ui_label()})", fg="green"))
    click.echo(format_db_info(db, profile=prof))


@cli.command("load-data")
@click.option("--price", required=True, type=click.Path(exists=True), help="Путь к прайс-листу .xlsx")
@click.option("--db", default="data/app.db", show_default=True)
def load_data_cmd(price: str, db: str) -> None:
    """Загружает прайс-лист в таблицу test_items."""
    click.echo(f"Загрузка прайс-листа: {price}")
    try:
        count = load_price_list_from_xlsx(price, db)
        click.echo(click.style(f"✓ Загружено {count} позиций.", fg="green"))
    except Exception as e:
        click.echo(click.style(f"Ошибка: {e}", fg="red"), err=True)


@cli.command("prepare-prod-db")
@click.option("--db", default="data/app.db", show_default=True)
@click.option(
    "--yes",
    "confirm",
    is_flag=True,
    help="Подтвердить очистку без интерактивного вопроса",
)
@click.option(
    "--no-backup",
    is_flag=True,
    help="Не создавать копию app.db.pre_prod_*.db",
)
def prepare_prod_db_cmd(db: str, confirm: bool, no_backup: bool) -> None:
    """Очищает марки и организации; прайс (test_items) и test_mappings оставляет.

    Для prod-установки на рабочем ПК: чистые справочники марок/орг.,
    прайс без изменений. Также удаляет заказы/расчёты/извлечения (ссылки на org/mark).
    Перед очисткой по умолчанию делает backup БД.
    """
    from .persistence.sqlite_repo import prepare_prod_db

    if not confirm:
        click.echo(
            "Будут удалены: cable_marks, organizations, orders, calculations, "
            "document_extractions, …\n"
            "Останутся: test_items (прайс), test_mappings, app_settings.\n"
            "Повторите с флагом --yes для выполнения."
        )
        raise SystemExit(1)
    try:
        result = prepare_prod_db(db, backup=not no_backup)
    except Exception as e:
        click.echo(click.style(f"Ошибка: {e}", fg="red"), err=True)
        raise SystemExit(1) from e
    click.echo(click.style("✓ БД подготовлена к prod", fg="green"))
    click.echo(f"  db: {result['db_path']}")
    if result.get("backup_path"):
        click.echo(f"  backup: {result['backup_path']}")
    click.echo(
        f"  сохранено: test_items={result['kept_test_items']}, "
        f"test_mappings={result['kept_test_mappings']}"
    )
    for table, n in (result.get("deleted") or {}).items():
        if n:
            click.echo(f"  удалено {table}: {n}")


@cli.command("calculate")
@click.option("--mark", required=True, help="Полная марка кабеля")
@click.option("--cores", type=int, default=1, show_default=True, help="Количество жил/элементов")
@click.option("--groups", type=int, default=1, show_default=True, help="Количество групп")
@click.option("--tests", required=True, help="Коды испытаний через запятую")
@click.option("--hour", "hours_list", multiple=True, help="Часы в формате ключ=значение, можно указывать несколько раз. Пример: --hour temp_low=48")
@click.option("--hours", default="{}", help='JSON-строка (устаревший способ). Лучше используй --hour')
@click.option("--qty", "qty_list", multiple=True, help="Количество испытаний: код=число. Пример: --qty испытание_напряжением=3")
@click.option("--discount", default=0.0, type=float, show_default=True, help="Скидка, %")
@click.option("--markup", default=0.0, type=float, show_default=True, help="Наценка, %")
@click.option("--armor/--no-armor", default=None, help="Бронированный кабель (влияет на сложность образца)")
@click.option("--no-minimum", is_flag=True, help="Не применять минимальный заказ (базовая стоимость)")
@click.option("--output", default="out", show_default=True, type=click.Path(), help="Папка для результатов")
@click.option("--save-to-db", is_flag=True, default=True, help="Сохранить расчёт в БД")
@click.option("--db", default="data/app.db", show_default=True)
def calculate_cmd(
    mark: str,
    cores: int,
    groups: int,
    tests: str,
    hours_list: tuple[str, ...],
    hours: str,
    qty_list: tuple[str, ...],
    discount: float,
    markup: float,
    armor: bool | None,
    no_minimum: bool,
    output: str,
    save_to_db: bool,
    db: str,
) -> None:
    """Ручной расчёт стоимости для одной марки."""

    click.echo(f"Расчёт марки: {mark}")

    hours_dict: dict[str, float] = build_default_hours_map(db)

    # Новый удобный способ (--hour temp_low=48 --hour humidity=120)
    if hours_list:
        for item in hours_list:
            if "=" in item:
                key, value = item.split("=", 1)
                try:
                    hours_dict[key.strip()] = float(value.strip())
                except ValueError:
                    click.echo(click.style(f"⚠ Не удалось распарсить --hour {item}", fg="yellow"))
    else:
        try:
            hours_dict.update(json.loads(hours))
        except json.JSONDecodeError:
            if hours != "{}":
                click.echo(click.style("⚠ Не удалось распарсить --hours как JSON. Используйте --hour key=value", fg="yellow"))

    test_list = [t.strip() for t in tests.split(",") if t.strip()]
    quantities: dict[str, int] = {}
    for item in qty_list:
        if "=" in item:
            key, value = item.split("=", 1)
            try:
                quantities[key.strip()] = max(1, int(value.strip()))
            except ValueError:
                click.echo(click.style(f"⚠ Не удалось распарсить --qty {item}", fg="yellow"))

    try:
        calc = calculate_cost(
            mark,
            test_list,
            hours_dict,
            db,
            quantities=quantities or None,
            discount_percent=discount,
            markup_percent=markup,
            has_armor=armor,
            apply_minimum=not no_minimum,
        )
        print_breakdown(calc)

        if save_to_db:
            calc_id = save_calculation(calc, db)
            click.echo(click.style(f"✓ Расчёт сохранён в БД (id={calc_id})", fg="green"))

        out_dir = Path(output)
        out_dir.mkdir(parents=True, exist_ok=True)

    except Exception as e:
        click.echo(click.style(f"Ошибка расчёта: {e}", fg="red"), err=True)
        raise


@cli.command("extract-pdf")
@click.option("--pdf", required=True, type=click.Path(exists=True), help="Путь к PDF или Word (.docx)")
@click.option(
    "--output",
    "output_path",
    default=None,
    type=click.Path(),
    help="Путь для JSON-результата (по умолчанию data/extracted/<имя>.json)",
)
@click.option("--show-marks", is_flag=True, help="Показать найденные марки кабелей")
@click.option("--full-text", is_flag=True, help="Вывести полный извлечённый текст")
@click.option("--no-ocr", is_flag=True, help="Не запускать OCR для сканов")
@click.option("--ocr-dpi", default=200, show_default=True, type=int, help="DPI для OCR сканов")
@click.option(
    "--ocr-engine",
    type=click.Choice(["auto", "tesseract", "easyocr"], case_sensitive=False),
    default="auto",
    show_default=True,
    help="OCR: auto | tesseract | easyocr (PyTorch CV, A/B spike 35v)",
)
@click.option("--no-ocr-cache", is_flag=True, help="Не читать/писать кэш OCR (data/ocr_cache/)")
@click.option("--no-save-marks", is_flag=True, help="Не сохранять марки в БД")
@click.option("--no-save-orgs", is_flag=True, help="Не сохранять организации в БД")
@click.option(
    "--dry-run",
    is_flag=True,
    help="Сохранить только JSON, без записи в БД (марки, организации, document_extractions)",
)
@click.option(
    "--validate",
    is_flag=True,
    help="Вывести отчёт валидатора парсинга; код выхода 1 при блокировке подтверждения",
)
@click.option("--db", default="data/app.db", show_default=True)
def extract_pdf_cmd(
    pdf: str,
    output_path: Optional[str],
    show_marks: bool,
    full_text: bool,
    no_ocr: bool,
    ocr_dpi: int,
    ocr_engine: str,
    no_ocr_cache: bool,
    no_save_marks: bool,
    no_save_orgs: bool,
    dry_run: bool,
    validate: bool,
    db: str,
) -> None:
    """Извлекает текст, таблицы, марки и организации из PDF или Word."""
    pdf_file = Path(pdf)

    try:
        result = extract_from_document(
            pdf_file,
            use_ocr=not no_ocr,
            ocr_dpi=ocr_dpi,
            use_ocr_cache=not no_ocr_cache,
            ocr_engine=ocr_engine,
        )
    except Exception as e:
        click.echo(click.style(f"Ошибка извлечения: {e}", fg="red"), err=True)
        raise SystemExit(1) from e

    click.echo(f"Файл: {pdf_file.name}")
    click.echo(f"Страниц: {result.page_count}")
    click.echo(f"Символов текста: {len(result.text)}")
    click.echo(f"Таблиц: {len(result.tables)}")
    click.echo(f"Найдено марок: {len(result.cable_marks)}")
    if result.customer_name:
        click.echo(f"Заказчик: {result.customer_name}")
    if result.manufacturer_name and result.manufacturer_name != result.customer_name:
        click.echo(f"Производитель: {result.manufacturer_name}")
    if result.organizations:
        click.echo(click.style("\nОрганизации:", bold=True))
        for org in result.organizations:
            click.echo(f"  [{org.role}] {org.name}")
            if org.inn:
                click.echo(f"     ИНН/КПП: {org.inn}/{org.kpp or '—'}")
            if org.address:
                click.echo(f"     Адрес: {org.postal_code or ''} {org.address}".strip())

    if result.is_scanned:
        if result.ocr_used:
            eng = result.ocr_engine or ocr_engine
            label = "easyocr (PyTorch CV)" if eng == "easyocr" else eng
            click.echo(click.style(f"Скан распознан через OCR ({label}).", fg="cyan"))
        elif no_ocr:
            click.echo(
                click.style(
                    "⚠ PDF — скан без текстового слоя. Запусти без --no-ocr для распознавания.",
                    fg="yellow",
                )
            )
        else:
            click.echo(
                click.style(
                    "⚠ PDF — скан, но OCR не дал текста. Проверь установку Tesseract/easyocr.",
                    fg="yellow",
                )
            )

    if show_marks or result.cable_marks:
        if result.cable_marks:
            click.echo(click.style("\nМарки кабелей:", bold=True))
            for i, match in enumerate(result.cable_marks, 1):
                click.echo(f"  {i}. {match.mark}")
                if match.context and show_marks:
                    click.echo(f"     …{match.context}…")
        elif show_marks:
            click.echo("Марки не найдены.")

    if full_text:
        click.echo(click.style("\n--- Текст ---\n", bold=True))
        click.echo(result.text if result.text else "(пусто)")

    validation_report = validate_extraction(result) if validate else None
    skip_db = dry_run or no_save_marks

    out = Path(output_path) if output_path else Path("data/extracted") / f"{pdf_file.stem}.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        result.model_dump_json(indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    click.echo(click.style(f"\n✓ Результат сохранён: {out}", fg="green"))

    if validation_report is not None:
        click.echo(click.style("\n--- Валидация ---\n", bold=True))
        click.echo(format_validation_report(validation_report, source_name=pdf_file.name))

    if dry_run:
        click.echo(click.style("Режим --dry-run: запись в БД пропущена.", fg="cyan"))

    if not skip_db and result.cable_marks:
        migrate_db(db)
        stats = save_cable_marks_from_matches(
            result.cable_marks,
            source=str(pdf_file.resolve()),
            db_path=db,
        )
        click.echo(
            click.style(
                f"✓ Марки в БД: сохранено {stats['saved']}, ошибок {stats['errors']}",
                fg="green",
            )
        )

    if not dry_run and not no_save_orgs and result.organizations:
        migrate_db(db)
        org_ids = save_organizations_from_extraction(
            result.organizations,
            source=str(pdf_file.resolve()),
            db_path=db,
        )
        save_document_extraction(
            source_path=str(pdf_file.resolve()),
            source_type=result.source_type,
            text=result.text,
            marks_count=len(result.cable_marks),
            customer_org_id=org_ids.get("customer_org_id"),
            manufacturer_org_id=org_ids.get("manufacturer_org_id"),
            db_path=db,
        )
        click.echo(
            click.style(
                f"✓ Организации в БД: заказчик id={org_ids.get('customer_org_id')}, "
                f"производитель id={org_ids.get('manufacturer_org_id')}",
                fg="green",
            )
        )

    if validation_report is not None and validation_report.block_confirm:
        raise SystemExit(1)


@cli.command("list-organizations")
@click.option("--search", default=None, help="Поиск по названию, ИНН, адресу")
@click.option("--type", "org_type", default=None, help="Тип: manufacturer, testing_center, …")
@click.option("--limit", default=50, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_organizations_cmd(search: str | None, org_type: str | None, limit: int, db: str) -> None:
    """Список организаций из справочника БД."""
    migrate_db(db)
    rows = list_organizations(search=search, org_type=org_type, limit=limit, db_path=db)
    if not rows:
        click.echo("Организации не найдены.")
        return
    for row in rows:
        acc = "аккред." if row.get("is_accredited") else "не аккред."
        click.echo(
            f"{row['id']:>4}  {row['name']}  [{row.get('org_type', 'unknown')}, {acc}]"
        )
        if row.get("inn"):
            click.echo(f"       ИНН {row['inn']}" + (f"/{row['kpp']}" if row.get("kpp") else ""))
        if row.get("address"):
            click.echo(f"       {row.get('postal_code', '')} {row['address']}".strip())
        if row.get("fsa_registry_number"):
            click.echo(f"       ФСА: {row['fsa_registry_number']}")


@cli.command("process")
@click.option("--input", required=True, type=click.Path(exists=True), help="Путь к PDF/документу")
@click.option("--output", default="data/extracted", show_default=True)
@click.option("--show-marks", is_flag=True, default=True, help="Показать найденные марки")
def process_cmd(input: str, output: str, show_marks: bool) -> None:
    """Обработка заявки из PDF: извлечение марок и сохранение JSON."""
    pdf_file = Path(input)
    click.echo(f"Обработка документа: {pdf_file}")

    try:
        result = extract_from_document(pdf_file)
    except Exception as e:
        click.echo(click.style(f"Ошибка: {e}", fg="red"), err=True)
        raise SystemExit(1) from e

    if result.customer_name:
        click.echo(f"Заказчик: {result.customer_name}")

    if result.is_scanned and result.ocr_used:
        click.echo(click.style("Скан распознан через OCR.", fg="cyan"))

    if show_marks and result.cable_marks:
        click.echo(click.style("\nНайденные марки:", bold=True))
        for i, match in enumerate(result.cable_marks, 1):
            click.echo(f"  {i}. {match.mark}")

    out_dir = Path(output)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{pdf_file.stem}.json"
    out_file.write_text(
        result.model_dump_json(indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    click.echo(click.style(f"✓ Сохранено: {out_file}", fg="green"))


@cli.command("history")
@click.option("--limit", default=10, show_default=True, type=int)
@click.option("--db", default="data/app.db")
def history_cmd(limit: int, db: str) -> None:
    """Показывает последние расчёты из БД."""
    records = get_recent_calculations(limit, db)
    if not records:
        click.echo("История пуста.")
        return

    click.echo(click.style(f"\nПоследние {len(records)} расчётов:\n", bold=True))
    for r in records:
        click.echo(
            f"#{r['id']:>3} | {r['created_at'][:16]} | {r['mark'][:50]:<50} | "
            f"{r['total_cost_with_vat']:>10.2f} ₽ | {r['source']}"
        )
        
        
'''---Управление справочником испытаний (Итерация 2)---'''

@cli.command("list-tests")
@click.option("--category", help="Фильтр по категории")
@click.option("--search", help="Поиск по названию или коду")
@click.option("--limit", default=100, show_default=True)
def list_tests(category: Optional[str], search: Optional[str], limit: int):
    """Выводит список испытаний из справочника."""
    items = list_test_items(category=category, search=search, limit=limit)

    if not items:
        click.echo("Испытания не найдены.")
        return

    click.echo(f"{'Код':<28} {'Наименование':<55} {'Стоимость':>10}  Правило")
    click.echo("-" * 110)

    for item in items:
        click.echo(
            f"{item['code']:<28} {item['name'][:53]:<55} "
            f"{item['base_cost']:>10.0f}  {item['rule_type']}"
        )


@cli.command("add-test-item")
@click.option("--code", required=True, help="Уникальный код (slug)")
@click.option("--name", required=True, help="Полное наименование")
@click.option("--base-cost", "base_cost", required=True, type=float, help="Стоимость без НДС")
@click.option("--category", required=True, help="Категория")
@click.option("--method", default=None)
@click.option(
    "--rule-type",
    "rule_type",
    type=click.Choice(["fixed", "per_core", "per_group", "time_based"]),
    default="fixed",
    show_default=True,
)
@click.option("--hours-key", default=None, help="Ключ часов для time_based")
@click.option("--default-hours", default=None, type=float, help="Часы выдержки по умолчанию")
@click.option("--cost-per-hour", default=None, type=float, help="Стоимость за час выдержки")
def add_test_item_cmd(
    code, name, base_cost, category, method, rule_type,
    hours_key, default_hours, cost_per_hour,
):
    """Добавляет одно испытание вручную."""
    rule_params: dict[str, Any] = {}
    if rule_type == "time_based":
        rule_params = {
            "hours_key": hours_key or code,
            "default_hours": default_hours or 2.0,
            "cost_per_hour": cost_per_hour or 0.0,
        }
    item = TestItemCreate(
        code=code,
        name=name,
        base_cost=base_cost,
        category=category,
        method=method,
        rule_type=rule_type,
        rule_params=rule_params,
    )
    new_id = add_test_item(item)
    click.echo(f"✓ Добавлено испытание (id={new_id}) | {code}")


@cli.command("suggest-tests")
@click.option(
    "--requirements",
    "requirements_text",
    default=None,
    help="Текст контролируемых показателей / требований",
)
@click.option("--mark", default=None, help="Условное обозначение (для поиска в последнем JSON)")
@click.option("--db", default="data/app.db", show_default=True)
def suggest_tests_cmd(
    requirements_text: str | None,
    mark: str | None,
    db: str,
) -> None:
    """Предлагает коды испытаний по тексту требований из заявки."""
    migrate_db(db)
    suggestions = []

    if requirements_text:
        suggestions = map_requirements_to_tests(requirements_text, db_path=db)
    elif mark:
        last = get_last_document_extraction(db)
        if not last:
            raise click.ClickException("Нет извлечённых заявок в БД. Сначала extract-pdf.")
        import json as _json
        from pathlib import Path as _Path

        src = last.get("source_path") or ""
        stem = _Path(src).stem
        json_path = _Path("data/extracted") / f"{stem}.json"
        if not json_path.exists():
            raise click.ClickException(f"JSON не найден: {json_path}")
        from .models import PdfExtractionResult

        result = PdfExtractionResult.model_validate(
            _json.loads(json_path.read_text(encoding="utf-8"))
        )
        match = next((m for m in result.cable_marks if mark.lower() in m.mark.lower()), None)
        if not match:
            raise click.ClickException(f"Марка «{mark}» не найдена в {json_path.name}")
        suggestions = suggest_tests_for_mark(match, db_path=db)
    else:
        raise click.ClickException("Укажите --requirements или --mark")

    if not suggestions:
        click.echo("Испытания не определены по тексту требований.")
        return

    click.echo(click.style("Предложенные испытания:\n", bold=True))
    for s in suggestions:
        src = "БД" if s.source == "database" else "правило"
        pat = f"  «{s.matched_pattern}»" if s.matched_pattern else ""
        click.echo(f"  {s.code:<22} {s.confidence:>4.0%}  {s.name[:45]:<45}  [{src}]{pat}")


@cli.command("list-test-mappings")
@click.option("--test-code", default=None, help="Фильтр по коду испытания")
@click.option("--limit", default=50, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_test_mappings_cmd(test_code: str | None, limit: int, db: str) -> None:
    """Справочник маппинга «фраза требования → испытание»."""
    migrate_db(db)
    rows = list_test_mappings(test_code=test_code, limit=limit, db_path=db)
    if not rows:
        click.echo("Маппинги не найдены.")
        return
    for row in rows:
        click.echo(
            f"{row['id']:>4}  {row['test_code']:<22}  "
            f"×{row.get('usage_count', 0):<3}  {row['requirement_pattern']}"
        )


@cli.command("add-test-mapping")
@click.option("--pattern", required=True, help="Фраза из заявки (подстрока, без учёта регистра)")
@click.option("--test-code", required=True, help="Код испытания из test_items")
@click.option("--note", default=None)
@click.option("--db", default="data/app.db", show_default=True)
def add_test_mapping_cmd(pattern: str, test_code: str, note: str | None, db: str) -> None:
    """Добавляет маппинг требования на испытание."""
    migrate_db(db)
    mapping_id = add_test_mapping(pattern, test_code, note=note, db_path=db)
    click.echo(click.style(f"✓ Маппинг сохранён (id={mapping_id})", fg="green"))


@cli.command("update-test-mapping")
@click.option("--id", "mapping_id", required=True, type=int, help="ID записи test_mappings")
@click.option("--pattern", default=None, help="Новая фраза")
@click.option("--test-code", default=None, help="Новый код испытания")
@click.option("--note", default=None, help="Примечание")
@click.option("--db", default="data/app.db", show_default=True)
def update_test_mapping_cmd(
    mapping_id: int,
    pattern: str | None,
    test_code: str | None,
    note: str | None,
    db: str,
) -> None:
    """Обновляет маппинг по id."""
    migrate_db(db)
    try:
        update_test_mapping(
            mapping_id,
            requirement_pattern=pattern,
            test_code=test_code,
            note=note,
            db_path=db,
        )
    except ValueError as exc:
        raise click.ClickException(str(exc)) from exc
    click.echo(click.style(f"✓ Маппинг id={mapping_id} обновлён", fg="green"))


@cli.command("delete-test-mapping")
@click.option("--id", "mapping_id", required=True, type=int, help="ID записи test_mappings")
@click.option("--db", default="data/app.db", show_default=True)
def delete_test_mapping_cmd(mapping_id: int, db: str) -> None:
    """Удаляет маппинг по id."""
    migrate_db(db)
    if not delete_test_mapping(mapping_id, db_path=db):
        raise click.ClickException(f"Маппинг id={mapping_id} не найден")
    click.echo(click.style(f"✓ Маппинг id={mapping_id} удалён", fg="green"))


@cli.command("list-cable-marks")
@click.option("--search", default=None, help="Поиск по марке")
@click.option("--limit", default=50, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_cable_marks_cmd(search: Optional[str], limit: int, db: str) -> None:
    """Список накопленных марок кабелей из БД."""
    rows = list_cable_marks(search=search, limit=limit, db_path=db)
    if not rows:
        click.echo("Марки не найдены.")
        return
    click.echo(
        f"{'Усл. обозначение':<45} {'Марка':<12} {'ТПЖ':>4} {'Размер':>10}  Документ"
    )
    click.echo("-" * 110)
    for row in rows:
        unit = "мм²" if row.get("size_unit") == "mm2" else "мм"
        click.echo(
            f"{row['full_mark'][:44]:<45} {row['brand'][:11]:<12} "
            f"{row['cores_count']:>4} "
            f"{row['characteristic_size']:>8}{unit:<2}  "
            f"{(row.get('document') or '')[:30]}"
        )


@cli.command("set-climatic-hours")
@click.option("--temp-low", default=None, type=float, help="Пониженная температура, ч")
@click.option("--temp-high", default=None, type=float, help="Повышенная температура, ч")
@click.option("--temp-cycling", default=None, type=float, help="Изменение температур, ч")
@click.option("--humidity", default=None, type=float, help="Повышенная влажность, ч")
@click.option("--solar-radiation", default=None, type=float, help="Солнечная радиация, ч")
@click.option("--db", default="data/app.db", show_default=True)
def set_climatic_hours_cmd(
    temp_low: Optional[float],
    temp_high: Optional[float],
    temp_cycling: Optional[float],
    humidity: Optional[float],
    solar_radiation: Optional[float],
    db: str,
) -> None:
    """Настройка времени выдержки климатических испытаний."""
    migrate_db(db)
    current = get_climatic_settings(db) or ClimaticTestSettings()
    settings = ClimaticTestSettings(
        temp_low=temp_low if temp_low is not None else current.temp_low,
        temp_high=temp_high if temp_high is not None else current.temp_high,
        temp_cycling=temp_cycling if temp_cycling is not None else current.temp_cycling,
        humidity=humidity if humidity is not None else current.humidity,
        solar_radiation=solar_radiation if solar_radiation is not None else current.solar_radiation,
    )
    save_climatic_settings(settings, db)
    click.echo(
        f"✓ Выдержка: temp_low={settings.temp_low}, temp_high={settings.temp_high}, "
        f"temp_cycling={settings.temp_cycling}, humidity={settings.humidity}, "
        f"solar_radiation={settings.solar_radiation} ч"
    )


@cli.command("import-tests")
@click.option("--file", "file_path", required=True, type=click.Path(exists=True), help="Путь к Excel-файлу")
@click.option("--dry-run", is_flag=True, help="Только проверить файл, ничего не записывать")
@click.option("--sheet", default=None, help="Название листа (по умолчанию первый)")
def import_tests(file_path: str, dry_run: bool, sheet: Optional[str]):
    """Пакетная загрузка / обновление испытаний из Excel."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    items_to_import: list[TestItemCreate] = []
    errors: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[1]:
            continue

        try:
            name = str(row[1]).strip()
            if not name:
                continue

            # === Генерация code, если он не указан ===
            raw_code = str(row[0]).strip() if row[0] else ""
            code = raw_code if raw_code else _slugify(name)

            item = TestItemCreate(
                code=code,
                name=name,
                base_cost=float(row[2]) if row[2] is not None else 0.0,
                category=str(row[3]).strip() if row[3] else "Без категории",
                method=str(row[4]).strip() if row[4] else None,
                rule_type=str(row[5]).strip() if row[5] else "fixed",
                rule_params=json.loads(str(row[6])) if row[6] else {},
            )
            items_to_import.append(item)

        except Exception as e:
            errors.append(f"Строка {row_idx}: {e}")

    # === Вывод результатов ===
    click.echo(f"Найдено записей для импорта: {len(items_to_import)}")
    if errors:
        click.echo(f"Ошибок при чтении: {len(errors)}")
        for err in errors[:10]:  # показываем первые 10 ошибок
            click.echo(f"  - {err}")
        if len(errors) > 10:
            click.echo(f"  ... и ещё {len(errors) - 10} ошибок")

    if dry_run:
        click.echo("\nРежим --dry-run: изменения в базу не записаны.")
        return

    if not items_to_import:
        click.echo("Нет данных для загрузки.")
        return

    # Загружаем в БД
    stats = bulk_upsert_test_items(items_to_import)
    click.echo(f"\n✓ Успешно обработано: {stats['processed']}")
    if stats.get("errors", 0) > 0:
        click.echo(f"⚠ Ошибок при записи: {stats['errors']}")


@cli.command("generate-kp")
@click.option("--customer", required=True, help="Заказчик / изготовитель")
@click.option(
    "--subject",
    default="Проведение периодических испытаний",
    show_default=True,
    help="Вид испытаний для вводной КП (по умолчанию — периодические)",
)
@click.option("--calc-ids", required=True, help="ID расчётов через запятую (например: 1,2,3,4)")
@click.option("--note", default=None, help="Дополнительный текст")
@click.option(
    "--output",
    "output_path",
    default=None,
    type=click.Path(),
    help="Путь к .docx (по умолчанию data/generated/КП_...)",
)
@click.option("--db", default="data/app.db", show_default=True)
def generate_kp_cmd(
    customer: str,
    subject: str,
    calc_ids: str,
    note: Optional[str],
    output_path: Optional[str],
    db: str,
) -> None:
    """Формирует коммерческое предложение (Word) по выбранным расчётам."""
    ids = [int(x.strip()) for x in calc_ids.split(",") if x.strip()]
    if not ids:
        raise click.ClickException("Укажите хотя бы один ID расчёта в --calc-ids")

    if output_path:
        out = Path(output_path)
    else:
        safe = re.sub(r'[<>:"/\\|?*]', "", customer)[:40] or "заказчик"
        from .config import GENERATED_DIR_DEFAULT

        out = GENERATED_DIR_DEFAULT / f"КП_{safe}.docx"

    try:
        path = generate_kp_from_db(
            customer=customer,
            subject=subject,
            calculation_ids=ids,
            output_path=out,
            db_path=db,
            note=note,
        )
    except Exception as exc:
        raise click.ClickException(str(exc)) from exc

    migrate_db(db)
    last_doc = get_last_document_extraction(db)
    order_id = create_order_from_kp(
        customer_name=customer,
        manufacturer_name=last_doc.get("manufacturer_name") if last_doc else None,
        subject=subject,
        note=note,
        calculation_ids=ids,
        kp_output_path=str(path),
        document_extraction_id=int(last_doc["id"]) if last_doc else None,
        db_path=db,
    )
    click.echo(click.style(f"✓ КП сохранено: {path}", fg="green"))
    click.echo(click.style(f"✓ Заказ №{order_id} создан", fg="green"))


@cli.command("generate-application")
@click.option("--order-id", required=True, type=int, help="ID заказа в БД")
@click.option(
    "--output",
    "output_path",
    default=None,
    type=click.Path(),
    help="Путь к .docx (по умолчанию data/generated/Заявка_...)",
)
@click.option("--db", default="data/app.db", show_default=True)
def generate_application_cmd(order_id: int, output_path: Optional[str], db: str) -> None:
    """Формирует заявку на испытания (Word) по сохранённому заказу."""
    migrate_db(db)
    try:
        path = generate_application_from_order(
            order_id,
            output_path=Path(output_path) if output_path else None,
            db_path=db,
        )
    except Exception as exc:
        raise click.ClickException(str(exc)) from exc
    click.echo(click.style(f"✓ Заявка сохранена: {path}", fg="green"))


@cli.command("list-generated-documents")
@click.option("--order-id", default=None, type=int, help="Фильтр по заказу")
@click.option("--type", "doc_type", type=click.Choice(["kp", "application"]), default=None)
@click.option("--limit", default=20, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_generated_documents_cmd(
    order_id: int | None,
    doc_type: str | None,
    limit: int,
    db: str,
) -> None:
    """История сгенерированных файлов (КП, заявки на испытания)."""
    migrate_db(db)
    rows = list_generated_documents(order_id=order_id, doc_type=doc_type, limit=limit, db_path=db)
    if not rows:
        click.echo("Сгенерированные документы не найдены.")
        return
    for row in rows:
        click.echo(
            f"{row['id']:>4}  заказ №{row.get('order_id') or '—':>4}  "
            f"{row['doc_type']:<12}  {(row.get('created_at') or '')[:16]}  "
            f"{row['file_path']}"
        )


@cli.command("list-applications")
@click.option("--order-id", default=None, type=int, help="Фильтр по заказу")
@click.option("--limit", default=20, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_applications_cmd(order_id: int | None, limit: int, db: str) -> None:
    """История сформированных заявок на испытания из БД."""
    migrate_db(db)
    rows = list_test_applications(order_id=order_id, limit=limit, db_path=db)
    if not rows:
        click.echo("Заявки на испытания не найдены.")
        return
    for row in rows:
        click.echo(
            f"{row['id']:>4}  заказ №{row['order_id']}  "
            f"{(row.get('created_at') or '')[:16]}  "
            f"{(row.get('test_type') or '—'):18}  "
            f"марок: {row.get('marks_count') or 0}"
        )
        click.echo(f"       {(row.get('customer_name') or '—')[:50]}")
        click.echo(f"       {row.get('output_path') or '—'}")


@cli.command("list-orders")
@click.option("--limit", default=20, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_orders_cmd(limit: int, db: str) -> None:
    """Список сохранённых заказов (КП)."""
    migrate_db(db)
    rows = list_orders(limit=limit, db_path=db)
    if not rows:
        click.echo("Заказы не найдены.")
        return
    for row in rows:
        click.echo(
            f"{row['id']:>4}  {(row.get('created_at') or '')[:16]}  "
            f"{(row.get('customer_name') or '—')[:35]:35}  "
            f"марок: {row.get('marks_count') or 0}  "
            f"{float(row.get('total_with_vat') or 0):,.2f} ₽".replace(",", " ")
        )


@cli.command("ingest-training-doc")
@click.option("--file", "file_path", required=True, type=click.Path(exists=True))
@click.option("--type", "document_type", default=None, help="letter_periodic, direction_il, …")
@click.option("--family", "document_family", default=None, help="periodic_letter_v1, lan_letter_v1, …")
@click.option("--keep-in-place", is_flag=True, help="Не переносить из inbox в registered/")
@click.option("--skip-extract", is_flag=True, help="Только регистрация в БД, без extract_from_document")
@click.option("--db", default="data/app.db", show_default=True)
def ingest_training_doc_cmd(
    file_path: str,
    document_type: Optional[str],
    document_family: Optional[str],
    keep_in_place: bool,
    skip_extract: bool,
    db: str,
) -> None:
    """Регистрирует документ в training_documents (Фаза 1)."""
    from .persistence.training_repo import ingest_training_document

    migrate_db(db)
    doc = ingest_training_document(
        Path(file_path),
        document_type=document_type,
        document_family=document_family,
        move_to_registered=not keep_in_place,
        run_extract=not skip_extract,
        db_path=db,
    )
    click.echo(
        f"✓ training_documents id={doc['id']}  {doc['file_name']}  "
        f"type={doc.get('document_type') or '—'}  family={doc.get('document_family') or '—'}"
    )


@cli.command("ingest-training-inbox")
@click.option("--keep-in-place", is_flag=True)
@click.option("--db", default="data/app.db", show_default=True)
def ingest_training_inbox_cmd(keep_in_place: bool, db: str) -> None:
    """Пакетный ingest всех PDF/DOCX из data/training/documents/inbox/."""
    from .persistence.training_repo import ingest_inbox_batch, seed_document_families

    migrate_db(db)
    seed_document_families(db_path=db)
    stats = ingest_inbox_batch(move_to_registered=not keep_in_place, db_path=db)
    click.echo(f"OK: {stats['ok']}  FAIL: {stats['fail']}")


@cli.command("import-label")
@click.option(
    "--document-id",
    default=None,
    type=int,
    help="ID в training_documents; если не указан — ищется по source_file в JSON",
)
@click.option("--file", "label_file", required=True, type=click.Path(exists=True))
@click.option(
    "--type",
    "label_type",
    default=None,
    type=click.Choice(["marks", "organizations", "requirements", "ocr_page", "full_json"]),
)
@click.option("--db", default="data/app.db", show_default=True)
def import_label_cmd(
    document_id: Optional[int],
    label_file: str,
    label_type: Optional[str],
    db: str,
) -> None:
    """Импортирует JSON-разметку в training_labels."""
    from .persistence.training_repo import import_label_file, resolve_document_id_for_label

    migrate_db(db)
    payload = json.loads(Path(label_file).read_text(encoding="utf-8"))
    resolved_id = document_id or resolve_document_id_for_label(payload, label_file, db_path=db)
    label_id = import_label_file(
        document_id,
        Path(label_file),
        label_type=label_type,  # type: ignore[arg-type]
        db_path=db,
    )
    click.echo(f"✓ training_labels id={label_id} для document_id={resolved_id}")


@cli.command("eval-extraction")
@click.option(
    "--labels-dir",
    default="data/training/labels/marks",
    show_default=True,
    type=click.Path(),
)
@click.option(
    "--output",
    "output_path",
    default=None,
    type=click.Path(),
    help="JSON-отчёт (по умолчанию data/training/exports/reports/eval_marks_<дата>.json)",
)
@click.option("--no-ocr-cache", is_flag=True, help="Свежий OCR при сравнении")
@click.option(
    "--ocr-engine",
    type=click.Choice(["auto", "tesseract", "easyocr"], case_sensitive=False),
    default="auto",
    show_default=True,
    help="OCR engine для A/B: auto | tesseract | easyocr (PyTorch CV)",
)
@click.option("--db", default="data/app.db", show_default=True)
def eval_extraction_cmd(
    labels_dir: str,
    output_path: Optional[str],
    no_ocr_cache: bool,
    ocr_engine: str,
    db: str,
) -> None:
    """Сравнивает извлечённые марки с эталонами в labels/marks/."""
    from datetime import date

    from .config import TRAINING_EXPORTS_REPORTS_DIR
    from .validation.eval_extraction import eval_marks_labels_dir

    migrate_db(db)
    report = eval_marks_labels_dir(
        Path(labels_dir),
        use_ocr_cache=not no_ocr_cache,
        ocr_engine=ocr_engine,
        db_path=db,
    )
    out = (
        Path(output_path)
        if output_path
        else TRAINING_EXPORTS_REPORTS_DIR
        / f"eval_marks_{date.today().isoformat()}_{ocr_engine}.json"
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    click.echo(
        f"Файлов: {report['files_evaluated']}/{report['files_total']}  "
        f"recall (micro): {report['micro_recall']:.0%}  "
        f"recall (macro): {report['macro_recall']:.0%}"
    )
    click.echo(f"Отчёт: {out}")


@cli.command("ocr-benchmark")
@click.option("--pdf", "pdf_path", required=True, type=click.Path(exists=True))
@click.option("--page", default=1, show_default=True, type=int, help="Номер страницы (1-based)")
@click.option("--dpi", default=200, show_default=True, type=int)
@click.option(
    "--output",
    "output_path",
    default=None,
    type=click.Path(),
    help="JSON-отчёт (по умолчанию data/training/exports/reports/ocr_benchmark_<stem>_p<N>_<дата>.json)",
)
@click.option(
    "--batch",
    "batch_paths",
    multiple=True,
    type=click.Path(exists=True),
    help="Дополнительные PDF для пакетного отчёта (вместе с --pdf)",
)
def ocr_benchmark_cmd(
    pdf_path: str,
    page: int,
    dpi: int,
    output_path: Optional[str],
    batch_paths: tuple[str, ...],
) -> None:
    """Сравнивает OCR raw vs preprocess v1 на странице скана."""
    from datetime import date

    from .config import TRAINING_EXPORTS_REPORTS_DIR
    from .extraction.ocr.benchmark import (
        benchmark_pdf_page,
        benchmark_scans_batch,
        save_benchmark_report,
    )

    paths = [Path(pdf_path), *[Path(p) for p in batch_paths]]
    if len(paths) == 1:
        report = benchmark_pdf_page(paths[0], page=page, dpi=dpi)
        out = save_benchmark_report(
            report,
            Path(output_path) if output_path else None,
        )
        raw = report["variants"]["raw"]
        pre = report["variants"]["preprocessed"]
        click.echo(
            f"Страница {page}/{report['page_count']}  "
            f"raw: conf={raw['mean_confidence']:.0%} chars={raw['text_chars']}  "
            f"pre: conf={pre['mean_confidence']:.0%} chars={pre['text_chars']}"
        )
        if report.get("cer_delta") is not None:
            click.echo(f"CER delta (pre - raw): {report['cer_delta']:+.2%}")
        click.echo(f"Отчёт: {out}")
        return

    batch_report = benchmark_scans_batch(paths, page=page, dpi=dpi)
    out = (
        Path(output_path)
        if output_path
        else TRAINING_EXPORTS_REPORTS_DIR / f"ocr_benchmark_batch_{date.today().isoformat()}.json"
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(batch_report, ensure_ascii=False, indent=2), encoding="utf-8")
    click.echo(
        f"Файлов: {batch_report['files_ok']}/{batch_report['files_total']}  "
        f"улучшено preprocess: {batch_report['preprocess_improved_count']}"
    )
    click.echo(f"Отчёт: {out}")


@cli.command("sync-corrections")
@click.option("--dir", "corrections_dir", default=None, type=click.Path())
@click.option("--db", default="data/app.db", show_default=True)
def sync_corrections_cmd(corrections_dir: Optional[str], db: str) -> None:
    """Синхронизирует data/training/corrections/*.jsonl → training_corrections."""
    from .persistence.training_repo import sync_corrections_from_dir

    migrate_db(db)
    stats = sync_corrections_from_dir(
        Path(corrections_dir) if corrections_dir else None,
        db_path=db,
    )
    click.echo(f"Файлов: {stats['files']}  строк: {stats['rows']}  пропущено: {stats['skipped']}")


@cli.command("index-rag")
@click.option(
    "--folder",
    required=True,
    type=click.Path(exists=True),
    help="Папка корпуса (tu, protocols, pmi, gost, templates)",
)
@click.option(
    "--kind",
    "doc_kind",
    default=None,
    type=click.Choice(["tu", "protocol", "gost", "method", "pmi", "internal", "template"]),
)
@click.option("--db", default="data/app.db", show_default=True)
def index_rag_cmd(folder: str, doc_kind: Optional[str], db: str) -> None:
    """Регистрирует файлы корпуса в rag_documents (без embeddings — Фаза 4)."""
    from .persistence.training_repo import index_rag_folder

    migrate_db(db)
    stats = index_rag_folder(
        Path(folder),
        doc_kind=doc_kind,  # type: ignore[arg-type]
        db_path=db,
    )
    click.echo(f"Проиндексировано: {stats['indexed']}  пропущено: {stats['skipped']}")


@cli.command("seed-training")
@click.option("--db", default="data/app.db", show_default=True)
def seed_training_cmd(db: str) -> None:
    """Семейства YAML → document_families; эталоны data/extracted → training."""
    from .persistence.training_repo import import_extracted_fixtures, seed_document_families

    migrate_db(db)
    families = seed_document_families(db_path=db)
    fixture_ids = import_extracted_fixtures(db_path=db)
    click.echo(f"Семейств: {families}  эталонов JSON: {len(fixture_ids)}")


@cli.command("list-training-docs")
@click.option("--status", "label_status", default=None)
@click.option("--limit", default=30, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_training_docs_cmd(label_status: Optional[str], limit: int, db: str) -> None:
    """Список training_documents."""
    from .persistence.training_repo import list_training_documents

    migrate_db(db)
    rows = list_training_documents(label_status=label_status, limit=limit, db_path=db)
    if not rows:
        click.echo("Нет документов.")
        return
    for row in rows:
        click.echo(
            f"{row['id']:>4}  {row['label_status']:10}  "
            f"{(row.get('document_type') or '—'):12}  {row['file_name']}"
        )


@cli.command("list-rag")
@click.option(
    "--kind",
    "doc_kind",
    default=None,
    type=click.Choice(["tu", "protocol", "gost", "pmi", "internal", "template"]),
)
@click.option("--limit", default=40, show_default=True)
@click.option("--db", default="data/app.db", show_default=True)
def list_rag_cmd(doc_kind: Optional[str], limit: int, db: str) -> None:
    """Список rag_documents (корпус ТУ/ГОСТ/ПМИ/протоколы)."""
    from .persistence.training_repo import list_rag_documents

    migrate_db(db)
    rows = list_rag_documents(doc_kind=doc_kind, limit=limit, db_path=db)
    if not rows:
        click.echo("Корпус пуст. Запустите: index-rag --folder data/training/rag_corpus")
        return
    for row in rows:
        click.echo(
            f"{row['id']:>4}  {row['doc_kind']:10}  "
            f"{(row.get('title') or '—')[:45]:45}  {row.get('file_path') or ''}"
        )


@cli.command("list-parse-snapshots")
@click.option("--limit", default=50, show_default=True, type=int)
def list_parse_snapshots_cmd(limit: int) -> None:
    """Список сохранённых снимков парсинга (data/parse_snapshots/)."""
    from .parse_compare import list_snapshots

    rows = list_snapshots(limit=limit)
    if not rows:
        click.echo("Снимков нет. Сохраните из GUI (вкладка «Сравнение») или save-parse-snapshot.")
        return
    for row in rows:
        click.echo(
            f"{row['id']}  marks={row['marks_count']:>2}  q={row['quality_score']:.2f}  "
            f"{(row.get('ocr_engine') or '—'):10}  {(row.get('label') or '')[:50]}"
        )


@cli.command("compare-parse-snapshots")
@click.argument("snapshot_a")
@click.argument("snapshot_b")
@click.option("--output", "output_path", default=None, type=click.Path(), help="JSON отчёт сравнения")
def compare_parse_snapshots_cmd(snapshot_a: str, snapshot_b: str, output_path: Optional[str]) -> None:
    """Сравнивает два снимка парсинга по id или пути к JSON."""
    from .parse_compare import compare_snapshots, load_snapshot

    a = load_snapshot(snapshot_a)
    b = load_snapshot(snapshot_b)
    report = compare_snapshots(a, b)
    if output_path:
        Path(output_path).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        click.echo(f"Отчёт: {output_path}")
    marks = report["marks"]
    click.echo(f"A: {a.label} ({a.ocr_engine})  marks={marks['count_a']}")
    click.echo(f"B: {b.label} ({b.ocr_engine})  marks={marks['count_b']}")
    click.echo(f"Пересечение: {marks['intersection']}  Jaccard: {marks['jaccard']:.2%}")
    click.echo(f"Только A: {len(marks['only_a'])}  Только B: {len(marks['only_b'])}")
    click.echo(f"Quality A/B: {report['quality']['a']} / {report['quality']['b']}  winner={report['quality']['winner']}")
    if marks["only_a"]:
        click.echo("only A (norm): " + ", ".join(marks["only_a"][:12]))
    if marks["only_b"]:
        click.echo("only B (norm): " + ", ".join(marks["only_b"][:12]))


@cli.command("save-parse-snapshot")
@click.option("--json", "json_path", required=True, type=click.Path(exists=True), help="JSON PdfExtractionResult")
@click.option("--label", default="", help="Подпись снимка")
@click.option("--notes", default="", help="Заметка")
@click.option("--dpi", default=None, type=int)
def save_parse_snapshot_cmd(json_path: str, label: str, notes: str, dpi: Optional[int]) -> None:
    """Сохраняет снимок парсинга из JSON извлечения."""
    from .models import PdfExtractionResult
    from .parse_compare import save_snapshot_from_extraction

    data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    result = PdfExtractionResult.model_validate(data)
    snap = save_snapshot_from_extraction(result, label=label, notes=notes, ocr_dpi=dpi)
    click.echo(click.style(f"✓ Снимок {snap.id}", fg="green"))
    click.echo(f"  marks={snap.metrics.marks_count}  quality={snap.metrics.quality_score}  engine={snap.ocr_engine}")


@cli.command("import-test-program")
@click.option("--file", "file_path", required=True, type=click.Path(exists=True), help="DOCX программы")
@click.option("--db", default="data/app.db", show_default=True)
@click.option("--no-match-price", is_flag=True, help="Не сопоставлять с прайсом")
def import_test_program_cmd(file_path: str, db: str, no_match_price: bool) -> None:
    """Импорт программы испытаний из Word (.docx) в БД."""
    from .generation.program_importer import import_program_from_docx
    from .persistence.sqlite_repo import migrate_db

    migrate_db(db)
    try:
        result = import_program_from_docx(
            file_path, db_path=db, match_price=not no_match_price
        )
    except Exception as e:
        click.echo(click.style(f"Ошибка: {e}", fg="red"), err=True)
        raise SystemExit(1) from e
    click.echo(click.style(f"✓ Программа id={result['program_id']}", fg="green"))
    click.echo(f"  name:  {result['name'][:80]}")
    click.echo(f"  type:  {result.get('test_type') or '—'}")
    click.echo(f"  mark:  {result.get('cable_mark_text') or '—'}")
    click.echo(f"  tu:    {result.get('tu_ref') or '—'}")
    click.echo(
        f"  items: {result['items_count']}  "
        f"(price matched={result['matched']}, unmatched={result['unmatched']})"
    )


@cli.command("list-test-programs")
@click.option("--search", default=None)
@click.option("--limit", default=50, show_default=True, type=int)
@click.option("--db", default="data/app.db", show_default=True)
def list_test_programs_cmd(search: Optional[str], limit: int, db: str) -> None:
    """Список программ испытаний в БД."""
    from .persistence.sqlite_repo import list_test_programs, migrate_db

    migrate_db(db)
    rows = list_test_programs(search=search, limit=limit, db_path=db)
    if not rows:
        click.echo("Программ нет. import-test-program --file …")
        return
    for r in rows:
        click.echo(
            f"{r['id']:>4}  items={r.get('items_count', 0):>3}  "
            f"{(r.get('test_type') or '—')[:20]:20}  {(r.get('name') or '')[:60]}"
        )


@cli.command("show-test-program")
@click.option("--id", "program_id", required=True, type=int)
@click.option("--db", default="data/app.db", show_default=True)
def show_test_program_cmd(program_id: int, db: str) -> None:
    """Показать программу и позиции."""
    from .persistence.sqlite_repo import get_test_program, migrate_db

    migrate_db(db)
    prog = get_test_program(program_id, db_path=db)
    if not prog:
        click.echo("Не найдено", err=True)
        raise SystemExit(1)
    click.echo(f"#{prog['id']} {prog['name']}")
    click.echo(f"  type={prog.get('test_type')}  mark={prog.get('cable_mark_text')}")
    click.echo(f"  tu={prog.get('tu_ref')}")
    click.echo(f"  source={prog.get('source_path')}")
    for it in prog.get("items") or []:
        click.echo(
            f"  {it['sort_order']:>3}. {it['name'][:55]:55}  "
            f"req={it.get('requirement_clause') or '—':12}  "
            f"meth={it.get('method_clause') or '—':12}  "
            f"price={it.get('price_test_code') or '—'}"
        )


@cli.command("delete-test-program")
@click.option("--id", "program_id", required=True, type=int)
@click.option("--db", default="data/app.db", show_default=True)
@click.option("--yes", is_flag=True)
def delete_test_program_cmd(program_id: int, db: str, yes: bool) -> None:
    """Удалить программу испытаний."""
    from .persistence.sqlite_repo import delete_test_program, migrate_db

    if not yes:
        click.echo("Повторите с --yes")
        raise SystemExit(1)
    migrate_db(db)
    ok = delete_test_program(program_id, db_path=db)
    click.echo("✓ удалено" if ok else "не найдено")


@cli.command("match-program-price")
@click.option("--id", "program_id", required=True, type=int)
@click.option("--db", default="data/app.db", show_default=True)
@click.option(
    "--overwrite",
    is_flag=True,
    help="Перезаписать уже заданные price_test_code (исправить ошибочные)",
)
@click.option("--verbose", "-v", is_flag=True, help="Показать каждую позицию")
def match_program_price_cmd(
    program_id: int, db: str, overwrite: bool, verbose: bool
) -> None:
    """Сопоставить позиции программы с кодами прайса (S4)."""
    from .persistence.sqlite_repo import match_program_items_to_price, migrate_db

    migrate_db(db)
    stats = match_program_items_to_price(
        program_id, db_path=db, overwrite=overwrite
    )
    click.echo(
        f"{stats.get('summary', '')}  "
        f"matched={stats['matched']} unmatched={stats['unmatched']}"
        + ("  [overwrite]" if overwrite else "")
    )
    if verbose:
        for d in stats.get("details") or []:
            code = d.get("code") or "—"
            click.echo(
                f"  [{d.get('method', '?')}] {(d.get('name') or '')[:55]} → {code}"
            )


@cli.command("list-norm-documents")
@click.option("--kind", default=None, help="tu|gost|iec|pmi|other")
@click.option("--db", default="data/app.db", show_default=True)
def list_norm_documents_cmd(kind: Optional[str], db: str) -> None:
    """Нормативные документы (S5, задел базы требований)."""
    from .persistence.sqlite_repo import list_norm_documents, migrate_db

    migrate_db(db)
    for r in list_norm_documents(kind=kind, db_path=db):
        click.echo(f"{r['id']:>3}  {r['kind']:6}  {r['doc_id']:28}  {r['title'][:50]}")


@cli.command("list-acceptance-items")
@click.option(
    "--doc",
    "doc_id",
    default=None,
    help="doc_id ТУ, напр. «ТУ 27.31.11-131-47273194-2025»",
)
@click.option("--norm-id", "norm_id", default=None, type=int, help="id norm_documents")
@click.option(
    "--billable/--all",
    "billable_only",
    default=False,
    help="Только billable=1 (по умолчанию все)",
)
@click.option("--db", default="data/app.db", show_default=True)
def list_acceptance_items_cmd(
    doc_id: Optional[str],
    norm_id: Optional[int],
    billable_only: bool,
    db: str,
) -> None:
    """Строки каталога приёмки ТУ (acceptance_items, волна 1)."""
    from .persistence.sqlite_repo import list_acceptance_items, migrate_db

    migrate_db(db)
    rows = list_acceptance_items(
        doc_id=doc_id,
        norm_document_id=norm_id,
        billable=True if billable_only else None,
        db_path=db,
    )
    if not rows:
        click.echo("Нет acceptance_items (migrate-db + seed или импорт волны 2).")
        return
    for r in rows:
        bill = "💰" if r.get("billable") else "—"
        cat = (r.get("test_category") or "—")[:8]
        code = r.get("price_test_code") or "n/a"
        click.echo(
            f"{r['id']:>4}  {bill}  {cat:8}  {(r.get('doc_id') or '')[:32]:32}  "
            f"{(r.get('name_exact') or '')[:48]:48}  code={code}"
        )


@cli.command("show-acceptance-item")
@click.option("--id", "item_id", type=int, required=True, help="id acceptance_items")
@click.option("--db", default="data/app.db", show_default=True)
def show_acceptance_item_cmd(item_id: int, db: str) -> None:
    """Карточка acceptance_item: clauses + внешние ГОСТ + regime."""
    from .persistence.sqlite_repo import get_acceptance_item, migrate_db

    migrate_db(db)
    item = get_acceptance_item(item_id, db_path=db)
    if not item:
        click.echo(f"Не найдено id={item_id}", err=True)
        raise SystemExit(1)
    click.echo(f"id={item['id']}  status={item.get('status')}")
    click.echo(f"doc: {item.get('doc_id')}  ({item.get('doc_title')})")
    click.echo(f"name: {item.get('name_exact')}")
    click.echo(
        f"category={item.get('test_category') or '—'}  "
        f"group={item.get('group_code') or '—'}  "
        f"billable={bool(item.get('billable'))}  "
        f"price={item.get('price_test_code') or 'n/a'}"
    )
    click.echo("clauses:")
    for c in item.get("clauses") or []:
        click.echo(
            f"  [{c.get('role')}] п.{c.get('clause')}  "
            f"kind={c.get('clause_kind')}  {(c.get('title') or '')[:60]}"
        )
    if not item.get("clauses"):
        click.echo("  (нет)")
    click.echo("method_external:")
    for e in item.get("method_external") or []:
        click.echo(
            f"  {e.get('ext_doc_id')}  "
            f"{e.get('ext_clause_or_method') or ''}  "
            f"{(e.get('note') or '')[:40]}"
        )
    if not item.get("method_external"):
        click.echo("  (нет)")
    if item.get("regime"):
        click.echo(f"regime: {item['regime']}")
    elif item.get("regime_json"):
        click.echo(f"regime_json: {item['regime_json']}")
    if item.get("notes"):
        click.echo(f"notes: {item['notes']}")


@cli.command("show-norm-catalog")
@click.option(
    "--doc",
    "doc_id",
    default=None,
    help="doc_id, напр. «ТУ 27.31.11-131-47273194-2025»",
)
@click.option("--norm-id", "norm_id", default=None, type=int)
@click.option("--db", default="data/app.db", show_default=True)
def show_norm_catalog_cmd(
    doc_id: Optional[str],
    norm_id: Optional[int],
    db: str,
) -> None:
    """ТУ целиком: метаданные + acceptance_items с пунктами (JOIN)."""
    from .persistence.sqlite_repo import migrate_db, show_norm_catalog

    if not doc_id and norm_id is None:
        click.echo("Укажите --doc или --norm-id", err=True)
        raise SystemExit(2)
    migrate_db(db)
    try:
        cat = show_norm_catalog(doc_id=doc_id, norm_document_id=norm_id, db_path=db)
    except ValueError as e:
        click.echo(str(e), err=True)
        raise SystemExit(2) from e
    if not cat:
        click.echo("Документ не найден", err=True)
        raise SystemExit(1)
    click.echo(
        f"{cat.get('doc_id')}  kind={cat.get('kind')}  "
        f"status={cat.get('status')}  format={cat.get('source_format') or '—'}"
    )
    click.echo(f"title: {cat.get('title')}")
    if cat.get("manufacturer_hint"):
        click.echo(f"manufacturer: {cat['manufacturer_hint']}")
    if cat.get("file_path"):
        click.echo(f"file_path (local only): {cat['file_path']}")
    items = cat.get("acceptance_items") or []
    click.echo(f"acceptance_items: {len(items)}")
    for it in items:
        reqs = [
            c["clause"]
            for c in (it.get("clauses") or [])
            if c.get("role") == "requirement"
        ]
        meths = [
            c["clause"]
            for c in (it.get("clauses") or [])
            if c.get("role") == "method_internal"
        ]
        exts = [e.get("ext_doc_id") for e in (it.get("method_external") or [])]
        bill = "billable" if it.get("billable") else "n/a-price"
        click.echo(
            f"  [{it['id']}] {it.get('name_exact')}  ({bill})\n"
            f"      req={','.join(reqs) or '—'}  "
            f"method_tu={','.join(meths) or '—'}  "
            f"ext={','.join(x for x in exts if x) or '—'}"
        )


@cli.command("add-acceptance-item")
@click.option("--doc", "doc_id", required=True, help="doc_id ТУ (уже в norm_documents)")
@click.option("--name", "name_exact", required=True, help="Точное имя из таблицы приёмки")
@click.option(
    "--req",
    "req_clauses",
    multiple=True,
    help="Пункт требований (повторяемый; один пункт за раз, не диапазон)",
)
@click.option(
    "--method",
    "method_clauses",
    multiple=True,
    help="Пункт методов ТУ 5.x (повторяемый)",
)
@click.option("--category", default=None, help="psi|periodic|type|other (опц.)")
@click.option("--code", default=None, help="price_test_code или пусто")
@click.option(
    "--billable/--no-billable",
    default=True,
    help="В прайсе / не в прайсе (маркировка — --no-billable)",
)
@click.option("--sort", "sort_order", default=0, type=int)
@click.option(
    "--ext-doc",
    default=None,
    help="Внешний ГОСТ/IEC (один; детальнее — волна 2/CLI позже)",
)
@click.option("--ext-method", default=None, help="Метод/п. внешнего НД")
@click.option("--db", default="data/app.db", show_default=True)
def add_acceptance_item_cmd(
    doc_id: str,
    name_exact: str,
    req_clauses: tuple[str, ...],
    method_clauses: tuple[str, ...],
    category: Optional[str],
    code: Optional[str],
    billable: bool,
    sort_order: int,
    ext_doc: Optional[str],
    ext_method: Optional[str],
    db: str,
) -> None:
    """Вручную добавить строку каталога приёмки (HITL / до парсера docx)."""
    from .persistence.sqlite_repo import add_acceptance_item, migrate_db, upsert_norm_document

    migrate_db(db)
    # если doc ещё нет — минимальная карточка (без пути к ТУ)
    upsert_norm_document(doc_id, doc_id, kind="tu", status="draft", db_path=db)
    ext_list = []
    if ext_doc:
        ext_list.append(
            {"ext_doc_id": ext_doc, "ext_clause_or_method": ext_method or ""}
        )
    item_id = add_acceptance_item(
        doc_id=doc_id,
        name_exact=name_exact,
        requirement_clauses=list(req_clauses),
        method_clauses=list(method_clauses),
        test_category=category,
        price_test_code=code,
        billable=billable,
        sort_order=sort_order,
        method_external=ext_list or None,
        db_path=db,
    )
    click.echo(click.style(f"✓ acceptance_item id={item_id}", fg="green"))


@cli.command("import-acceptance-docx")
@click.option(
    "--file",
    "file_path",
    required=True,
    type=click.Path(exists=True, dir_okay=False),
    help="Локальный .docx ТУ (не коммитить в git)",
)
@click.option("--doc-id", default=None, help="Переопределить doc_id в БД")
@click.option(
    "--replace/--no-replace",
    default=True,
    help="Удалить старые acceptance_items этого ТУ перед импортом",
)
@click.option(
    "--match-price/--no-match-price",
    default=False,
    help="После импорта попытаться проставить price_test_code",
)
@click.option("--db", default="data/app.db", show_default=True)
def import_acceptance_docx_cmd(
    file_path: str,
    doc_id: Optional[str],
    replace: bool,
    match_price: bool,
    db: str,
) -> None:
    """Импорт таблицы приёмки из docx (поколение A) → acceptance_items."""
    from .generation.acceptance_table_import import (
        import_acceptance_docx,
        try_match_prices,
    )
    from .persistence.sqlite_repo import migrate_db

    migrate_db(db)
    try:
        result = import_acceptance_docx(
            file_path, db_path=db, replace=replace, doc_id=doc_id
        )
    except Exception as e:
        click.echo(click.style(f"✗ {e}", fg="red"), err=True)
        raise SystemExit(1) from e
    click.echo(
        click.style(
            f"✓ {result['doc_id']}: items={result['items']} "
            f"(deleted_before={result.get('deleted_before', 0)})",
            fg="green",
        )
    )
    for w in result.get("warnings") or []:
        click.echo(click.style(f"  ! {w}", fg="yellow"))
    if match_price:
        m = try_match_prices(doc_id=result["doc_id"], db_path=db)
        click.echo(f"  price match: {m['matched']}/{m['scanned']}")


@cli.command("import-acceptance-etalons")
@click.option(
    "--replace/--no-replace",
    default=True,
    help="Перезаписать acceptance_items эталонов",
)
@click.option(
    "--match-price/--no-match-price",
    default=True,
    help="Проставить price_test_code где удаётся (HITL потом)",
)
@click.option("--db", default="data/app.db", show_default=True)
def import_acceptance_etalons_cmd(
    replace: bool,
    match_price: bool,
    db: str,
) -> None:
    """Импорт эталонов ТЗ v3: 131 + 141 (docx) + 005 (raw_text fallback)."""
    from .generation.acceptance_table_import import (
        import_etalon_batch,
        try_match_prices,
    )
    from .persistence.sqlite_repo import migrate_db

    migrate_db(db)
    results = import_etalon_batch(db_path=db, replace=replace)
    total = 0
    for r in results:
        if r.get("error"):
            click.echo(click.style(f"✗ {r.get('doc_id')}: {r['error']}", fg="red"))
            continue
        total += int(r.get("items") or 0)
        note = r.get("note") or r.get("source_format") or ""
        click.echo(
            click.style(
                f"✓ {r['doc_id']}: items={r['items']}  {note}",
                fg="green",
            )
        )
        for w in r.get("warnings") or []:
            click.echo(click.style(f"  ! {w}", fg="yellow"))
        if match_price and r.get("doc_id"):
            m = try_match_prices(doc_id=r["doc_id"], db_path=db)
            click.echo(f"  price match: {m['matched']}/{m['scanned']}")
    click.echo(f"Итого строк: {total}")


@cli.command("list-requirements")
@click.option("--doc-id", "norm_id", default=None, type=int, help="id norm_documents")
@click.option("--db", default="data/app.db", show_default=True)
def list_requirements_cmd(norm_id: Optional[int], db: str) -> None:
    """Пункты требований (примеры + будущий импорт из ТУ)."""
    from .persistence.sqlite_repo import list_requirements, migrate_db

    migrate_db(db)
    for r in list_requirements(norm_document_id=norm_id, db_path=db):
        click.echo(
            f"{r['doc_id']:24}  п.{(r.get('clause') or '—'):8}  "
            f"{(r.get('title') or '')[:50]}"
        )


@cli.command("list-test-aliases")
@click.option("--db", default="data/app.db", show_default=True)
def list_test_aliases_cmd(db: str) -> None:
    """Синонимы названий испытаний → канон / код прайса."""
    from .persistence.sqlite_repo import list_test_aliases, migrate_db

    migrate_db(db)
    for r in list_test_aliases(db_path=db):
        click.echo(
            f"{r['alias_norm'][:30]:30} → {r['canonical_name'][:40]:40}  "
            f"code={r.get('price_test_code') or '—'}"
        )


@cli.command("add-test-alias")
@click.option("--alias", required=True)
@click.option("--canonical", required=True, help="Каноническое имя")
@click.option("--code", default=None, help="Код test_items")
@click.option("--db", default="data/app.db", show_default=True)
def add_test_alias_cmd(alias: str, canonical: str, code: Optional[str], db: str) -> None:
    """Добавить синоним испытания."""
    from .persistence.sqlite_repo import add_test_alias, migrate_db

    migrate_db(db)
    i = add_test_alias(alias, canonical, price_test_code=code, db_path=db)
    click.echo(click.style(f"✓ alias id={i}", fg="green"))


@cli.command("import-norm-text")
@click.option(
    "--file",
    "file_path",
    default=None,
    type=click.Path(exists=True),
    help="Один .txt файл ТУ",
)
@click.option(
    "--dir",
    "dir_path",
    default=None,
    type=click.Path(exists=True, file_okay=False),
    help="Каталог raw_text/*.txt (пакетный импорт S5)",
)
@click.option("--kind", default="tu", show_default=True, help="tu|gost|iec|other")
@click.option("--max-clauses", default=80, show_default=True, type=int)
@click.option(
    "--limit",
    default=0,
    show_default=True,
    type=int,
    help="Макс. файлов при --dir (0 = все)",
)
@click.option("--db", default="data/app.db", show_default=True)
def import_norm_text_cmd(
    file_path: Optional[str],
    dir_path: Optional[str],
    kind: str,
    max_clauses: int,
    limit: int,
    db: str,
) -> None:
    """Импорт пунктов требований из локального .txt ТУ (raw_text, не git)."""
    from pathlib import Path as P

    from .generation.norm_text_import import import_norm_from_text_file
    from .persistence.sqlite_repo import migrate_db

    if not file_path and not dir_path:
        click.echo("Укажите --file или --dir", err=True)
        raise SystemExit(2)

    migrate_db(db)
    paths: list[P] = []
    if file_path:
        paths.append(P(file_path))
    if dir_path:
        found = sorted(P(dir_path).glob("*.txt"))
        if limit and limit > 0:
            found = found[:limit]
        paths.extend(found)
    if not paths:
        click.echo("Нет .txt для импорта", err=True)
        raise SystemExit(1)

    total_clauses = 0
    ok = 0
    for path in paths:
        try:
            result = import_norm_from_text_file(
                path, kind=kind, db_path=db, max_clauses=max_clauses
            )
        except Exception as e:
            click.echo(click.style(f"✗ {path.name}: {e}", fg="red"), err=True)
            continue
        ok += 1
        total_clauses += int(result.get("clauses") or 0)
        click.echo(
            click.style(f"✓ {result['doc_id']}", fg="green")
            + f"  clauses≈{result['clauses']}  id={result['norm_document_id']}"
        )
    if len(paths) > 1:
        click.echo(f"Итого: файлов {ok}/{len(paths)}, clauses≈{total_clauses}")


@cli.command("import-aliases-yaml")
@click.option(
    "--file",
    "file_path",
    default="data/knowledge/manufacturer_v1/test_synonyms.yaml",
    show_default=True,
    type=click.Path(),
)
@click.option("--db", default="data/app.db", show_default=True)
def import_aliases_yaml_cmd(file_path: str, db: str) -> None:
    """Импорт test_synonyms.yaml → test_aliases (локальный knowledge, не git)."""
    from pathlib import Path as P

    from .generation.norm_text_import import import_aliases_from_synonyms_yaml
    from .persistence.sqlite_repo import migrate_db

    migrate_db(db)
    if not P(file_path).is_file():
        click.echo(f"Файл не найден: {file_path}", err=True)
        raise SystemExit(1)
    n = import_aliases_from_synonyms_yaml(file_path, db_path=db)
    click.echo(click.style(f"✓ импортировано/обновлено aliases: {n}", fg="green"))


@cli.command("export-protocol-meta")
@click.option("--order-id", type=int, required=True, help="ID заказа")
@click.option(
    "--output",
    "-o",
    default=None,
    type=click.Path(),
    help="Путь к JSON (по умолчанию data/generated/protocol_meta_order…json)",
)
@click.option("--db", default="data/app.db", show_default=True)
def export_protocol_meta_cmd(order_id: int, output: Optional[str], db: str) -> None:
    """JSON-каркас протокола для protocol_generator (без измеренных значений)."""
    from .generation.protocol_meta_export import export_protocol_meta_for_order

    path = export_protocol_meta_for_order(
        order_id,
        output_path=output,
        db_path=db,
    )
    click.echo(click.style(f"✓ JSON: {path}", fg="green"))
    click.echo("Дальше (на машине с protocol_generator):")
    click.echo(f'  cd D:\\My_projects\\protocol_generator')
    click.echo(f'  .\\venv\\Scripts\\python.exe main.py "{path}"')


@cli.command("export-prod-data")
@click.option(
    "--output",
    "-o",
    default=None,
    type=click.Path(),
    help="Путь к .zip (по умолчанию data/training/exports/prod_data_<station>_<дата>.zip)",
)
@click.option("--db", default="data/app.db", show_default=True)
@click.option("--full", is_flag=True, help="Весь архив, не только дельта с прошлого экспорта")
@click.option("--note", default="", help="Комментарий оператора в manifest")
def export_prod_data_cmd(
    output: Optional[str],
    db: str,
    full: bool,
    note: str,
) -> None:
    """Экспорт данных prod (правки, снимки, ассистент) на машину разработки."""
    from datetime import datetime

    from .training.prod_data import export_prod_data, get_prod_station_id

    migrate_db(db)
    station = get_prod_station_id(db)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out = (
        Path(output)
        if output
        else Path("data/training/exports") / f"prod_data_{station}_{stamp}.zip"
    )
    result = export_prod_data(
        out,
        db_path=db,
        delta_only=not full,
        operator_note=note,
    )
    m = result["manifest"]
    click.echo(click.style(f"✓ Пакет: {result['path']}", fg="green"))
    click.echo(
        f"  station: {m.get('station_id') or m.get('host_id')}  delta: {m.get('delta_only')}"
    )
    for k, v in (m.get("counts") or {}).items():
        click.echo(f"  {k}: {v}")


@cli.command("import-prod-data")
@click.argument("archive", type=click.Path(exists=True, dir_okay=False))
@click.option("--db", default="data/app.db", show_default=True)
@click.option("--no-sync", is_flag=True, help="Не вызывать sync-corrections в БД")
def import_prod_data_cmd(archive: str, db: str, no_sync: bool) -> None:
    """Импорт пакета данных prod с рабочего ПК."""
    from .training.prod_data import import_prod_data

    migrate_db(db)
    result = import_prod_data(
        archive,
        db_path=db,
        sync_db=not no_sync,
    )
    m = result.get("manifest") or {}
    station = m.get("station_id") or m.get("host_id")
    click.echo(click.style("✓ Импорт завершён", fg="green"))
    click.echo(f"  источник: {m.get('host_name')} ({station})")
    click.echo(f"  экспорт:  {m.get('exported_at')}")
    for k, v in result["stats"].items():
        click.echo(f"  {k}: {v}")
    if result.get("sync_corrections"):
        click.echo(f"  sync_corrections: {result['sync_corrections']}")


@cli.command("demo-ocr-marks")
@click.option("--db", default="data/app.db", show_default=True)
@click.option(
    "--record/--no-record",
    default=True,
    help="Записать feedback (помогло/нет) в corrections + assistant_sessions",
)
@click.option(
    "--save-report/--no-save-report",
    default=True,
    help="JSON-отчёт в data/training/exports/reports/",
)
def demo_ocr_marks_cmd(db: str, record: bool, save_report: bool) -> None:
    """S2.5: демо MarkCorrector (+ LLM если включён) на 3 OCR-марках.

    Печатает таблицу raw → suggested → source → helped (yes/partial/no).
    LLM opt-in: Настройки GUI или ASSISTANT_LLM_ENABLED=1.
    """
    from .assistant.demo_marks import (
        format_demo_table,
        run_ocr_marks_demo,
        save_demo_report,
    )
    from .persistence.sqlite_repo import migrate_db

    migrate_db(db)
    report = run_ocr_marks_demo(db_path=db, record_feedback=record)
    click.echo(format_demo_table(report))
    if save_report:
        path = save_demo_report(report)
        click.echo(click.style(f"✓ Отчёт: {path}", fg="green"))
    c = report.get("counts") or {}
    if c.get("helped_yes", 0) + c.get("helped_partial", 0) >= 2:
        click.echo(click.style("S2.5: демо OK (детерминированный слой помогает)", fg="green"))
    else:
        click.echo(
            click.style(
                "S2.5: мало совпадений с эталоном — смотрите reason/source в таблице",
                fg="yellow",
            )
        )


@cli.command("assistant-llm-status")
@click.option("--db", default="data/app.db", show_default=True)
def assistant_llm_status_cmd(db: str) -> None:
    """Проверка настроек LLM и доступности Ollama."""
    from .assistant.llm_provider import check_ollama_health
    from .persistence.sqlite_repo import get_assistant_llm_settings, migrate_db

    migrate_db(db)
    settings = get_assistant_llm_settings(db)
    click.echo(f"enabled: {settings.enabled}")
    click.echo(f"model:   {settings.model}")
    click.echo(f"url:     {settings.base_url}")
    click.echo(f"models:  {settings.ollama_models_dir}")
    health = check_ollama_health(settings)
    if health.ok:
        click.echo(click.style(f"Ollama: {health.message}", fg="green"))
        for name in health.models[:10]:
            click.echo(f"  - {name}")
    else:
        click.echo(click.style(f"Ollama: {health.message}", fg="red"))


@cli.command("assistant-llm-test")
@click.argument("mark")
@click.option("--db", default="data/app.db", show_default=True)
@click.option("--enable/--no-enable", default=None, help="Принудительно вкл/выкл LLM для теста")
def assistant_llm_test_cmd(mark: str, db: str, enable: bool | None) -> None:
    """Тест подсказки ассистента для одной марки (детерминированный + LLM)."""
    from .assistant.mark_corrector import suggest_mark_correction
    from .persistence.sqlite_repo import get_assistant_llm_settings, migrate_db, save_assistant_llm_settings

    migrate_db(db)
    if enable is not None:
        settings = get_assistant_llm_settings(db)
        settings.enabled = enable
        save_assistant_llm_settings(settings, db)
    suggestion = suggest_mark_correction(mark, db_path=db)
    click.echo(f"raw:        {suggestion.raw}")
    click.echo(f"suggested:  {suggestion.suggested}")
    click.echo(f"changed:    {suggestion.changed}")
    click.echo(f"confidence: {suggestion.confidence:.0%}")
    click.echo(f"source:     {suggestion.source}")
    click.echo(f"reason:     {suggestion.reason}")


@cli.command("gui")
def gui_cmd() -> None:
    """Запускает графический интерфейс (tkinter)."""
    setup_logging(level="INFO")
    from .ui.gui import main

    main()


if __name__ == "__main__":
    cli()