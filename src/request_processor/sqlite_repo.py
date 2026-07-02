"""
sqlite_repo.py — слой работы с базой данных SQLite.

Это "репозиторий" (Repository pattern). 
Он полностью отвечает за:
- Создание таблиц
- Загрузку прайс-листа из Excel
- Сохранение и получение расчётов
- Seed демо-данных для быстрого старта

Почему отдельный файл:
- Вся работа с БД в одном месте → легко менять на PostgreSQL позже
- Чистый sqlite3 (без SQLAlchemy) — просто и понятно на старте
- Все запросы параметризованы (защита от SQL-инъекций)
- Используем контекстный менеджер для безопасной работы с соединением

Структура таблиц (Итерация 1):
- test_items — справочник испытаний (из прайс-листа)
- calculations — заголовок одного расчёта (марка + итоги)
- calculation_lines — детальные строки (каждое испытание в расчёте)
"""

from __future__ import annotations

import json
import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from .models import (
    Calculation,
    CalculationLine,
    CableMarkRecord,
    ClimaticTestSettings,
    OrganizationExtract,
    TestItem,
    TestItemUpdate,
    TestItemCreate,
)
from .organization_extractor import normalize_org_name
from .cable_mark_parser import parse_cable_mark_record
from .climatic_tests import CLIMATIC_TESTS, climatic_settings_fields
from .test_rules import DEFAULT_PRICE_XLSX, infer_rule_type

# Корень проекта (не зависит от текущей рабочей директории при запуске GUI/CLI)
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH_DEFAULT = PROJECT_ROOT / "data" / "app.db"
GENERATED_DIR_DEFAULT = PROJECT_ROOT / "data" / "generated"


def resolve_db_path(db_path: str | Path = DB_PATH_DEFAULT) -> Path:
    """Абсолютный путь к БД; относительные пути — от cwd, кроме стандартного data/app.db."""
    p = Path(db_path)
    if p.is_absolute():
        return p
    if p.as_posix() == "data/app.db":
        return (PROJECT_ROOT / p).resolve()
    return (Path.cwd() / p).resolve()


@contextmanager
def get_connection(db_path: str | Path = DB_PATH_DEFAULT):
    """
    Контекстный менеджер подключения к SQLite.
    
    Преимущества:
    - Автоматически commit при успехе и close при любом исходе
    - Включает foreign_keys (для целостности данных)
    - row_factory = sqlite3.Row → можно обращаться как к dict
    """
    db_path = resolve_db_path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path), detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """
    Создаёт структуру базы данных + seed демо-данных.
    
    Вызывается командой: python -m cli init-db
    """
    schema = """
    CREATE TABLE IF NOT EXISTS test_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        base_cost REAL NOT NULL,
        category TEXT,
        method TEXT,
        rule_type TEXT DEFAULT 'fixed',
        rule_params TEXT DEFAULT '{}'
    );

    CREATE TABLE IF NOT EXISTS calculations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mark TEXT NOT NULL,
        parsed_mark TEXT NOT NULL,           -- JSON строки CableMark
        total_cost_without_vat REAL NOT NULL,
        vat_rate REAL NOT NULL DEFAULT 0.22,
        total_cost_with_vat REAL NOT NULL,
        source TEXT DEFAULT 'manual',
        output_path TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS calculation_lines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        calculation_id INTEGER NOT NULL,
        test_item_id INTEGER,
        test_name TEXT NOT NULL,
        base_cost REAL NOT NULL,
        multiplier REAL DEFAULT 1.0,
        hours REAL,
        final_cost REAL NOT NULL,
        note TEXT,
        FOREIGN KEY (calculation_id) REFERENCES calculations(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS cable_marks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        full_mark TEXT NOT NULL UNIQUE,
        brand TEXT NOT NULL,
        fire_class TEXT,
        cores_count INTEGER NOT NULL,
        structural_element_type TEXT,
        structural_elements_count INTEGER,
        characteristic_size REAL NOT NULL,
        size_unit TEXT NOT NULL DEFAULT 'mm2',
        document TEXT,
        source TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS app_settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS organizations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        name_normalized TEXT NOT NULL,
        address TEXT,
        postal_code TEXT,
        phone TEXT,
        email TEXT,
        inn TEXT,
        kpp TEXT,
        is_accredited INTEGER NOT NULL DEFAULT 0,
        fsa_registry_number TEXT,
        org_type TEXT NOT NULL DEFAULT 'unknown',
        source TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS document_extractions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source_path TEXT NOT NULL,
        source_type TEXT NOT NULL,
        customer_org_id INTEGER,
        manufacturer_org_id INTEGER,
        subject TEXT,
        raw_text_length INTEGER,
        marks_count INTEGER NOT NULL DEFAULT 0,
        extracted_at TEXT NOT NULL,
        FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
        FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
    );

    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_org_id INTEGER,
        manufacturer_org_id INTEGER,
        subject TEXT,
        note TEXT,
        status TEXT NOT NULL DEFAULT 'kp_generated',
        total_without_vat REAL NOT NULL DEFAULT 0,
        total_with_vat REAL NOT NULL DEFAULT 0,
        vat_rate REAL NOT NULL DEFAULT 0.22,
        document_extraction_id INTEGER,
        kp_output_path TEXT,
        application_path TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
        FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id),
        FOREIGN KEY (document_extraction_id) REFERENCES document_extractions(id)
    );

    CREATE TABLE IF NOT EXISTS order_marks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        calculation_id INTEGER NOT NULL,
        cable_mark_id INTEGER,
        manufacturer_org_id INTEGER,
        mark TEXT NOT NULL,
        total_without_vat REAL NOT NULL,
        total_with_vat REAL NOT NULL,
        FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
        FOREIGN KEY (calculation_id) REFERENCES calculations(id),
        FOREIGN KEY (cable_mark_id) REFERENCES cable_marks(id),
        FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
    );

    CREATE INDEX IF NOT EXISTS idx_test_items_code ON test_items(code);
    CREATE INDEX IF NOT EXISTS idx_calculations_created_at ON calculations(created_at);
    CREATE INDEX IF NOT EXISTS idx_cable_marks_brand ON cable_marks(brand);
    CREATE UNIQUE INDEX IF NOT EXISTS idx_organizations_dedup
        ON organizations(COALESCE(inn, ''), name_normalized);
    CREATE INDEX IF NOT EXISTS idx_organizations_name ON organizations(name_normalized);
    CREATE INDEX IF NOT EXISTS idx_orders_created_at ON orders(created_at);
    CREATE INDEX IF NOT EXISTS idx_order_marks_order_id ON order_marks(order_id);
    """

    with get_connection(db_path) as conn:
        conn.executescript(schema)

    migrate_db(db_path)
    _seed_demo_tests(db_path)
    _seed_default_settings(db_path)
    print(f"База данных инициализирована: {db_path}")


def migrate_db(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """Добавляет новые таблицы в существующие БД без пересоздания."""
    with get_connection(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS cable_marks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_mark TEXT NOT NULL UNIQUE,
                brand TEXT NOT NULL,
                fire_class TEXT,
                cores_count INTEGER NOT NULL,
                structural_element_type TEXT,
                structural_elements_count INTEGER,
                characteristic_size REAL NOT NULL,
                size_unit TEXT NOT NULL DEFAULT 'mm2',
                document TEXT,
                source TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS organizations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                name_normalized TEXT NOT NULL,
                address TEXT,
                postal_code TEXT,
                phone TEXT,
                email TEXT,
                inn TEXT,
                kpp TEXT,
                is_accredited INTEGER NOT NULL DEFAULT 0,
                fsa_registry_number TEXT,
                org_type TEXT NOT NULL DEFAULT 'unknown',
                source TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS document_extractions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_path TEXT NOT NULL,
                source_type TEXT NOT NULL,
                customer_org_id INTEGER,
                manufacturer_org_id INTEGER,
                subject TEXT,
                raw_text_length INTEGER,
                marks_count INTEGER NOT NULL DEFAULT 0,
                extracted_at TEXT NOT NULL,
                FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
                FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
            );
            CREATE INDEX IF NOT EXISTS idx_cable_marks_brand ON cable_marks(brand);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_organizations_dedup
                ON organizations(COALESCE(inn, ''), name_normalized);
            CREATE INDEX IF NOT EXISTS idx_organizations_name ON organizations(name_normalized);
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_org_id INTEGER,
                manufacturer_org_id INTEGER,
                subject TEXT,
                note TEXT,
                status TEXT NOT NULL DEFAULT 'kp_generated',
                total_without_vat REAL NOT NULL DEFAULT 0,
                total_with_vat REAL NOT NULL DEFAULT 0,
                vat_rate REAL NOT NULL DEFAULT 0.22,
                document_extraction_id INTEGER,
                kp_output_path TEXT,
                application_path TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
                FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id),
                FOREIGN KEY (document_extraction_id) REFERENCES document_extractions(id)
            );
            CREATE TABLE IF NOT EXISTS order_marks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                calculation_id INTEGER NOT NULL,
                cable_mark_id INTEGER,
                manufacturer_org_id INTEGER,
                mark TEXT NOT NULL,
                total_without_vat REAL NOT NULL,
                total_with_vat REAL NOT NULL,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
                FOREIGN KEY (calculation_id) REFERENCES calculations(id),
                FOREIGN KEY (cable_mark_id) REFERENCES cable_marks(id),
                FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
            );
            CREATE INDEX IF NOT EXISTS idx_orders_created_at ON orders(created_at);
            CREATE INDEX IF NOT EXISTS idx_order_marks_order_id ON order_marks(order_id);
            """
        )
    _migrate_orders_columns(db_path)
    sync_climatic_tests(db_path)
    sync_test_rule_types(db_path)


def _migrate_orders_columns(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    with get_connection(db_path) as conn:
        cols = {row[1] for row in conn.execute("PRAGMA table_info(orders)").fetchall()}
        if "application_path" not in cols:
            conn.execute("ALTER TABLE orders ADD COLUMN application_path TEXT")


def get_calculation_lines(
    calculation_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    with get_connection(db_path) as conn:
        rows = conn.execute(
            """
            SELECT test_name, base_cost, multiplier, hours, final_cost, note
            FROM calculation_lines
            WHERE calculation_id = ?
            ORDER BY id
            """,
            (calculation_id,),
        ).fetchall()
        return [dict(row) for row in rows]


def get_cable_mark_document(
    full_mark: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> str | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT document FROM cable_marks WHERE full_mark = ? LIMIT 1",
            (full_mark,),
        ).fetchone()
        return row["document"] if row and row["document"] else None


def update_order_application_path(
    order_id: int,
    application_path: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            "UPDATE orders SET application_path = ?, updated_at = ? WHERE id = ?",
            (application_path, datetime.now().isoformat(), order_id),
        )


def sync_test_rule_types(db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Пересчитывает rule_type по названию/категории (per_core, per_group, time_based)."""
    updated = 0
    with get_connection(db_path) as conn:
        rows = conn.execute(
            "SELECT id, code, name, category FROM test_items"
        ).fetchall()
        for row in rows:
            rule_type, rule_params = infer_rule_type(
                row["name"], row["category"], row["code"]
            )
            conn.execute(
                """
                UPDATE test_items
                SET rule_type = ?, rule_params = ?
                WHERE id = ?
                """,
                (
                    rule_type,
                    json.dumps(rule_params, ensure_ascii=False),
                    row["id"],
                ),
            )
            updated += 1
    return updated


def sync_climatic_tests(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """Приводит климатические испытания к time_based с актуальными названиями."""
    for spec in CLIMATIC_TESTS:
        item = TestItem(
            code=spec["code"],
            name=spec["name"],
            base_cost=spec["base_cost"],
            category="Внешние воздействующие факторы",
            rule_type="time_based",
            rule_params={
                "hours_key": spec["hours_key"],
                "default_hours": spec["default_hours"],
                "cost_per_hour": spec["cost_per_hour"],
            },
        )
        insert_test_item(item, db_path)


CLIMATIC_SETTINGS_KEY = "climatic_test_hours"


def _seed_default_settings(db_path: str | Path) -> None:
    if get_climatic_settings(db_path) is not None:
        return
    save_climatic_settings(ClimaticTestSettings(), db_path)


def get_climatic_settings(db_path: str | Path = DB_PATH_DEFAULT) -> ClimaticTestSettings | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT value FROM app_settings WHERE key = ?", (CLIMATIC_SETTINGS_KEY,)
        ).fetchone()
        if not row:
            return None
        return ClimaticTestSettings(**json.loads(row["value"]))


def save_climatic_settings(
    settings: ClimaticTestSettings,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO app_settings (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (CLIMATIC_SETTINGS_KEY, settings.model_dump_json()),
        )


def build_default_hours_map(db_path: str | Path = DB_PATH_DEFAULT) -> dict[str, float]:
    """Собирает часы выдержки из настроек + default_hours из time_based испытаний."""
    hours: dict[str, float] = {}
    settings = get_climatic_settings(db_path)
    if settings:
        for key, _ in climatic_settings_fields():
            hours[key] = float(getattr(settings, key))
    for item in get_all_test_items(db_path):
        if item.rule_type != "time_based":
            continue
        key = item.rule_params.get("hours_key", item.code)
        if key not in hours and "default_hours" in item.rule_params:
            hours[key] = float(item.rule_params["default_hours"])
    return hours


def upsert_cable_mark(
    record: CableMarkRecord,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Добавляет марку без полных дублей (по full_mark)."""
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO cable_marks (
                full_mark, brand, fire_class, cores_count,
                structural_element_type, structural_elements_count,
                characteristic_size, size_unit, document, source, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(full_mark) DO UPDATE SET
                brand = excluded.brand,
                fire_class = excluded.fire_class,
                cores_count = excluded.cores_count,
                structural_element_type = excluded.structural_element_type,
                structural_elements_count = excluded.structural_elements_count,
                characteristic_size = excluded.characteristic_size,
                size_unit = excluded.size_unit,
                document = COALESCE(excluded.document, cable_marks.document),
                source = COALESCE(excluded.source, cable_marks.source)
            """,
            (
                record.full_mark,
                record.brand,
                record.fire_class,
                record.cores_count,
                record.structural_element_type,
                record.structural_elements_count,
                record.characteristic_size,
                record.size_unit,
                record.document,
                record.source,
                (record.created_at or datetime.now()).isoformat(),
            ),
        )
        if cursor.lastrowid:
            return cursor.lastrowid
        row = conn.execute(
            "SELECT id FROM cable_marks WHERE full_mark = ?", (record.full_mark,)
        ).fetchone()
        return int(row["id"]) if row else 0


def save_cable_marks_from_matches(
    matches: list,
    *,
    source: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, int]:
    """Сохраняет найденные в PDF марки в накопительную таблицу."""
    stats = {"saved": 0, "errors": 0}
    for match in matches:
        try:
            record = parse_cable_mark_record(
                match.mark,
                document=getattr(match, "document", None),
                context=getattr(match, "context", None),
            )
            record.source = source
            upsert_cable_mark(record, db_path)
            stats["saved"] += 1
        except Exception:
            stats["errors"] += 1
    return stats


def upsert_organization(
    extract: OrganizationExtract,
    *,
    source: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Сохраняет организацию без дублей (по ИНН или нормализованному названию)."""
    now = datetime.now().isoformat()
    name_normalized = normalize_org_name(extract.name)
    inn_key = extract.inn or ""

    with get_connection(db_path) as conn:
        row = None
        if extract.inn:
            row = conn.execute(
                "SELECT * FROM organizations WHERE inn = ?",
                (extract.inn,),
            ).fetchone()
        if row is None:
            row = conn.execute(
                "SELECT * FROM organizations WHERE name_normalized = ? AND COALESCE(inn, '') = ?",
                (name_normalized, inn_key),
            ).fetchone()

        if row:
            org_id = int(row["id"])
            conn.execute(
                """
                UPDATE organizations SET
                    name = ?,
                    address = COALESCE(?, address),
                    postal_code = COALESCE(?, postal_code),
                    phone = COALESCE(?, phone),
                    email = COALESCE(?, email),
                    inn = COALESCE(?, inn),
                    kpp = COALESCE(?, kpp),
                    is_accredited = MAX(is_accredited, ?),
                    fsa_registry_number = COALESCE(?, fsa_registry_number),
                    org_type = CASE WHEN ? = 'unknown' THEN org_type ELSE ? END,
                    source = COALESCE(?, source),
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    extract.name,
                    extract.address,
                    extract.postal_code,
                    extract.phone,
                    extract.email,
                    extract.inn,
                    extract.kpp,
                    int(extract.is_accredited),
                    extract.fsa_registry_number,
                    extract.org_type,
                    extract.org_type,
                    source,
                    now,
                    org_id,
                ),
            )
            return org_id

        cursor = conn.execute(
            """
            INSERT INTO organizations (
                name, name_normalized, address, postal_code, phone, email,
                inn, kpp, is_accredited, fsa_registry_number, org_type,
                source, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                extract.name,
                name_normalized,
                extract.address,
                extract.postal_code,
                extract.phone,
                extract.email,
                extract.inn,
                extract.kpp,
                int(extract.is_accredited),
                extract.fsa_registry_number,
                extract.org_type,
                source,
                now,
                now,
            ),
        )
        return int(cursor.lastrowid or 0)


def save_organizations_from_extraction(
    organizations: list[OrganizationExtract],
    *,
    source: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, int | None]:
    """Сохраняет организации из заявки; возвращает id заказчика и производителя."""
    customer_id: int | None = None
    manufacturer_id: int | None = None

    for org in organizations:
        org_id = upsert_organization(org, source=source, db_path=db_path)
        if org.role == "customer" and customer_id is None:
            customer_id = org_id
        if org.role == "manufacturer" and manufacturer_id is None:
            manufacturer_id = org_id

    if customer_id is None and organizations:
        customer_id = upsert_organization(organizations[0], source=source, db_path=db_path)
    if manufacturer_id is None and len(organizations) > 1:
        manufacturer_id = upsert_organization(organizations[1], source=source, db_path=db_path)
    elif manufacturer_id is None and customer_id is not None:
        manufacturer_id = customer_id

    return {"customer_org_id": customer_id, "manufacturer_org_id": manufacturer_id}


def save_document_extraction(
    *,
    source_path: str,
    source_type: str,
    text: str,
    marks_count: int,
    customer_org_id: int | None = None,
    manufacturer_org_id: int | None = None,
    subject: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO document_extractions (
                source_path, source_type, customer_org_id, manufacturer_org_id,
                subject, raw_text_length, marks_count, extracted_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_path,
                source_type,
                customer_org_id,
                manufacturer_org_id,
                subject,
                len(text),
                marks_count,
                datetime.now().isoformat(),
            ),
        )
        return int(cursor.lastrowid or 0)


def list_organizations(
    search: str | None = None,
    org_type: str | None = None,
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    query = "SELECT * FROM organizations"
    params: list[Any] = []
    conditions: list[str] = []
    if search:
        conditions.append("(name LIKE ? OR inn LIKE ? OR address LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
    if org_type:
        conditions.append("org_type = ?")
        params.append(org_type)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY updated_at DESC LIMIT ?"
    params.append(limit)

    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def get_organization_by_id(
    org_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    with get_connection(db_path) as conn:
        row = conn.execute("SELECT * FROM organizations WHERE id = ?", (org_id,)).fetchone()
        return dict(row) if row else None


def update_organization(
    org_id: int,
    *,
    name: str,
    address: str | None = None,
    postal_code: str | None = None,
    phone: str | None = None,
    email: str | None = None,
    inn: str | None = None,
    kpp: str | None = None,
    is_accredited: bool = False,
    fsa_registry_number: str | None = None,
    org_type: str = "unknown",
    db_path: str | Path = DB_PATH_DEFAULT,
) -> bool:
    """Обновляет организацию по id (ручное редактирование в GUI)."""
    now = datetime.now().isoformat()
    name_normalized = normalize_org_name(name)
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            UPDATE organizations SET
                name = ?,
                name_normalized = ?,
                address = ?,
                postal_code = ?,
                phone = ?,
                email = ?,
                inn = ?,
                kpp = ?,
                is_accredited = ?,
                fsa_registry_number = ?,
                org_type = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                name.strip(),
                name_normalized,
                address,
                postal_code,
                phone,
                email,
                inn,
                kpp,
                int(is_accredited),
                fsa_registry_number,
                org_type,
                now,
                org_id,
            ),
        )
        return cursor.rowcount > 0


def get_last_document_extraction(
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    """Последняя обработанная заявка (для панели сводки в GUI)."""
    with get_connection(db_path) as conn:
        row = conn.execute(
            """
            SELECT d.*,
                   c.name AS customer_name,
                   m.name AS manufacturer_name
            FROM document_extractions d
            LEFT JOIN organizations c ON c.id = d.customer_org_id
            LEFT JOIN organizations m ON m.id = d.manufacturer_org_id
            ORDER BY d.extracted_at DESC
            LIMIT 1
            """
        ).fetchone()
        return dict(row) if row else None


def find_organization_id_by_name(
    name: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int | None:
    if not name or not name.strip():
        return None
    normalized = normalize_org_name(name)
    with get_connection(db_path) as conn:
        row = conn.execute(
            """
            SELECT id FROM organizations
            WHERE name_normalized = ? OR name = ?
            ORDER BY updated_at DESC
            LIMIT 1
            """,
            (normalized, name.strip()),
        ).fetchone()
        return int(row["id"]) if row else None


def _find_cable_mark_id(mark: str, db_path: str | Path) -> int | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id FROM cable_marks WHERE full_mark = ? LIMIT 1",
            (mark,),
        ).fetchone()
        return int(row["id"]) if row else None


def create_order_from_kp(
    *,
    customer_name: str,
    manufacturer_name: str | None = None,
    customer_org_id: int | None = None,
    manufacturer_org_id: int | None = None,
    subject: str,
    note: str | None = None,
    calculation_ids: list[int],
    kp_output_path: str,
    document_extraction_id: int | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """
    Создаёт заказ после формирования КП.
    Заказчик — на уровне заказа; каждая марка связана с производителем.
    """
    if not calculation_ids:
        raise ValueError("Нет расчётов для заказа")

    if customer_org_id is None and customer_name:
        customer_org_id = find_organization_id_by_name(customer_name, db_path)
    if manufacturer_org_id is None and manufacturer_name:
        manufacturer_org_id = find_organization_id_by_name(manufacturer_name, db_path)
    if manufacturer_org_id is None and document_extraction_id:
        with get_connection(db_path) as conn:
            row = conn.execute(
                "SELECT manufacturer_org_id FROM document_extractions WHERE id = ?",
                (document_extraction_id,),
            ).fetchone()
            if row and row["manufacturer_org_id"]:
                manufacturer_org_id = int(row["manufacturer_org_id"])

    rows = get_calculations_for_kp(calculation_ids, db_path=db_path)
    if not rows:
        raise ValueError("Расчёты не найдены")

    total_without = round(sum(float(r["total_cost_without_vat"]) for r in rows), 2)
    total_with = round(sum(float(r["total_cost_with_vat"]) for r in rows), 2)
    vat_rate = float(rows[0].get("vat_rate") or 0.22)
    now = datetime.now().isoformat()

    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO orders (
                customer_org_id, manufacturer_org_id, subject, note, status,
                total_without_vat, total_with_vat, vat_rate,
                document_extraction_id, kp_output_path, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 'kp_generated', ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                customer_org_id,
                manufacturer_org_id,
                subject,
                note,
                total_without,
                total_with,
                vat_rate,
                document_extraction_id,
                kp_output_path,
                now,
                now,
            ),
        )
        order_id = int(cursor.lastrowid or 0)

        for row in rows:
            calc_id = int(row["id"])
            mark = row["mark"]
            cable_mark_id = _find_cable_mark_id(mark, db_path)
            conn.execute(
                """
                INSERT INTO order_marks (
                    order_id, calculation_id, cable_mark_id, manufacturer_org_id,
                    mark, total_without_vat, total_with_vat
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    order_id,
                    calc_id,
                    cable_mark_id,
                    manufacturer_org_id,
                    mark,
                    float(row["total_cost_without_vat"]),
                    float(row["total_cost_with_vat"]),
                ),
            )
        return order_id


def list_orders(
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    with get_connection(db_path) as conn:
        rows = conn.execute(
            """
            SELECT o.*,
                   c.name AS customer_name,
                   m.name AS manufacturer_name,
                   (SELECT COUNT(*) FROM order_marks om WHERE om.order_id = o.id) AS marks_count
            FROM orders o
            LEFT JOIN organizations c ON c.id = o.customer_org_id
            LEFT JOIN organizations m ON m.id = o.manufacturer_org_id
            ORDER BY o.created_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def get_order_details(
    order_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    with get_connection(db_path) as conn:
        order_row = conn.execute(
            """
            SELECT o.*,
                   c.name AS customer_name,
                   c.inn AS customer_inn,
                   c.address AS customer_address,
                   c.phone AS customer_phone,
                   m.name AS manufacturer_name,
                   m.inn AS manufacturer_inn,
                   m.address AS manufacturer_address,
                   d.source_path AS source_document
            FROM orders o
            LEFT JOIN organizations c ON c.id = o.customer_org_id
            LEFT JOIN organizations m ON m.id = o.manufacturer_org_id
            LEFT JOIN document_extractions d ON d.id = o.document_extraction_id
            WHERE o.id = ?
            """,
            (order_id,),
        ).fetchone()
        if not order_row:
            return None

        marks = conn.execute(
            """
            SELECT om.*,
                   mo.name AS manufacturer_name
            FROM order_marks om
            LEFT JOIN organizations mo ON mo.id = om.manufacturer_org_id
            WHERE om.order_id = ?
            ORDER BY om.id
            """,
            (order_id,),
        ).fetchall()

        result = dict(order_row)
        result["marks"] = [dict(m) for m in marks]
        return result


def list_cable_marks(
    search: str | None = None,
    limit: int = 200,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    query = "SELECT * FROM cable_marks"
    params: list[Any] = []
    if search:
        query += " WHERE full_mark LIKE ? OR brand LIKE ?"
        params.extend([f"%{search}%", f"%{search}%"])
    query += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)

    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def _seed_demo_tests(db_path: str | Path) -> None:
    """Добавляет демо-тесты с реалистичными параметрами."""
    demo = [
        TestItem(
            code="resistance_core",
            name="Электрическое сопротивление ТПЖ",
            base_cost=400,
            category="Электрические параметры НЧ",
            rule_type="per_core",
        ),
        TestItem(
            code="insulation_resistance",
            name="Электрическое сопротивление изоляции ТПЖ",
            base_cost=600,
            category="Электрические параметры НЧ",
            rule_type="per_core",
        ),
        TestItem(
            code="voltage_test",
            name="Испытание напряжением",
            base_cost=400,
            category="Электрические параметры НЧ",
        ),
    ]

    existing = get_all_test_items(db_path)
    if len(existing) < 5:
        for item in demo:
            insert_test_item(item, db_path)
    sync_climatic_tests(db_path)


def insert_test_item(item: TestItem, db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Вставляет или обновляет элемент прайс-листа по коду."""
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO test_items (code, name, base_cost, category, method, rule_type, rule_params)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                name=excluded.name,
                base_cost=excluded.base_cost,
                category=excluded.category,
                method=excluded.method,
                rule_type=excluded.rule_type,
                rule_params=excluded.rule_params
            """,
            (
                item.code,
                item.name,
                item.base_cost,
                item.category,
                item.method,
                item.rule_type,
                json.dumps(item.rule_params, ensure_ascii=False),
            ),
        )
        return cursor.lastrowid or 0


def get_test_item_by_code(code: str, db_path: str | Path = DB_PATH_DEFAULT) -> TestItem | None:
    """Получает TestItem по короткому коду (используется в калькуляторе)."""
    with get_connection(db_path) as conn:
        row = conn.execute("SELECT * FROM test_items WHERE code = ?", (code,)).fetchone()
        if not row:
            return None
        data = dict(row)
        data["rule_params"] = json.loads(data.get("rule_params") or "{}")
        return TestItem(**data)


def get_all_test_items(db_path: str | Path = DB_PATH_DEFAULT) -> list[TestItem]:
    """Возвращает все тесты (для отладки и load-data)."""
    with get_connection(db_path) as conn:
        rows = conn.execute("SELECT * FROM test_items ORDER BY id").fetchall()
        return [
            TestItem(**{**dict(row), "rule_params": json.loads(row["rule_params"] or "{}")})
            for row in rows
        ]


def save_calculation(calc: Calculation, db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Сохраняет расчёт + все строки детализации. Возвращает id расчёта."""
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO calculations
                (mark, parsed_mark, total_cost_without_vat, vat_rate,
                 total_cost_with_vat, source, output_path, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                calc.mark,
                calc.parsed_mark.model_dump_json(),
                calc.total_cost_without_vat,
                calc.vat_rate,
                calc.total_cost_with_vat,
                calc.source,
                calc.output_path,
                calc.created_at.isoformat(),
            ),
        )
        calc_id = cursor.lastrowid

        for line in calc.lines:
            conn.execute(
                """
                INSERT INTO calculation_lines
                    (calculation_id, test_item_id, test_name, base_cost,
                     multiplier, hours, final_cost, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    calc_id,
                    line.test_item_id,
                    line.test_name,
                    line.base_cost,
                    line.multiplier,
                    line.hours,
                    line.final_cost,
                    line.note,
                ),
            )
        return calc_id or 0


def get_recent_calculations(limit: int = 10, db_path: str | Path = DB_PATH_DEFAULT) -> list[dict[str, Any]]:
    """История последних расчётов (для команды history)."""
    with get_connection(db_path) as conn:
        rows = conn.execute(
            """
            SELECT id, mark, total_cost_with_vat, source, created_at
            FROM calculations ORDER BY created_at DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def get_calculations_for_kp(
    calculation_ids: list[int] | None = None,
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    """Возвращает расчёты для формирования КП (с итогами по марке)."""
    with get_connection(db_path) as conn:
        if calculation_ids:
            placeholders = ",".join("?" * len(calculation_ids))
            rows = conn.execute(
                f"""
                SELECT id, mark, total_cost_without_vat, total_cost_with_vat,
                       vat_rate, source, created_at, output_path
                FROM calculations
                WHERE id IN ({placeholders})
                ORDER BY id
                """,
                calculation_ids,
            ).fetchall()
            return [dict(row) for row in rows]
        rows = conn.execute(
            """
            SELECT id, mark, total_cost_without_vat, total_cost_with_vat,
                   vat_rate, source, created_at, output_path
            FROM calculations ORDER BY created_at DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def update_calculation_output_path(
    calculation_id: int,
    output_path: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            "UPDATE calculations SET output_path = ? WHERE id = ?",
            (output_path, calculation_id),
        )



'''---Загрузка прайс-листа из Excel (минимальная, версия)---'''


try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def _slugify(text: str) -> str:
    """Делает короткий код из русского названия теста."""
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "_", text)
    return text[:55]


def load_price_list_from_xlsx(
    xlsx_path: str | Path,
    db_path: str | Path = DB_PATH_DEFAULT,
    sheet_name: str = "Стоимость",
) -> int:
    """
    Загружает прайс-лист в таблицу test_items.
    
    Использует openpyxl. Берёт:
    - B (Наименование) → name + code (slug)
    - D (Стоимость) → base_cost
    - F (Категория)
    - G (Метод)
    
    rule_type определяется автоматически (per_core, per_group, time_based, fixed).
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl не установлен")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    count = 0
    with get_connection(db_path) as conn:
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[1]:
                continue
            name = str(row[1]).strip()
            if not name or name.lower() == "наименование":
                continue

            try:
                base_cost = float(row[3]) if row[3] is not None else 0.0
            except (TypeError, ValueError):
                base_cost = 0.0

            category = str(row[5]).strip() if row[5] else None
            method = str(row[6]).strip() if row[6] else None

            code = _slugify(name) or f"test_{idx}"
            rule_type, rule_params = infer_rule_type(name, category, code)

            item = TestItem(
                code=code,
                name=name,
                base_cost=base_cost,
                category=category,
                method=method,
                rule_type=rule_type,  # type: ignore[arg-type]
                rule_params=rule_params,
            )
            insert_test_item(item, db_path)
            count += 1

    print(f"Загружено {count} позиций прайс-листа")
    return count

'''---Новые функции управления справочником испытаний (Итерация 2)---'''

def add_test_item(item: TestItemCreate, db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Добавляет новое испытание в справочник test_items.
    
    Использует Pydantic-модель TestItemCreate для валидации входных данных.
    Если code уже существует — SQLite выбросит IntegrityError.
    
    Args:
        item: Валидированная модель с данными испытания
        db_path: Путь к базе данных
        
    Returns:
        id созданной записи в БД
    """
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO test_items 
            (code, name, base_cost, category, method, rule_type, rule_params)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.code,
                item.name,
                item.base_cost,
                item.category,
                item.method,
                item.rule_type,
                json.dumps(item.rule_params, ensure_ascii=False),
            ),
        )
        return cursor.lastrowid or 0


def update_test_item(
    code: str, 
    updates: TestItemUpdate, 
    db_path: str | Path = DB_PATH_DEFAULT
) -> bool:
    """Частично обновляет испытание по коду.
    
    Обновляет только те поля, которые реально были переданы в модели updates.
    Использует exclude_unset=True внутри Pydantic.
    
    Args:
        code: Уникальный код испытания (например, 'temp_low')
        updates: Модель TestItemUpdate с полями для обновления
        db_path: Путь к БД
        
    Returns:
        True, если была обновлена хотя бы одна запись
    """
    # Получаем только те поля, которые пользователь реально передал
    update_data = updates.model_dump(exclude_unset=True)
    if not update_data:
        return False

    # Если обновляем rule_params — сериализуем в JSON
    if "rule_params" in update_data and update_data["rule_params"] is not None:
        update_data["rule_params"] = json.dumps(update_data["rule_params"], ensure_ascii=False)

    set_clause = ", ".join(f"{k} = ?" for k in update_data.keys())
    values = list(update_data.values()) + [code]

    with get_connection(db_path) as conn:
        cursor = conn.execute(
            f"UPDATE test_items SET {set_clause} WHERE code = ?", 
            values
        )
        return cursor.rowcount > 0


def list_test_items(
    category: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 200,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict]:
    """Возвращает список испытаний с фильтрацией (для команды list-tests в CLI).
    
    Поддерживает фильтр по категории и поиск по названию/коду.
    """
    query = "SELECT * FROM test_items"
    params: list[Any] = []

    conditions: list[str] = []
    if category:
        conditions.append("category = ?")
        params.append(category)
    if search:
        conditions.append("(name LIKE ? OR code LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY category, name LIMIT ?"
    params.append(limit)

    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def bulk_upsert_test_items(
    items: list[TestItemCreate], 
    db_path: str | Path = DB_PATH_DEFAULT
) -> dict[str, int]:
    """Пакетное добавление/обновление испытаний (для команды import-tests).
    
    Использует INSERT OR REPLACE — если code уже есть, запись обновляется.
    Это удобно при повторной загрузке прайс-листа или его части.
    
    Returns:
        Статистика: сколько обработано и сколько было ошибок
    """
    stats = {"processed": 0, "errors": 0}

    with get_connection(db_path) as conn:
        for item in items:
            try:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO test_items 
                    (code, name, base_cost, category, method, rule_type, rule_params)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item.code,
                        item.name,
                        item.base_cost,
                        item.category,
                        item.method,
                        item.rule_type,
                        json.dumps(item.rule_params, ensure_ascii=False),
                    ),
                )
                stats["processed"] += 1
            except Exception:
                stats["errors"] += 1

        conn.commit()

    return stats