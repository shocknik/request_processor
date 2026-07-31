"""
sqlite_repo.py — слой работы с базой данных SQLite.

Это "репозиторий" (Repository pattern). 
Он полностью отвечает за:
- Создание таблиц
- Загрузку прайс-листа из Excel
- Сохранение и получение расчётов
- Seed демо-данных для быстрого старта

Почему отдельный файл:
- Вся работа с БД в одном месте → легко менять на PostgreSQL позже
- Чистый sqlite3 (без SQLAlchemy) — просто и понятно на старте
- Все запросы параметризованы (защита от SQL-инъекций)
- Используем контекстный менеджер для безопасной работы с соединением

Структура таблиц (Итерация 1):
- test_items — справочник испытаний (из прайс-листа)
- calculations — заголовок одного расчёта (марка + итоги)
- calculation_lines — детальные строки (каждое испытание в расчёте)
"""

from __future__ import annotations

import json
import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from ..models import (
    AssistantLlmSettings,
    Calculation,
    CalculationLine,
    CableMarkRecord,
    ClimaticTestSettings,
    DocumentPackSettings,
    OrganizationExtract,
    TestItem,
    TestItemUpdate,
    TestItemCreate,
)
from ..extraction.organization_extractor import normalize_org_name
from ..parsing.cable_mark_parser import parse_cable_mark_record
from ..calculation.climatic_tests import (
    CLIMATE_ITEM_ALIASES,
    CLIMATE_SLUG_BY_HOURS_KEY,
    CLIMATIC_TESTS,
    DEPRECATED_CLIMATE_ITEM_CODES,
    climatic_settings_fields,
)
from ..calculation.test_rules import DEFAULT_PRICE_XLSX, infer_rule_type

# Корень проекта (не зависит от текущей рабочей директории при запуске GUI/CLI)
from ..config import DB_PATH_DEFAULT, GENERATED_DIR_DEFAULT, PROJECT_ROOT
from ..logging_setup import get_logger

_log = get_logger("persistence")


def resolve_db_path(db_path: str | Path = DB_PATH_DEFAULT) -> Path:
    """Абсолютный путь к БД; относительные пути — от cwd, кроме стандартного data/app.db."""
    p = Path(db_path)
    if p.is_absolute():
        return p
    if p.as_posix() == "data/app.db":
        return (PROJECT_ROOT / p).resolve()
    return (Path.cwd() / p).resolve()


@contextmanager
def get_connection(db_path: str | Path = DB_PATH_DEFAULT):
    """
    Контекстный менеджер подключения к SQLite.
    
    Преимущества:
    - Автоматически commit при успехе и close при любом исходе
    - Включает foreign_keys (для целостности данных)
    - row_factory = sqlite3.Row → можно обращаться как к dict
    """
    db_path = resolve_db_path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path), detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """
    Создаёт структуру базы данных + seed демо-данных.
    
    Вызывается командой: python -m cli init-db
    """
    schema = """
    CREATE TABLE IF NOT EXISTS test_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        base_cost REAL NOT NULL,
        category TEXT,
        method TEXT,
        rule_type TEXT DEFAULT 'fixed',
        rule_params TEXT DEFAULT '{}'
    );

    CREATE TABLE IF NOT EXISTS calculations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mark TEXT NOT NULL,
        parsed_mark TEXT NOT NULL,           -- JSON строки CableMark
        total_cost_without_vat REAL NOT NULL,
        vat_rate REAL NOT NULL DEFAULT 0.22,
        total_cost_with_vat REAL NOT NULL,
        source TEXT DEFAULT 'manual',
        output_path TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS calculation_lines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        calculation_id INTEGER NOT NULL,
        test_item_id INTEGER,
        test_name TEXT NOT NULL,
        base_cost REAL NOT NULL,
        multiplier REAL DEFAULT 1.0,
        hours REAL,
        final_cost REAL NOT NULL,
        note TEXT,
        FOREIGN KEY (calculation_id) REFERENCES calculations(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS cable_marks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        full_mark TEXT NOT NULL UNIQUE,
        brand TEXT NOT NULL,
        fire_class TEXT,
        cores_count INTEGER NOT NULL,
        structural_element_type TEXT,
        structural_elements_count INTEGER,
        characteristic_size REAL NOT NULL,
        size_unit TEXT NOT NULL DEFAULT 'mm2',
        document TEXT,
        source TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS app_settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS organizations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        name_normalized TEXT NOT NULL,
        address TEXT,
        legal_address TEXT,
        actual_address TEXT,
        postal_code TEXT,
        phone TEXT,
        email TEXT,
        inn TEXT,
        kpp TEXT,
        is_accredited INTEGER NOT NULL DEFAULT 0,
        fsa_registry_number TEXT,
        org_type TEXT NOT NULL DEFAULT 'unknown',
        source TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS document_extractions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source_path TEXT NOT NULL,
        source_type TEXT NOT NULL,
        customer_org_id INTEGER,
        manufacturer_org_id INTEGER,
        subject TEXT,
        raw_text_length INTEGER,
        marks_count INTEGER NOT NULL DEFAULT 0,
        extracted_at TEXT NOT NULL,
        FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
        FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
    );

    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_org_id INTEGER,
        manufacturer_org_id INTEGER,
        subject TEXT,
        note TEXT,
        status TEXT NOT NULL DEFAULT 'kp_generated',
        total_without_vat REAL NOT NULL DEFAULT 0,
        total_with_vat REAL NOT NULL DEFAULT 0,
        vat_rate REAL NOT NULL DEFAULT 0.22,
        document_extraction_id INTEGER,
        kp_output_path TEXT,
        application_path TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
        FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id),
        FOREIGN KEY (document_extraction_id) REFERENCES document_extractions(id)
    );

    CREATE TABLE IF NOT EXISTS order_marks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        calculation_id INTEGER NOT NULL,
        cable_mark_id INTEGER,
        manufacturer_org_id INTEGER,
        mark TEXT NOT NULL,
        total_without_vat REAL NOT NULL,
        total_with_vat REAL NOT NULL,
        FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
        FOREIGN KEY (calculation_id) REFERENCES calculations(id),
        FOREIGN KEY (cable_mark_id) REFERENCES cable_marks(id),
        FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
    );

    CREATE INDEX IF NOT EXISTS idx_test_items_code ON test_items(code);
    CREATE INDEX IF NOT EXISTS idx_calculations_created_at ON calculations(created_at);
    CREATE INDEX IF NOT EXISTS idx_cable_marks_brand ON cable_marks(brand);
    CREATE UNIQUE INDEX IF NOT EXISTS idx_organizations_dedup
        ON organizations(COALESCE(inn, ''), name_normalized);
    CREATE INDEX IF NOT EXISTS idx_organizations_name ON organizations(name_normalized);
    CREATE INDEX IF NOT EXISTS idx_orders_created_at ON orders(created_at);
    CREATE INDEX IF NOT EXISTS idx_order_marks_order_id ON order_marks(order_id);

    CREATE TABLE IF NOT EXISTS test_applications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        output_path TEXT NOT NULL,
        template_path TEXT,
        application_number TEXT,
        test_type TEXT,
        customer_name TEXT,
        manufacturer_name TEXT,
        marks_count INTEGER NOT NULL DEFAULT 0,
        marks_snapshot TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_test_applications_order_id ON test_applications(order_id);
    CREATE INDEX IF NOT EXISTS idx_test_applications_created_at ON test_applications(created_at);
    """

    with get_connection(db_path) as conn:
        conn.executescript(schema)

    migrate_db(db_path)
    price = ensure_price_catalog(db_path)
    _seed_default_settings(db_path)
    _log.info(
        "init_db path=%s price_source=%s tests=%s",
        db_path,
        price.get("source"),
        price.get("after"),
        extra={"tag": "БД"},
    )
    print(f"База данных инициализирована: {db_path}")


def migrate_db(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """Добавляет новые таблицы в существующие БД без пересоздания."""
    with get_connection(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS cable_marks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_mark TEXT NOT NULL UNIQUE,
                brand TEXT NOT NULL,
                fire_class TEXT,
                cores_count INTEGER NOT NULL,
                structural_element_type TEXT,
                structural_elements_count INTEGER,
                characteristic_size REAL NOT NULL,
                size_unit TEXT NOT NULL DEFAULT 'mm2',
                document TEXT,
                source TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS organizations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                name_normalized TEXT NOT NULL,
                address TEXT,
                legal_address TEXT,
                actual_address TEXT,
                postal_code TEXT,
                phone TEXT,
                email TEXT,
                inn TEXT,
                kpp TEXT,
                is_accredited INTEGER NOT NULL DEFAULT 0,
                fsa_registry_number TEXT,
                org_type TEXT NOT NULL DEFAULT 'unknown',
                source TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS document_extractions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_path TEXT NOT NULL,
                source_type TEXT NOT NULL,
                customer_org_id INTEGER,
                manufacturer_org_id INTEGER,
                subject TEXT,
                raw_text_length INTEGER,
                marks_count INTEGER NOT NULL DEFAULT 0,
                extracted_at TEXT NOT NULL,
                FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
                FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
            );
            CREATE INDEX IF NOT EXISTS idx_cable_marks_brand ON cable_marks(brand);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_organizations_dedup
                ON organizations(COALESCE(inn, ''), name_normalized);
            CREATE INDEX IF NOT EXISTS idx_organizations_name ON organizations(name_normalized);
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_org_id INTEGER,
                manufacturer_org_id INTEGER,
                subject TEXT,
                note TEXT,
                status TEXT NOT NULL DEFAULT 'kp_generated',
                total_without_vat REAL NOT NULL DEFAULT 0,
                total_with_vat REAL NOT NULL DEFAULT 0,
                vat_rate REAL NOT NULL DEFAULT 0.22,
                document_extraction_id INTEGER,
                kp_output_path TEXT,
                application_path TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (customer_org_id) REFERENCES organizations(id),
                FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id),
                FOREIGN KEY (document_extraction_id) REFERENCES document_extractions(id)
            );
            CREATE TABLE IF NOT EXISTS order_marks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                calculation_id INTEGER NOT NULL,
                cable_mark_id INTEGER,
                manufacturer_org_id INTEGER,
                mark TEXT NOT NULL,
                total_without_vat REAL NOT NULL,
                total_with_vat REAL NOT NULL,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
                FOREIGN KEY (calculation_id) REFERENCES calculations(id),
                FOREIGN KEY (cable_mark_id) REFERENCES cable_marks(id),
                FOREIGN KEY (manufacturer_org_id) REFERENCES organizations(id)
            );
            CREATE INDEX IF NOT EXISTS idx_orders_created_at ON orders(created_at);
            CREATE INDEX IF NOT EXISTS idx_order_marks_order_id ON order_marks(order_id);
            CREATE TABLE IF NOT EXISTS test_applications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                output_path TEXT NOT NULL,
                template_path TEXT,
                application_number TEXT,
                test_type TEXT,
                customer_name TEXT,
                manufacturer_name TEXT,
                marks_count INTEGER NOT NULL DEFAULT 0,
                marks_snapshot TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_test_applications_order_id ON test_applications(order_id);
            CREATE INDEX IF NOT EXISTS idx_test_applications_created_at ON test_applications(created_at);
            CREATE TABLE IF NOT EXISTS test_mappings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                requirement_pattern TEXT NOT NULL UNIQUE,
                test_code TEXT NOT NULL,
                note TEXT,
                usage_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_test_mappings_code ON test_mappings(test_code);
            CREATE TABLE IF NOT EXISTS generated_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                calculation_id INTEGER,
                doc_type TEXT NOT NULL,
                file_path TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE SET NULL,
                FOREIGN KEY (calculation_id) REFERENCES calculations(id) ON DELETE SET NULL
            );
            CREATE INDEX IF NOT EXISTS idx_generated_documents_order_id
                ON generated_documents(order_id);
            CREATE INDEX IF NOT EXISTS idx_generated_documents_type
                ON generated_documents(doc_type);
            """
        )
    _migrate_orders_columns(db_path)
    _migrate_organizations_columns(db_path)
    _migrate_training_tables(db_path)
    _migrate_test_programs(db_path)
    _migrate_norm_requirements(db_path)
    _migrate_acceptance_catalog(db_path)
    _migrate_calculation_lines_quantity(db_path)
    apply_price_catalog_fixes(db_path)
    sync_climatic_tests(db_path)
    sync_test_rule_types(db_path)
    # Восстановить полный прайс, если БД «пустая» (частый кейс чистого install без -IncludeAppDb)
    ensure_price_catalog(db_path)
    sync_default_test_mappings(db_path)
    sync_mappings_from_test_item_names(db_path)
    seed_example_norm_requirements(db_path)
    seed_example_acceptance_catalog(db_path)


def _migrate_training_tables(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """Таблицы обучения и RAG (мастер-план 35c, Фаза 1)."""
    with get_connection(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS training_documents (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path       TEXT NOT NULL UNIQUE,
                file_hash       TEXT NOT NULL,
                file_name       TEXT NOT NULL,
                mime_type       TEXT,
                page_count      INTEGER,
                document_type   TEXT,
                document_family TEXT,
                source          TEXT DEFAULT 'operator',
                label_status    TEXT DEFAULT 'unlabeled',
                notes           TEXT,
                registered_at   TEXT NOT NULL,
                updated_at      TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_training_docs_type ON training_documents(document_type);
            CREATE INDEX IF NOT EXISTS idx_training_docs_label ON training_documents(label_status);
            CREATE INDEX IF NOT EXISTS idx_training_docs_hash ON training_documents(file_hash);

            CREATE TABLE IF NOT EXISTS training_labels (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                document_id     INTEGER NOT NULL REFERENCES training_documents(id),
                label_type      TEXT NOT NULL,
                label_version   INTEGER DEFAULT 1,
                payload_json    TEXT NOT NULL,
                labeled_by      TEXT DEFAULT 'operator',
                created_at      TEXT NOT NULL,
                is_active       INTEGER DEFAULT 1
            );
            CREATE INDEX IF NOT EXISTS idx_training_labels_doc
                ON training_labels(document_id, label_type);

            CREATE TABLE IF NOT EXISTS document_families (
                id              TEXT PRIMARY KEY,
                display_name    TEXT NOT NULL,
                document_type   TEXT NOT NULL,
                config_path     TEXT NOT NULL,
                sender_patterns TEXT,
                enabled         INTEGER DEFAULT 1,
                priority        INTEGER DEFAULT 100,
                created_at      TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS ocr_runs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                document_id     INTEGER REFERENCES training_documents(id),
                source_path     TEXT NOT NULL,
                engine          TEXT NOT NULL,
                dpi             INTEGER,
                preprocess      TEXT,
                page_count      INTEGER,
                mean_confidence REAL,
                duration_ms     INTEGER,
                cache_path      TEXT,
                created_at      TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS training_corrections (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                document_id     INTEGER REFERENCES training_documents(id),
                field_name      TEXT NOT NULL,
                original_value  TEXT,
                corrected_value TEXT,
                mark_context    TEXT,
                exported_from   TEXT,
                created_at      TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS rag_documents (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                title           TEXT NOT NULL,
                doc_kind        TEXT NOT NULL,
                file_path       TEXT,
                text_length     INTEGER,
                chunk_count     INTEGER DEFAULT 0,
                indexed_at      TEXT,
                metadata_json   TEXT
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_rag_documents_path
                ON rag_documents(file_path);

            CREATE TABLE IF NOT EXISTS rag_chunks (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                rag_document_id INTEGER NOT NULL REFERENCES rag_documents(id),
                chunk_index     INTEGER NOT NULL,
                chunk_text      TEXT NOT NULL,
                embedding_blob  BLOB,
                UNIQUE(rag_document_id, chunk_index)
            );

            CREATE TABLE IF NOT EXISTS assistant_sessions (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id        INTEGER,
                document_id     INTEGER,
                role            TEXT,
                message         TEXT NOT NULL,
                response        TEXT,
                model           TEXT,
                feedback        TEXT,
                created_at      TEXT NOT NULL
            );
            """
        )


def _migrate_test_programs(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """Программы испытаний (S4): заголовок + позиции."""
    with get_connection(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS test_programs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                name            TEXT NOT NULL,
                test_type       TEXT,
                cable_mark_text TEXT,
                tu_ref          TEXT,
                source_path     TEXT,
                notes           TEXT,
                created_at      TEXT NOT NULL,
                updated_at      TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_test_programs_name ON test_programs(name);

            CREATE TABLE IF NOT EXISTS test_program_items (
                id                   INTEGER PRIMARY KEY AUTOINCREMENT,
                program_id           INTEGER NOT NULL
                    REFERENCES test_programs(id) ON DELETE CASCADE,
                sort_order           INTEGER NOT NULL DEFAULT 0,
                name                 TEXT NOT NULL,
                requirement_doc      TEXT,
                requirement_clause   TEXT,
                method_doc           TEXT,
                method_clause        TEXT,
                price_test_code      TEXT,
                meta_json            TEXT,
                UNIQUE(program_id, sort_order, name)
            );
            CREATE INDEX IF NOT EXISTS idx_test_program_items_prog
                ON test_program_items(program_id);
            """
        )


def _migrate_norm_requirements(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """S5: нормативные документы, требования, aliases (задел под ПИ по ТУ)."""
    with get_connection(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS norm_documents (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                doc_id          TEXT NOT NULL UNIQUE,
                title           TEXT NOT NULL,
                kind            TEXT NOT NULL DEFAULT 'tu',
                -- kind: tu | gost | iec | method_std | pmi | other
                file_path       TEXT,
                notes           TEXT,
                created_at      TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS requirements (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                norm_document_id INTEGER NOT NULL
                    REFERENCES norm_documents(id) ON DELETE CASCADE,
                clause          TEXT NOT NULL,
                title           TEXT,
                body            TEXT,
                created_at      TEXT NOT NULL,
                UNIQUE(norm_document_id, clause)
            );
            CREATE INDEX IF NOT EXISTS idx_requirements_doc
                ON requirements(norm_document_id);

            CREATE TABLE IF NOT EXISTS requirement_test_links (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                requirement_id  INTEGER NOT NULL
                    REFERENCES requirements(id) ON DELETE CASCADE,
                price_test_code TEXT,
                program_item_hint TEXT,
                note            TEXT,
                UNIQUE(requirement_id, price_test_code)
            );

            CREATE TABLE IF NOT EXISTS test_aliases (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                alias_norm      TEXT NOT NULL UNIQUE,
                canonical_name  TEXT NOT NULL,
                price_test_code TEXT,
                source          TEXT DEFAULT 'manual',
                created_at      TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_test_aliases_code
                ON test_aliases(price_test_code);
            """
        )


def _migrate_acceptance_catalog(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """
    Волна 1 (ТЗ v3): каталог строк таблицы приёмки ТУ.

    Учитывает решения оператора:
    - group_code / test_category — опциональны;
    - пункты не диапазонами: связь item ↔ clause по одной;
    - внешний ГОСТ — отдельная таблица method_external_refs;
    - regime_json — плоский JSON на item (ветки по марке — v2);
    - ТУ-файлы не в git (file_path локальный).
    """
    with get_connection(db_path) as conn:
        # Расширение norm_documents / requirements (идемпотентно)
        nd_cols = {
            row[1] for row in conn.execute("PRAGMA table_info(norm_documents)").fetchall()
        }
        for col, decl in (
            ("edition_note", "TEXT"),
            ("source_format", "TEXT"),
            ("manufacturer_hint", "TEXT"),
            ("status", "TEXT NOT NULL DEFAULT 'draft'"),
        ):
            if col not in nd_cols:
                conn.execute(f"ALTER TABLE norm_documents ADD COLUMN {col} {decl}")

        req_cols = {
            row[1] for row in conn.execute("PRAGMA table_info(requirements)").fetchall()
        }
        if "clause_kind" not in req_cols:
            # requirement | method | note | ref
            conn.execute(
                "ALTER TABLE requirements ADD COLUMN clause_kind TEXT "
                "NOT NULL DEFAULT 'requirement'"
            )

        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS acceptance_items (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                norm_document_id    INTEGER NOT NULL
                    REFERENCES norm_documents(id) ON DELETE CASCADE,
                test_category       TEXT,
                -- psi | periodic | type | other | NULL (не жёстко)
                group_code          TEXT,
                -- С1/П1… опционально, в v1 часто NULL
                name_exact          TEXT NOT NULL,
                name_norm           TEXT NOT NULL,
                price_test_code     TEXT,
                billable            INTEGER NOT NULL DEFAULT 1,
                sort_order          INTEGER NOT NULL DEFAULT 0,
                regime_json         TEXT,
                notes               TEXT,
                status              TEXT NOT NULL DEFAULT 'draft',
                -- draft | reviewed | approved
                created_at          TEXT NOT NULL,
                UNIQUE(norm_document_id, name_norm, sort_order)
            );
            CREATE INDEX IF NOT EXISTS idx_acceptance_items_doc
                ON acceptance_items(norm_document_id);
            CREATE INDEX IF NOT EXISTS idx_acceptance_items_code
                ON acceptance_items(price_test_code);

            CREATE TABLE IF NOT EXISTS acceptance_item_clauses (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                acceptance_item_id  INTEGER NOT NULL
                    REFERENCES acceptance_items(id) ON DELETE CASCADE,
                requirement_id      INTEGER NOT NULL
                    REFERENCES requirements(id) ON DELETE CASCADE,
                role                TEXT NOT NULL,
                -- requirement | method_internal
                UNIQUE(acceptance_item_id, requirement_id, role)
            );
            CREATE INDEX IF NOT EXISTS idx_acceptance_item_clauses_item
                ON acceptance_item_clauses(acceptance_item_id);

            CREATE TABLE IF NOT EXISTS method_external_refs (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                acceptance_item_id  INTEGER NOT NULL
                    REFERENCES acceptance_items(id) ON DELETE CASCADE,
                ext_doc_id          TEXT NOT NULL,
                ext_doc_title       TEXT,
                ext_clause_or_method TEXT,
                note                TEXT,
                UNIQUE(acceptance_item_id, ext_doc_id, ext_clause_or_method)
            );
            CREATE INDEX IF NOT EXISTS idx_method_external_refs_item
                ON method_external_refs(acceptance_item_id);
            """
        )


def seed_example_norm_requirements(db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """1–2 примера норм (идемпотентно), чтобы UI/CLI не были пустыми."""
    now = datetime.now().isoformat()
    examples = [
        {
            "doc_id": "TU-16.K99-058-2014",
            "title": "ТУ 16.К99-058-2014 (пример структуры)",
            "kind": "tu",
            "clauses": [
                ("1.4.1", "Электрическое сопротивление жил", "resistance_core"),
                ("1.4.5", "Испытание напряжением", None),
            ],
        },
        {
            "doc_id": "GOST-7229-76",
            "title": "ГОСТ 7229-76 (метод, пример)",
            "kind": "gost",
            "clauses": [
                ("—", "Метод определения электрического сопротивления ТПЖ", "resistance_core"),
            ],
        },
    ]
    added = 0
    with get_connection(db_path) as conn:
        for ex in examples:
            cur = conn.execute(
                """
                INSERT INTO norm_documents (doc_id, title, kind, file_path, notes, created_at)
                VALUES (?, ?, ?, NULL, 'seed example S5', ?)
                ON CONFLICT(doc_id) DO NOTHING
                """,
                (ex["doc_id"], ex["title"], ex["kind"], now),
            )
            if cur.rowcount:
                added += 1
            row = conn.execute(
                "SELECT id FROM norm_documents WHERE doc_id = ?", (ex["doc_id"],)
            ).fetchone()
            if not row:
                continue
            nd_id = int(row["id"])
            for clause, title, code in ex["clauses"]:
                conn.execute(
                    """
                    INSERT INTO requirements
                        (norm_document_id, clause, title, body, created_at)
                    VALUES (?, ?, ?, NULL, ?)
                    ON CONFLICT(norm_document_id, clause) DO NOTHING
                    """,
                    (nd_id, clause, title, now),
                )
                if code:
                    req = conn.execute(
                        """
                        SELECT id FROM requirements
                        WHERE norm_document_id = ? AND clause = ?
                        """,
                        (nd_id, clause),
                    ).fetchone()
                    if req:
                        conn.execute(
                            """
                            INSERT INTO requirement_test_links
                                (requirement_id, price_test_code, program_item_hint, note)
                            VALUES (?, ?, ?, 'seed')
                            ON CONFLICT(requirement_id, price_test_code) DO NOTHING
                            """,
                            (int(req["id"]), code, title),
                        )
        # aliases examples (S4/S5: реальные коды прайса + ПМИ-формулировки)
        from ..mapping.program_price_matcher import PROGRAM_ALIAS_SEED

        for alias, canon, code in PROGRAM_ALIAS_SEED:
            conn.execute(
                """
                INSERT INTO test_aliases
                    (alias_norm, canonical_name, price_test_code, source, created_at)
                VALUES (?, ?, ?, 'seed', ?)
                ON CONFLICT(alias_norm) DO UPDATE SET
                    price_test_code = COALESCE(
                        excluded.price_test_code, test_aliases.price_test_code
                    ),
                    canonical_name = excluded.canonical_name
                """,
                (alias.lower(), canon, code, now),
            )
    return added


def _norm_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def ensure_requirement(
    norm_document_id: int,
    clause: str,
    *,
    title: str | None = None,
    body: str | None = None,
    clause_kind: str = "requirement",
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Создаёт или возвращает requirements.id. clause — один пункт (не диапазон)."""
    clause = (clause or "").strip()
    if not clause:
        raise ValueError("clause пустой")
    kind = (clause_kind or "requirement").strip() or "requirement"
    now = datetime.now().isoformat()
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO requirements
                (norm_document_id, clause, title, body, created_at, clause_kind)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(norm_document_id, clause) DO UPDATE SET
                title = COALESCE(excluded.title, requirements.title),
                body = COALESCE(excluded.body, requirements.body),
                clause_kind = COALESCE(excluded.clause_kind, requirements.clause_kind)
            """,
            (
                norm_document_id,
                clause,
                (title or "").strip() or None,
                body,
                now,
                kind,
            ),
        )
        row = conn.execute(
            """
            SELECT id FROM requirements
            WHERE norm_document_id = ? AND clause = ?
            """,
            (norm_document_id, clause),
        ).fetchone()
        if not row:
            raise RuntimeError("ensure_requirement failed")
        return int(row["id"])


def upsert_norm_document(
    doc_id: str,
    title: str,
    *,
    kind: str = "tu",
    file_path: str | None = None,
    notes: str | None = None,
    edition_note: str | None = None,
    source_format: str | None = None,
    manufacturer_hint: str | None = None,
    status: str = "draft",
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Создаёт/обновляет norm_documents, возвращает id."""
    now = datetime.now().isoformat()
    doc_id = (doc_id or "").strip()
    title = (title or "").strip() or doc_id
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO norm_documents (
                doc_id, title, kind, file_path, notes, created_at,
                edition_note, source_format, manufacturer_hint, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(doc_id) DO UPDATE SET
                title = excluded.title,
                kind = excluded.kind,
                file_path = COALESCE(excluded.file_path, norm_documents.file_path),
                notes = COALESCE(excluded.notes, norm_documents.notes),
                edition_note = COALESCE(excluded.edition_note, norm_documents.edition_note),
                source_format = COALESCE(
                    excluded.source_format, norm_documents.source_format
                ),
                manufacturer_hint = COALESCE(
                    excluded.manufacturer_hint, norm_documents.manufacturer_hint
                ),
                status = COALESCE(excluded.status, norm_documents.status)
            """,
            (
                doc_id,
                title,
                (kind or "tu").strip(),
                file_path,
                notes,
                now,
                edition_note,
                source_format,
                manufacturer_hint,
                (status or "draft").strip() or "draft",
            ),
        )
        row = conn.execute(
            "SELECT id FROM norm_documents WHERE doc_id = ?", (doc_id,)
        ).fetchone()
        if not row:
            raise RuntimeError("upsert_norm_document failed")
        return int(row["id"])


def add_acceptance_item(
    *,
    norm_document_id: int | None = None,
    doc_id: str | None = None,
    name_exact: str,
    requirement_clauses: list[str] | None = None,
    method_clauses: list[str] | None = None,
    test_category: str | None = None,
    group_code: str | None = None,
    price_test_code: str | None = None,
    billable: bool = True,
    sort_order: int = 0,
    regime: dict[str, Any] | str | None = None,
    notes: str | None = None,
    status: str = "draft",
    method_external: list[dict[str, str | None]] | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """
    Добавляет acceptance_item и связи с пунктами (по одному clause за раз).

    requirement_clauses / method_clauses — списки вида [\"2.5.1\"], не \"2.3.1-2.3.6\".
    """
    name_exact = (name_exact or "").strip()
    if not name_exact:
        raise ValueError("name_exact пустой")
    if norm_document_id is None:
        if not doc_id:
            raise ValueError("нужен norm_document_id или doc_id")
        with get_connection(db_path) as conn:
            row = conn.execute(
                "SELECT id FROM norm_documents WHERE doc_id = ?",
                (doc_id.strip(),),
            ).fetchone()
            if not row:
                raise ValueError(f"norm_documents не найден: {doc_id}")
            norm_document_id = int(row["id"])

    name_n = _norm_name(name_exact)
    if isinstance(regime, dict):
        regime_s = json.dumps(regime, ensure_ascii=False)
    else:
        regime_s = regime
    now = datetime.now().isoformat()

    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO acceptance_items (
                norm_document_id, test_category, group_code,
                name_exact, name_norm, price_test_code, billable,
                sort_order, regime_json, notes, status, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(norm_document_id, name_norm, sort_order) DO UPDATE SET
                name_exact = excluded.name_exact,
                test_category = excluded.test_category,
                group_code = excluded.group_code,
                price_test_code = excluded.price_test_code,
                billable = excluded.billable,
                regime_json = COALESCE(excluded.regime_json, acceptance_items.regime_json),
                notes = COALESCE(excluded.notes, acceptance_items.notes),
                status = excluded.status
            """,
            (
                norm_document_id,
                (test_category or "").strip() or None,
                (group_code or "").strip() or None,
                name_exact,
                name_n,
                (price_test_code or "").strip() or None,
                1 if billable else 0,
                int(sort_order),
                regime_s,
                notes,
                (status or "draft").strip() or "draft",
                now,
            ),
        )
        row = conn.execute(
            """
            SELECT id FROM acceptance_items
            WHERE norm_document_id = ? AND name_norm = ? AND sort_order = ?
            """,
            (norm_document_id, name_n, int(sort_order)),
        ).fetchone()
        if not row:
            raise RuntimeError("add_acceptance_item failed")
        item_id = int(row["id"])

    for cl in requirement_clauses or []:
        rid = ensure_requirement(
            norm_document_id,
            cl,
            title=name_exact,
            clause_kind="requirement",
            db_path=db_path,
        )
        _link_acceptance_clause(item_id, rid, "requirement", db_path=db_path)

    for cl in method_clauses or []:
        rid = ensure_requirement(
            norm_document_id,
            cl,
            title=f"Метод: {name_exact}",
            clause_kind="method",
            db_path=db_path,
        )
        _link_acceptance_clause(item_id, rid, "method_internal", db_path=db_path)

    for ext in method_external or []:
        add_method_external_ref(
            item_id,
            ext_doc_id=str(ext.get("ext_doc_id") or ext.get("doc") or ""),
            ext_doc_title=ext.get("ext_doc_title") or ext.get("title"),
            ext_clause_or_method=ext.get("ext_clause_or_method")
            or ext.get("method")
            or ext.get("clause"),
            note=ext.get("note"),
            db_path=db_path,
        )

    return item_id


def _link_acceptance_clause(
    acceptance_item_id: int,
    requirement_id: int,
    role: str,
    *,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    role = (role or "").strip()
    if role not in ("requirement", "method_internal"):
        raise ValueError(f"role недопустим: {role}")
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO acceptance_item_clauses
                (acceptance_item_id, requirement_id, role)
            VALUES (?, ?, ?)
            ON CONFLICT(acceptance_item_id, requirement_id, role) DO NOTHING
            """,
            (acceptance_item_id, requirement_id, role),
        )


def add_method_external_ref(
    acceptance_item_id: int,
    *,
    ext_doc_id: str,
    ext_doc_title: str | None = None,
    ext_clause_or_method: str | None = None,
    note: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    ext_doc_id = (ext_doc_id or "").strip()
    if not ext_doc_id:
        raise ValueError("ext_doc_id пустой")
    # Пустая строка вместо NULL — чтобы UNIQUE/ON CONFLICT стабильно срабатывали
    method = (ext_clause_or_method or "").strip()
    with get_connection(db_path) as conn:
        cur = conn.execute(
            """
            INSERT INTO method_external_refs (
                acceptance_item_id, ext_doc_id, ext_doc_title,
                ext_clause_or_method, note
            ) VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(acceptance_item_id, ext_doc_id, ext_clause_or_method)
            DO UPDATE SET
                ext_doc_title = COALESCE(
                    excluded.ext_doc_title, method_external_refs.ext_doc_title
                ),
                note = COALESCE(excluded.note, method_external_refs.note)
            """,
            (
                acceptance_item_id,
                ext_doc_id,
                (ext_doc_title or "").strip() or None,
                method,
                note,
            ),
        )
        if cur.lastrowid:
            return int(cur.lastrowid)
        row = conn.execute(
            """
            SELECT id FROM method_external_refs
            WHERE acceptance_item_id = ? AND ext_doc_id = ?
              AND IFNULL(ext_clause_or_method, '') = ?
            """,
            (acceptance_item_id, ext_doc_id, method),
        ).fetchone()
        return int(row["id"]) if row else 0


def list_acceptance_items(
    *,
    norm_document_id: int | None = None,
    doc_id: str | None = None,
    billable: bool | None = None,
    limit: int = 500,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    """Список acceptance_items с JOIN на norm_documents."""
    q = """
        SELECT a.*, n.doc_id, n.title AS doc_title, n.kind AS doc_kind,
               n.status AS doc_status, n.source_format
        FROM acceptance_items a
        JOIN norm_documents n ON n.id = a.norm_document_id
        WHERE 1=1
    """
    params: list[Any] = []
    if norm_document_id is not None:
        q += " AND a.norm_document_id = ?"
        params.append(norm_document_id)
    if doc_id:
        q += " AND n.doc_id = ?"
        params.append(doc_id.strip())
    if billable is not None:
        q += " AND a.billable = ?"
        params.append(1 if billable else 0)
    q += " ORDER BY n.doc_id, a.sort_order, a.id LIMIT ?"
    params.append(limit)
    with get_connection(db_path) as conn:
        return [dict(r) for r in conn.execute(q, params).fetchall()]


def get_acceptance_item(
    item_id: int,
    *,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    """Один item + clauses + external refs (для show CLI)."""
    with get_connection(db_path) as conn:
        row = conn.execute(
            """
            SELECT a.*, n.doc_id, n.title AS doc_title, n.kind AS doc_kind,
                   n.edition_note, n.source_format, n.manufacturer_hint,
                   n.status AS doc_status, n.file_path AS doc_file_path
            FROM acceptance_items a
            JOIN norm_documents n ON n.id = a.norm_document_id
            WHERE a.id = ?
            """,
            (item_id,),
        ).fetchone()
        if not row:
            return None
        out = dict(row)
        clauses = conn.execute(
            """
            SELECT c.role, r.id AS requirement_id, r.clause, r.title,
                   r.body, r.clause_kind
            FROM acceptance_item_clauses c
            JOIN requirements r ON r.id = c.requirement_id
            WHERE c.acceptance_item_id = ?
            ORDER BY c.role, r.clause
            """,
            (item_id,),
        ).fetchall()
        out["clauses"] = [dict(c) for c in clauses]
        exts = conn.execute(
            """
            SELECT * FROM method_external_refs
            WHERE acceptance_item_id = ?
            ORDER BY id
            """,
            (item_id,),
        ).fetchall()
        out["method_external"] = [dict(e) for e in exts]
        if out.get("regime_json"):
            try:
                out["regime"] = json.loads(out["regime_json"])
            except (TypeError, json.JSONDecodeError):
                out["regime"] = None
        return out


def show_norm_catalog(
    *,
    doc_id: str | None = None,
    norm_document_id: int | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    """Карточка ТУ: документ + все acceptance_items с краткими clause."""
    with get_connection(db_path) as conn:
        if norm_document_id is not None:
            doc = conn.execute(
                "SELECT * FROM norm_documents WHERE id = ?",
                (norm_document_id,),
            ).fetchone()
        elif doc_id:
            doc = conn.execute(
                "SELECT * FROM norm_documents WHERE doc_id = ?",
                (doc_id.strip(),),
            ).fetchone()
        else:
            raise ValueError("нужен doc_id или norm_document_id")
        if not doc:
            return None
        d = dict(doc)
        items = list_acceptance_items(
            norm_document_id=int(d["id"]),
            db_path=db_path,
            limit=2000,
        )
        enriched: list[dict[str, Any]] = []
        for it in items:
            full = get_acceptance_item(int(it["id"]), db_path=db_path)
            if full:
                enriched.append(full)
        d["acceptance_items"] = enriched
        return d


def seed_example_acceptance_catalog(db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """
    Идемпотентный seed эталона 131 (структура таблицы приёмки, без текста ТУ из файла).

    group_code не заполняем (решение v3). Маркировка — billable=0.
    """
    doc_key = "ТУ 27.31.11-131-47273194-2025"
    nd_id = upsert_norm_document(
        doc_key,
        "Кабели оптические огнестойкие (эталон каталога acceptance, seed)",
        kind="tu",
        notes="seed wave1; полный импорт docx — волна 2; файл ТУ только локально",
        source_format="docx_clean",
        manufacturer_hint="ООО НПП Спецкабель",
        status="draft",
        edition_note="seed structure from TZ v3 §9.1",
        db_path=db_path,
    )
    # Уже есть items? не дублируем сверх seed
    existing = list_acceptance_items(norm_document_id=nd_id, db_path=db_path)
    if len(existing) >= 3:
        return 0

    specs: list[dict[str, Any]] = [
        {
            "name_exact": "Прочность к растягивающему усилию",
            "test_category": "periodic",
            "group_code": None,
            "requirement_clauses": ["2.5.1"],
            "method_clauses": ["5.4.1"],
            "billable": True,
            "sort_order": 10,
            "regime": {
                "force_kn": 1.5,
                "sample_length_m": 10,
                "hold_min": 10,
                "source": "seed_tz_v3_9_1",
            },
            "method_external": [
                {
                    "ext_doc_id": "ГОСТ 12182.5-80",
                    "ext_clause_or_method": None,
                    "note": "seed; + контроль целостности — отдельно при review",
                }
            ],
            "notes": "seed: эталон строки периодики",
        },
        {
            "name_exact": "Измерение коэффициента затухания",
            "test_category": "psi",
            "requirement_clauses": ["2.4"],
            "method_clauses": ["5.3"],
            "billable": True,
            "sort_order": 20,
            "method_external": [
                {
                    "ext_doc_id": "ГОСТ Р МЭК 60793-1-40",
                    "ext_clause_or_method": "метод C",
                }
            ],
            "notes": "seed: ПСИ, пункты по одному",
        },
        {
            "name_exact": "Проверка маркировки и упаковки",
            "test_category": "psi",
            "requirement_clauses": ["2.8", "2.9"],
            "method_clauses": ["5.7.1"],
            "billable": False,
            "sort_order": 30,
            "notes": "seed: billable=false (решение v3, не прайс)",
        },
    ]
    added = 0
    for spec in specs:
        name = str(spec["name_exact"])
        if any(_norm_name(e["name_exact"]) == _norm_name(name) for e in existing):
            continue
        add_acceptance_item(
            norm_document_id=nd_id,
            name_exact=name,
            requirement_clauses=list(spec.get("requirement_clauses") or []),
            method_clauses=list(spec.get("method_clauses") or []),
            test_category=spec.get("test_category"),
            group_code=spec.get("group_code"),
            billable=bool(spec.get("billable", True)),
            sort_order=int(spec.get("sort_order") or 0),
            regime=spec.get("regime"),
            notes=spec.get("notes"),
            method_external=list(spec.get("method_external") or []),
            status="draft",
            db_path=db_path,
        )
        added += 1
    return added


def list_norm_documents(
    *,
    kind: str | None = None,
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    q = "SELECT * FROM norm_documents"
    params: list[Any] = []
    if kind:
        q += " WHERE kind = ?"
        params.append(kind)
    q += " ORDER BY kind, doc_id LIMIT ?"
    params.append(limit)
    with get_connection(db_path) as conn:
        return [dict(r) for r in conn.execute(q, params).fetchall()]


def list_requirements(
    *,
    norm_document_id: int | None = None,
    limit: int = 200,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    q = """
        SELECT r.*, n.doc_id, n.title AS doc_title, n.kind
        FROM requirements r
        JOIN norm_documents n ON n.id = r.norm_document_id
    """
    params: list[Any] = []
    if norm_document_id is not None:
        q += " WHERE r.norm_document_id = ?"
        params.append(norm_document_id)
    q += " ORDER BY n.doc_id, r.clause LIMIT ?"
    params.append(limit)
    with get_connection(db_path) as conn:
        return [dict(r) for r in conn.execute(q, params).fetchall()]


def list_test_aliases(
    *,
    limit: int = 200,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    with get_connection(db_path) as conn:
        rows = conn.execute(
            "SELECT * FROM test_aliases ORDER BY alias_norm LIMIT ?",
            (limit,),
        ).fetchall()
        return [dict(r) for r in rows]


def add_test_alias(
    alias: str,
    canonical_name: str,
    *,
    price_test_code: str | None = None,
    source: str = "manual",
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    now = datetime.now().isoformat()
    alias_norm = alias.strip().lower()
    with get_connection(db_path) as conn:
        cur = conn.execute(
            """
            INSERT INTO test_aliases
                (alias_norm, canonical_name, price_test_code, source, created_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(alias_norm) DO UPDATE SET
                canonical_name = excluded.canonical_name,
                price_test_code = COALESCE(excluded.price_test_code, test_aliases.price_test_code)
            """,
            (
                alias_norm,
                canonical_name.strip(),
                (price_test_code or "").strip() or None,
                source,
                now,
            ),
        )
        if cur.lastrowid:
            return int(cur.lastrowid)
        row = conn.execute(
            "SELECT id FROM test_aliases WHERE alias_norm = ?", (alias_norm,)
        ).fetchone()
        return int(row["id"]) if row else 0


def resolve_test_alias(
    phrase: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    """Ищет alias по точному совпадению или вхождению."""
    p = (phrase or "").strip().lower()
    if not p:
        return None
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM test_aliases WHERE alias_norm = ?", (p,)
        ).fetchone()
        if row:
            return dict(row)
        rows = conn.execute("SELECT * FROM test_aliases").fetchall()
        for r in rows:
            a = r["alias_norm"]
            if a and (a in p or p in a) and len(a) >= 4:
                return dict(r)
    return None


def _migrate_orders_columns(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    with get_connection(db_path) as conn:
        cols = {row[1] for row in conn.execute("PRAGMA table_info(orders)").fetchall()}
        if "application_path" not in cols:
            conn.execute("ALTER TABLE orders ADD COLUMN application_path TEXT")


def _migrate_organizations_columns(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    from ..extraction.organization_extractor import finalize_organization_address, sanitize_address
    from ..models import OrganizationExtract

    with get_connection(db_path) as conn:
        cols = {row[1] for row in conn.execute("PRAGMA table_info(organizations)").fetchall()}
        if "legal_address" not in cols:
            conn.execute("ALTER TABLE organizations ADD COLUMN legal_address TEXT")
        if "actual_address" not in cols:
            conn.execute("ALTER TABLE organizations ADD COLUMN actual_address TEXT")

        rows = conn.execute(
            "SELECT id, address, legal_address, actual_address FROM organizations"
        ).fetchall()
        for row in rows:
            raw = row["address"]
            legal = row["legal_address"]
            actual = row["actual_address"]
            org_row = conn.execute(
                "SELECT name FROM organizations WHERE id = ?", (row["id"],)
            ).fetchone()
            org_name = org_row["name"] if org_row else ""
            finalized = finalize_organization_address(
                OrganizationExtract(
                    name=org_name or "—",
                    address=raw,
                    legal_address=legal,
                    actual_address=actual,
                ),
                str(raw or ""),
            )
            clean = finalized.address or sanitize_address(raw)
            clean_legal = finalized.legal_address or sanitize_address(legal) if legal else clean
            clean_actual = finalized.actual_address or sanitize_address(actual) if actual else None
            if clean and (not legal or len(str(legal)) > 250):
                legal = clean_legal
            if not actual and clean_actual:
                actual = clean_actual
            if raw != clean or legal != row["legal_address"] or actual != row["actual_address"]:
                conn.execute(
                    """
                    UPDATE organizations SET
                        address = ?,
                        legal_address = COALESCE(?, legal_address),
                        actual_address = COALESCE(?, actual_address)
                    WHERE id = ?
                    """,
                    (clean or raw, legal, actual, row["id"]),
                )


def get_calculation_lines(
    calculation_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    with get_connection(db_path) as conn:
        rows = conn.execute(
            """
            SELECT test_name, base_cost, multiplier, quantity, hours, final_cost, note
            FROM calculation_lines
            WHERE calculation_id = ?
            ORDER BY id
            """,
            (calculation_id,),
        ).fetchall()
        return [dict(row) for row in rows]


def get_cable_mark_document(
    full_mark: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> str | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT document FROM cable_marks WHERE full_mark = ? LIMIT 1",
            (full_mark,),
        ).fetchone()
        return row["document"] if row and row["document"] else None


def update_order_application_path(
    order_id: int,
    application_path: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            "UPDATE orders SET application_path = ?, updated_at = ? WHERE id = ?",
            (application_path, datetime.now().isoformat(), order_id),
        )


def save_test_application(
    order_id: int,
    output_path: str,
    *,
    template_path: str | None = None,
    test_type: str | None = None,
    customer_name: str | None = None,
    manufacturer_name: str | None = None,
    marks_snapshot: list[dict[str, Any]] | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Сохраняет сформированную заявку на испытания в БД и обновляет orders.application_path."""
    import json

    now = datetime.now().isoformat()
    snapshot = marks_snapshot or []
    with get_connection(db_path) as conn:
        cur = conn.execute(
            """
            INSERT INTO test_applications (
                order_id, output_path, template_path, application_number,
                test_type, customer_name, manufacturer_name,
                marks_count, marks_snapshot, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                order_id,
                output_path,
                template_path,
                str(order_id),
                test_type,
                customer_name,
                manufacturer_name,
                len(snapshot),
                json.dumps(snapshot, ensure_ascii=False),
                now,
            ),
        )
        conn.execute(
            "UPDATE orders SET application_path = ?, updated_at = ? WHERE id = ?",
            (output_path, now, order_id),
        )
        app_id = int(cur.lastrowid)
    save_generated_document(
        doc_type="application",
        file_path=output_path,
        order_id=order_id,
        db_path=db_path,
    )
    return app_id


def list_test_applications(
    order_id: int | None = None,
    limit: int = 50,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    """История сформированных заявок на испытания."""
    import json

    query = "SELECT * FROM test_applications"
    params: list[Any] = []
    if order_id is not None:
        query += " WHERE order_id = ?"
        params.append(order_id)
    query += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)

    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        result: list[dict[str, Any]] = []
        for row in rows:
            item = dict(row)
            raw = item.get("marks_snapshot")
            if raw:
                try:
                    item["marks_snapshot"] = json.loads(raw)
                except json.JSONDecodeError:
                    item["marks_snapshot"] = []
            else:
                item["marks_snapshot"] = []
            result.append(item)
        return result


def sync_test_rule_types(db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Пересчитывает rule_type по названию/категории (per_core, per_group, time_based)."""
    updated = 0
    with get_connection(db_path) as conn:
        rows = conn.execute(
            "SELECT id, code, name, category FROM test_items"
        ).fetchall()
        for row in rows:
            rule_type, rule_params = infer_rule_type(
                row["name"], row["category"], row["code"]
            )
            conn.execute(
                """
                UPDATE test_items
                SET rule_type = ?, rule_params = ?
                WHERE id = ?
                """,
                (
                    rule_type,
                    json.dumps(rule_params, ensure_ascii=False),
                    row["id"],
                ),
            )
            updated += 1
    return updated


def _migrate_calculation_lines_quantity(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    with get_connection(db_path) as conn:
        cols = {row[1] for row in conn.execute("PRAGMA table_info(calculation_lines)").fetchall()}
        if "quantity" not in cols:
            conn.execute(
                "ALTER TABLE calculation_lines ADD COLUMN quantity INTEGER NOT NULL DEFAULT 1"
            )


_PRICE_NAME_FIXES: dict[str, str] = {
    "измерение_толщины_облочкишланга": "Измерение толщины оболочки/шланга",
    "измерение_шагакратности_скртки": "Измерение шага/кратности скрутки",
    "водопоглащение": "Водопоглощение",
    "прочность_и_относительное_удлинение_элемента_конструкци": (
        "Прочность и относительное удлинение элемента конструкции"
    ),
    "хромотография_определние_фтора_в_тч": "Хроматография (определение фтора в т.ч.)",
    "установка_соединителейпод_vna_aesa": "Установка соединителей (под VNA, AESA)",
}

_OPTICAL_PER_CORE_CODES = (
    "измерение_затухания_оптического_волокнаодного",
    "определение_целостности_оптического_волокнаодного",
)


def apply_price_catalog_fixes(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """
    Синхронизация прайса по шаблону Obsidian §39:
    убрать EN-дубли климатики, исправить опечатки, оптика → per_core.
    """
    with get_connection(db_path) as conn:
        for en_code, slug in CLIMATE_ITEM_ALIASES.items():
            conn.execute(
                "UPDATE test_mappings SET test_code = ? WHERE test_code = ?",
                (slug, en_code),
            )
        for code in DEPRECATED_CLIMATE_ITEM_CODES:
            conn.execute("DELETE FROM test_items WHERE code = ?", (code,))

    for code, name in _PRICE_NAME_FIXES.items():
        if get_test_item_by_code(code, db_path):
            update_test_item(code, TestItemUpdate(name=name), db_path)

    for code in _OPTICAL_PER_CORE_CODES:
        if get_test_item_by_code(code, db_path):
            update_test_item(
                code,
                TestItemUpdate(rule_type="per_core", rule_params={}),
                db_path,
            )


def sync_climatic_tests(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """Обновляет slug-коды климатики: time_based + rule_params (без EN-дублей)."""
    for spec in CLIMATIC_TESTS:
        slug = CLIMATE_SLUG_BY_HOURS_KEY.get(spec["hours_key"])
        if not slug or not get_test_item_by_code(slug, db_path):
            continue
        update_test_item(
            slug,
            TestItemUpdate(
                rule_type="time_based",
                rule_params={
                    "hours_key": spec["hours_key"],
                    "default_hours": spec["default_hours"],
                    "cost_per_hour": spec["cost_per_hour"],
                },
            ),
            db_path,
        )


_DEFAULT_TEST_MAPPINGS: list[tuple[str, str, str | None]] = [
    # Климатика (slug-коды прайса)
    ("воздействию солнечного", "стойкость_к_солнечной_радиации", "Направления в ИЛ"),
    ("солнечного излучения", "стойкость_к_солнечной_радиации", "ГОСТ 20.57.406"),
    ("солнечной радиации", "стойкость_к_солнечной_радиации", "Климатика"),
    ("20.57.406", "стойкость_к_солнечной_радиации", "ГОСТ солнечного излучения"),
    ("метод 211-1", "стойкость_к_солнечной_радиации", "ГОСТ 20.57.406 метод 211-1"),
    ("повышенной влажности", "стойкость_к_повышенной_влажности_воздуха", "Климатика"),
    ("влажности воздуха", "стойкость_к_повышенной_влажности_воздуха", "Климатика"),
    ("пониженной температуры", "стойкость_к_пониженной_температуре", "Климатика"),
    ("пониженной температуре", "стойкость_к_пониженной_температуре", "Климатика"),
    ("повышенной температуры", "стойкость_к_повышенной_температуре", "Климатика"),
    ("повышенной температуре", "стойкость_к_повышенной_температуре", "Климатика"),
    ("изменению температур", "стойкость_к_изменению_температуррезкоеплавное", "Климатика"),
    ("изменению температуры", "стойкость_к_изменению_температуррезкоеплавное", "Климатика"),
    ("циклическ", "стойкость_к_изменению_температуррезкоеплавное", "Климатика"),
    ("отрицательной температур", "стойкость_к_пониженной_температуре", "Морозостойкость"),
    ("простому изгибу", "стойкость_к_простому_изгибу_100_циклов", "Механика"),
    # Электрические (коды из прайса)
    ("электрическое сопротивление тпж", "электрическое_сопротивление_тпж", "Прайс НЧ"),
    ("сопротивление изоляции", "электрическое_сопротивление_изоляции_тпж", "Прайс НЧ"),
    ("сопротивление изоляции тпж", "электрическое_сопротивление_изоляции_тпж", "Прайс НЧ"),
    ("испытание напряжением", "испытание_напряжением", "Прайс НЧ"),
    ("испытание напряжение", "испытание_напряжением", "Прайс НЧ"),
    ("емкости", "измерение_емкостииндуктивности", "ВЧ параметры"),
    ("индуктивности", "измерение_емкостииндуктивности", "ВЧ параметры"),
    ("затухания экранирования", "измерение_затухания_экранирования", "ВЧ"),
    ("затухания излучения", "измерение_затухания_излучения", "ВЧ"),
    ("огнестойкость", "огнестойкость", "Пожарная безопасность"),
    ("ультрафиолет", "стойкость_к_солнечной_радиации", "УФ → солнечная радиация"),
]


def sync_default_test_mappings(db_path: str | Path = DB_PATH_DEFAULT) -> None:
    """Заполняет test_mappings базовыми фразами (идемпотентно)."""
    now = datetime.now().isoformat()
    with get_connection(db_path) as conn:
        for pattern, test_code, note in _DEFAULT_TEST_MAPPINGS:
            conn.execute(
                """
                INSERT INTO test_mappings (requirement_pattern, test_code, note, usage_count, created_at)
                VALUES (?, ?, ?, 0, ?)
                ON CONFLICT(requirement_pattern) DO NOTHING
                """,
                (pattern.lower(), test_code, note, now),
            )


def sync_mappings_from_test_item_names(db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Добавляет маппинги «название из прайса → code» (для явных перечней в письмах).

    Идемпотентно. Возвращает число **новых** строк.
    """
    now = datetime.now().isoformat()
    added = 0
    with get_connection(db_path) as conn:
        rows = conn.execute("SELECT code, name FROM test_items").fetchall()
        for row in rows:
            code = (row["code"] or "").strip()
            name = (row["name"] or "").strip()
            if not code or not name or len(name) < 4:
                continue
            pattern = name.lower()
            cur = conn.execute(
                """
                INSERT INTO test_mappings (requirement_pattern, test_code, note, usage_count, created_at)
                VALUES (?, ?, ?, 0, ?)
                ON CONFLICT(requirement_pattern) DO NOTHING
                """,
                (pattern, code, "auto: имя из прайса", now),
            )
            if cur.rowcount:
                added += 1
            # короткий паттерн без «определение/измерение/проверка»
            short = re.sub(
                r"^(определение|измерение|проверка|испытание)\s+",
                "",
                pattern,
                flags=re.IGNORECASE,
            ).strip()
            if short and short != pattern and len(short) >= 6:
                cur2 = conn.execute(
                    """
                    INSERT INTO test_mappings (requirement_pattern, test_code, note, usage_count, created_at)
                    VALUES (?, ?, ?, 0, ?)
                    ON CONFLICT(requirement_pattern) DO NOTHING
                    """,
                    (short, code, "auto: короткое имя прайса", now),
                )
                if cur2.rowcount:
                    added += 1
    return added


# --- Программы испытаний (S4) ---


def create_test_program(
    *,
    name: str,
    test_type: str | None = None,
    cable_mark_text: str | None = None,
    tu_ref: str | None = None,
    source_path: str | None = None,
    notes: str | None = None,
    items: list[dict[str, Any]] | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Создаёт программу и позиции. items: dict с name, requirement_*, method_*, price_test_code."""
    now = datetime.now().isoformat()
    with get_connection(db_path) as conn:
        cur = conn.execute(
            """
            INSERT INTO test_programs
                (name, test_type, cable_mark_text, tu_ref, source_path, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name.strip(),
                (test_type or "").strip() or None,
                (cable_mark_text or "").strip() or None,
                (tu_ref or "").strip() or None,
                source_path,
                notes,
                now,
                now,
            ),
        )
        program_id = int(cur.lastrowid)
        for i, it in enumerate(items or []):
            conn.execute(
                """
                INSERT INTO test_program_items
                    (program_id, sort_order, name, requirement_doc, requirement_clause,
                     method_doc, method_clause, price_test_code, meta_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    program_id,
                    int(it.get("sort_order", i + 1)),
                    str(it.get("name") or "").strip() or f"Пункт {i + 1}",
                    it.get("requirement_doc"),
                    it.get("requirement_clause"),
                    it.get("method_doc"),
                    it.get("method_clause"),
                    it.get("price_test_code"),
                    json.dumps(it.get("meta") or {}, ensure_ascii=False)
                    if it.get("meta")
                    else None,
                ),
            )
        return program_id


def list_test_programs(
    *,
    search: str | None = None,
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    query = """
        SELECT p.*,
               (SELECT COUNT(*) FROM test_program_items i WHERE i.program_id = p.id) AS items_count
        FROM test_programs p
    """
    params: list[Any] = []
    if search:
        query += " WHERE p.name LIKE ? OR IFNULL(p.cable_mark_text,'') LIKE ? OR IFNULL(p.tu_ref,'') LIKE ?"
        params.extend([f"%{search}%"] * 3)
    query += " ORDER BY p.updated_at DESC LIMIT ?"
    params.append(limit)
    with get_connection(db_path) as conn:
        return [dict(r) for r in conn.execute(query, params).fetchall()]


def get_test_program(
    program_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM test_programs WHERE id = ?", (program_id,)
        ).fetchone()
        if not row:
            return None
        items = conn.execute(
            """
            SELECT * FROM test_program_items
            WHERE program_id = ?
            ORDER BY sort_order, id
            """,
            (program_id,),
        ).fetchall()
        data = dict(row)
        data["items"] = [dict(i) for i in items]
        return data


def delete_test_program(
    program_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> bool:
    with get_connection(db_path) as conn:
        conn.execute("DELETE FROM test_program_items WHERE program_id = ?", (program_id,))
        cur = conn.execute("DELETE FROM test_programs WHERE id = ?", (program_id,))
        return cur.rowcount > 0


def update_program_item_price_code(
    item_id: int,
    price_test_code: str | None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> bool:
    with get_connection(db_path) as conn:
        cur = conn.execute(
            "UPDATE test_program_items SET price_test_code = ? WHERE id = ?",
            ((price_test_code or "").strip() or None, item_id),
        )
        return cur.rowcount > 0


def match_program_items_to_price(
    program_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
    *,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Проставляет price_test_code (S4 polish: phrase rules + fuzzy + aliases).

    Returns: {matched, unmatched, total, rate, summary, details}
    rate = matched/total в [0..1]; summary — «сопоставлено N/M (xx%)».
    overwrite=True — перезаписать уже заданные (и ошибочные) коды.
    """
    from ..mapping.program_price_matcher import (
        match_rate_summary,
        resolve_program_item_price_code,
    )

    prog = get_test_program(program_id, db_path=db_path)
    if not prog:
        return {
            "matched": 0,
            "unmatched": 0,
            "total": 0,
            "rate": 0.0,
            "summary": match_rate_summary(0, 0),
            "details": [],
        }
    price_items = list_test_items(limit=500, db_path=db_path)
    context = " ".join(
        filter(
            None,
            [
                prog.get("name") or "",
                prog.get("cable_mark_text") or "",
                prog.get("tu_ref") or "",
                prog.get("notes") or "",
            ],
        )
    )
    matched = 0
    unmatched = 0
    details: list[dict[str, Any]] = []
    for item in prog["items"]:
        existing = (item.get("price_test_code") or "").strip() or None
        name = (item.get("name") or "").strip()
        if existing and not overwrite:
            matched += 1
            details.append(
                {
                    "id": item.get("id"),
                    "name": name,
                    "code": existing,
                    "method": "kept",
                }
            )
            continue
        hit = resolve_program_item_price_code(
            name,
            price_items=price_items,
            db_path=db_path,
            program_context=context,
        )
        if hit:
            update_program_item_price_code(int(item["id"]), hit.code, db_path=db_path)
            matched += 1
            details.append(
                {
                    "id": item.get("id"),
                    "name": name,
                    "code": hit.code,
                    "method": hit.method,
                    "note": hit.note,
                    "score": hit.score,
                }
            )
        else:
            if existing and overwrite:
                update_program_item_price_code(int(item["id"]), None, db_path=db_path)
            unmatched += 1
            details.append(
                {
                    "id": item.get("id"),
                    "name": name,
                    "code": None,
                    "method": "none",
                }
            )
    total = matched + unmatched
    rate = (matched / total) if total else 0.0
    return {
        "matched": matched,
        "unmatched": unmatched,
        "total": total,
        "rate": rate,
        "summary": match_rate_summary(matched, total),
        "details": details,
    }


def list_test_mappings(
    *,
    test_code: str | None = None,
    limit: int = 200,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    query = "SELECT * FROM test_mappings"
    params: list[Any] = []
    if test_code:
        query += " WHERE test_code = ?"
        params.append(test_code)
    query += " ORDER BY usage_count DESC, requirement_pattern LIMIT ?"
    params.append(limit)
    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def add_test_mapping(
    requirement_pattern: str,
    test_code: str,
    *,
    note: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Добавляет или обновляет маппинг «фраза → код испытания»."""
    pattern = requirement_pattern.strip().lower()
    if not pattern:
        raise ValueError("Пустой шаблон требования")
    code = test_code.strip()
    if not code:
        raise ValueError("Пустой код испытания")
    now = datetime.now().isoformat()
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO test_mappings (requirement_pattern, test_code, note, usage_count, created_at)
            VALUES (?, ?, ?, 0, ?)
            ON CONFLICT(requirement_pattern) DO UPDATE SET
                test_code = excluded.test_code,
                note = COALESCE(excluded.note, test_mappings.note)
            """,
            (pattern, code, note, now),
        )
        row = conn.execute(
            "SELECT id FROM test_mappings WHERE requirement_pattern = ?",
            (pattern,),
        ).fetchone()
        return int(row["id"]) if row else 0


def get_test_mapping(
    mapping_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM test_mappings WHERE id = ?",
            (mapping_id,),
        ).fetchone()
        return dict(row) if row else None


def update_test_mapping(
    mapping_id: int,
    *,
    requirement_pattern: str | None = None,
    test_code: str | None = None,
    note: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    """Обновляет запись маппинга по id."""
    current = get_test_mapping(mapping_id, db_path)
    if not current:
        raise ValueError(f"Маппинг id={mapping_id} не найден")
    pattern = (requirement_pattern or current["requirement_pattern"]).strip().lower()
    code = (test_code or current["test_code"]).strip()
    if not pattern:
        raise ValueError("Пустой шаблон требования")
    if not code:
        raise ValueError("Пустой код испытания")
    new_note = note if note is not None else current.get("note")
    with get_connection(db_path) as conn:
        conn.execute(
            """
            UPDATE test_mappings
            SET requirement_pattern = ?, test_code = ?, note = ?
            WHERE id = ?
            """,
            (pattern, code, new_note, mapping_id),
        )


def delete_test_mapping(
    mapping_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> bool:
    """Удаляет маппинг. Возвращает False, если запись не найдена."""
    with get_connection(db_path) as conn:
        cur = conn.execute("DELETE FROM test_mappings WHERE id = ?", (mapping_id,))
        return cur.rowcount > 0


def save_generated_document(
    *,
    doc_type: str,
    file_path: str,
    order_id: int | None = None,
    calculation_id: int | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Журнал сгенерированных файлов (КП, заявка на испытания)."""
    if doc_type not in ("kp", "application"):
        raise ValueError(f"Неизвестный тип документа: {doc_type}")
    now = datetime.now().isoformat()
    resolved = str(Path(file_path).resolve())
    with get_connection(db_path) as conn:
        cur = conn.execute(
            """
            INSERT INTO generated_documents (
                order_id, calculation_id, doc_type, file_path, created_at
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (order_id, calculation_id, doc_type, resolved, now),
        )
        return int(cur.lastrowid or 0)


def list_generated_documents(
    *,
    order_id: int | None = None,
    doc_type: str | None = None,
    limit: int = 50,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    query = "SELECT * FROM generated_documents"
    params: list[Any] = []
    conditions: list[str] = []
    if order_id is not None:
        conditions.append("order_id = ?")
        params.append(order_id)
    if doc_type:
        conditions.append("doc_type = ?")
        params.append(doc_type)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)
    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def record_mapping_usage(
    mapping_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    """Увеличивает счётчик срабатывания маппинга (для обучения на подтверждениях оператора)."""
    with get_connection(db_path) as conn:
        conn.execute(
            "UPDATE test_mappings SET usage_count = usage_count + 1 WHERE id = ?",
            (mapping_id,),
        )


CLIMATIC_SETTINGS_KEY = "climatic_test_hours"
DOCUMENT_PACK_SETTINGS_KEY = "document_pack_settings"
ASSISTANT_LLM_SETTINGS_KEY = "assistant_llm_settings"


def _seed_default_settings(db_path: str | Path) -> None:
    if get_climatic_settings(db_path) is not None:
        return
    save_climatic_settings(ClimaticTestSettings(), db_path)


def get_climatic_settings(db_path: str | Path = DB_PATH_DEFAULT) -> ClimaticTestSettings | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT value FROM app_settings WHERE key = ?", (CLIMATIC_SETTINGS_KEY,)
        ).fetchone()
        if not row:
            return None
        return ClimaticTestSettings(**json.loads(row["value"]))


def save_climatic_settings(
    settings: ClimaticTestSettings,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO app_settings (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (CLIMATIC_SETTINGS_KEY, settings.model_dump_json()),
        )


def get_document_pack_settings(
    db_path: str | Path = DB_PATH_DEFAULT,
) -> DocumentPackSettings:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT value FROM app_settings WHERE key = ?", (DOCUMENT_PACK_SETTINGS_KEY,)
        ).fetchone()
        if not row:
            return DocumentPackSettings()
        return DocumentPackSettings(**json.loads(row["value"]))


def save_document_pack_settings(
    settings: DocumentPackSettings,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO app_settings (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (DOCUMENT_PACK_SETTINGS_KEY, settings.model_dump_json()),
        )


def get_assistant_llm_settings(
    db_path: str | Path = DB_PATH_DEFAULT,
) -> AssistantLlmSettings:
    from ..assistant.llm_provider import default_llm_settings, resolve_llm_settings

    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT value FROM app_settings WHERE key = ?", (ASSISTANT_LLM_SETTINGS_KEY,)
        ).fetchone()
        if not row:
            return resolve_llm_settings(default_llm_settings())
        return resolve_llm_settings(AssistantLlmSettings(**json.loads(row["value"])))


def save_assistant_llm_settings(
    settings: AssistantLlmSettings,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            """
            INSERT INTO app_settings (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (ASSISTANT_LLM_SETTINGS_KEY, settings.model_dump_json()),
        )


def prepare_prod_db(
    db_path: str | Path = DB_PATH_DEFAULT,
    *,
    backup: bool = True,
) -> dict[str, Any]:
    """
    Готовит БД к prod (рабочий ПК) с «чистыми» марками и организациями.

    **Сохраняет:**
    - ``test_items`` — прайс и правила расчёта (fixed/per_core/…)
    - ``test_mappings`` — фразы требований → коды испытаний
    - ``app_settings`` — климатика, LLM, пути пакетов, station_id
    - training/RAG-таблицы (если есть)

    **Очищает:**
    - ``cable_marks``, ``organizations``
    - связанные операционные данные: заказы, расчёты, извлечения,
      заявки, generated_documents, assistant_sessions
      (иначе остаются «осиротевшие» ссылки на удалённые org/mark)

    Returns:
        dict с путём backup (если был), счётчиками удалённых строк
        и сохранёнными test_items / test_mappings.
    """
    path = Path(db_path)
    if not path.exists():
        raise FileNotFoundError(f"БД не найдена: {path}")

    backup_path: Path | None = None
    if backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = path.with_name(f"{path.stem}.pre_prod_{stamp}{path.suffix}")
        import shutil

        shutil.copy2(path, backup_path)

    # Порядок: дочерние таблицы → родители (на случай включённых FK)
    clear_tables = (
        "calculation_lines",
        "order_marks",
        "test_applications",
        "generated_documents",
        "orders",
        "calculations",
        "document_extractions",
        "assistant_sessions",
        "cable_marks",
        "organizations",
    )

    deleted: dict[str, int] = {}
    with get_connection(path) as conn:
        conn.execute("PRAGMA foreign_keys = OFF")
        for table in clear_tables:
            exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (table,),
            ).fetchone()
            if not exists:
                continue
            count = conn.execute(f"SELECT COUNT(*) AS n FROM {table}").fetchone()["n"]
            conn.execute(f"DELETE FROM {table}")
            deleted[table] = int(count)
        conn.execute("PRAGMA foreign_keys = ON")
        kept_items = conn.execute("SELECT COUNT(*) AS n FROM test_items").fetchone()["n"]
        kept_maps = conn.execute("SELECT COUNT(*) AS n FROM test_mappings").fetchone()["n"]

    return {
        "db_path": str(path.resolve()),
        "backup_path": str(backup_path.resolve()) if backup_path else None,
        "deleted": deleted,
        "kept_test_items": int(kept_items),
        "kept_test_mappings": int(kept_maps),
    }


def push_recent_pack_path(
    pack_dir: str | Path,
    db_path: str | Path = DB_PATH_DEFAULT,
    *,
    limit: int = 5,
) -> None:
    """Добавляет путь пакета в историю (последние limit уникальных)."""
    path = str(Path(pack_dir).resolve())
    settings = get_document_pack_settings(db_path)
    recent = [p for p in settings.recent_paths if p != path]
    recent.insert(0, path)
    settings.recent_paths = recent[:limit]
    save_document_pack_settings(settings, db_path)


def build_default_hours_map(db_path: str | Path = DB_PATH_DEFAULT) -> dict[str, float]:
    """Собирает часы выдержки из настроек + default_hours из time_based испытаний."""
    hours: dict[str, float] = {}
    settings = get_climatic_settings(db_path)
    if settings:
        for key, _ in climatic_settings_fields():
            hours[key] = float(getattr(settings, key))
    for item in get_all_test_items(db_path):
        if item.rule_type != "time_based":
            continue
        key = item.rule_params.get("hours_key", item.code)
        if key not in hours and "default_hours" in item.rule_params:
            hours[key] = float(item.rule_params["default_hours"])
    return hours


def upsert_cable_mark(
    record: CableMarkRecord,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Добавляет марку без полных дублей (по full_mark)."""
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO cable_marks (
                full_mark, brand, fire_class, cores_count,
                structural_element_type, structural_elements_count,
                characteristic_size, size_unit, document, source, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(full_mark) DO UPDATE SET
                brand = excluded.brand,
                fire_class = excluded.fire_class,
                cores_count = excluded.cores_count,
                structural_element_type = excluded.structural_element_type,
                structural_elements_count = excluded.structural_elements_count,
                characteristic_size = excluded.characteristic_size,
                size_unit = excluded.size_unit,
                document = COALESCE(excluded.document, cable_marks.document),
                source = COALESCE(excluded.source, cable_marks.source)
            """,
            (
                record.full_mark,
                record.brand,
                record.fire_class,
                record.cores_count,
                record.structural_element_type,
                record.structural_elements_count,
                record.characteristic_size,
                record.size_unit,
                record.document,
                record.source,
                (record.created_at or datetime.now()).isoformat(),
            ),
        )
        if cursor.lastrowid:
            return cursor.lastrowid
        row = conn.execute(
            "SELECT id FROM cable_marks WHERE full_mark = ?", (record.full_mark,)
        ).fetchone()
        return int(row["id"]) if row else 0


def save_cable_marks_from_matches(
    matches: list,
    *,
    source: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, int]:
    """Сохраняет найденные в PDF марки в накопительную таблицу."""
    stats = {"saved": 0, "errors": 0}
    for match in matches:
        try:
            record = parse_cable_mark_record(
                match.mark,
                document=getattr(match, "document", None),
                context=getattr(match, "context", None),
            )
            record.source = source
            upsert_cable_mark(record, db_path)
            stats["saved"] += 1
        except Exception:
            stats["errors"] += 1
    return stats


def save_cable_marks_from_validations(
    validations: list,
    *,
    source: str | None = None,
    only_accepted: bool = True,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, int]:
    """Сохраняет подтверждённые марки с полями, заданными оператором."""
    from ..models import MarkValidation

    stats = {"saved": 0, "errors": 0}
    for item in validations:
        if only_accepted and isinstance(item, MarkValidation) and not item.accepted:
            continue
        try:
            if isinstance(item, MarkValidation):
                record = item.to_cable_mark_record(source=source)
            else:
                record = parse_cable_mark_record(
                    item.mark,
                    document=getattr(item, "document", None),
                    context=getattr(item, "context", None),
                )
                record.source = source
            upsert_cable_mark(record, db_path)
            stats["saved"] += 1
        except Exception:
            stats["errors"] += 1
    return stats


def upsert_organization(
    extract: OrganizationExtract,
    *,
    source: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Сохраняет организацию без дублей (по ИНН или нормализованному названию)."""
    from ..extraction.organization_extractor import finalize_organization_address

    extract = finalize_organization_address(extract)
    now = datetime.now().isoformat()
    name_normalized = normalize_org_name(extract.name)
    inn_key = extract.inn or ""

    with get_connection(db_path) as conn:
        row = None
        if extract.inn:
            row = conn.execute(
                "SELECT * FROM organizations WHERE inn = ?",
                (extract.inn,),
            ).fetchone()
        if row is None:
            row = conn.execute(
                "SELECT * FROM organizations WHERE name_normalized = ? AND COALESCE(inn, '') = ?",
                (name_normalized, inn_key),
            ).fetchone()

        legal = extract.legal_address or extract.address
        actual = extract.actual_address or legal

        if row:
            org_id = int(row["id"])
            conn.execute(
                """
                UPDATE organizations SET
                    name = ?,
                    address = COALESCE(?, address),
                    legal_address = COALESCE(?, legal_address),
                    actual_address = COALESCE(?, actual_address),
                    postal_code = COALESCE(?, postal_code),
                    phone = COALESCE(?, phone),
                    email = COALESCE(?, email),
                    inn = COALESCE(?, inn),
                    kpp = COALESCE(?, kpp),
                    is_accredited = MAX(is_accredited, ?),
                    fsa_registry_number = COALESCE(?, fsa_registry_number),
                    org_type = CASE WHEN ? = 'unknown' THEN org_type ELSE ? END,
                    source = COALESCE(?, source),
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    extract.name,
                    extract.address,
                    legal,
                    actual,
                    extract.postal_code,
                    extract.phone,
                    extract.email,
                    extract.inn,
                    extract.kpp,
                    int(extract.is_accredited),
                    extract.fsa_registry_number,
                    extract.org_type,
                    extract.org_type,
                    source,
                    now,
                    org_id,
                ),
            )
            return org_id

        cursor = conn.execute(
            """
            INSERT INTO organizations (
                name, name_normalized, address, legal_address, actual_address,
                postal_code, phone, email,
                inn, kpp, is_accredited, fsa_registry_number, org_type,
                source, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                extract.name,
                name_normalized,
                extract.address,
                legal,
                actual,
                extract.postal_code,
                extract.phone,
                extract.email,
                extract.inn,
                extract.kpp,
                int(extract.is_accredited),
                extract.fsa_registry_number,
                extract.org_type,
                source,
                now,
                now,
            ),
        )
        return int(cursor.lastrowid or 0)


def find_similar_organizations(
    name: str,
    *,
    min_ratio: float = 0.82,
    limit: int = 8,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    """
    Кандидаты-дубликаты по нормализованному имени (SequenceMatcher + substring).

    Каждый элемент: {id, name, name_normalized, org_type, inn, score}.
    Exact match (score=1.0) идёт первым.
    """
    from difflib import SequenceMatcher

    raw = (name or "").strip()
    if len(raw) < 2:
        return []
    key = normalize_org_name(raw)
    if not key:
        return []

    with get_connection(db_path) as conn:
        rows = conn.execute(
            "SELECT id, name, name_normalized, org_type, inn FROM organizations"
        ).fetchall()

    scored: list[dict[str, Any]] = []
    for row in rows:
        other_key = (row["name_normalized"] or normalize_org_name(row["name"] or "")).strip()
        if not other_key:
            continue
        if other_key == key:
            score = 1.0
        else:
            score = SequenceMatcher(None, key, other_key).ratio()
            if key in other_key or other_key in key:
                score = max(score, 0.88)
        if score >= min_ratio:
            scored.append(
                {
                    "id": int(row["id"]),
                    "name": row["name"],
                    "name_normalized": other_key,
                    "org_type": row["org_type"],
                    "inn": row["inn"],
                    "score": round(score, 3),
                }
            )
    scored.sort(key=lambda x: (-x["score"], x["name"] or ""))
    return scored[:limit]


def create_organization(
    *,
    name: str,
    address: str | None = None,
    postal_code: str | None = None,
    phone: str | None = None,
    email: str | None = None,
    inn: str | None = None,
    kpp: str | None = None,
    is_accredited: bool = False,
    fsa_registry_number: str | None = None,
    org_type: str = "unknown",
    source: str | None = "manual",
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """Создаёт организацию (CRUD «+ Добавить»). Без fuzzy — вызывающий решает."""
    extract = OrganizationExtract(
        name=name.strip(),
        address=address,
        legal_address=address,
        actual_address=address,
        postal_code=postal_code,
        phone=phone,
        email=email,
        inn=inn,
        kpp=kpp,
        is_accredited=is_accredited,
        fsa_registry_number=fsa_registry_number,
        org_type=org_type if org_type in (
            "manufacturer",
            "certification_body",
            "testing_center",
            "dealer",
            "unknown",
        ) else "unknown",
        role="unknown",
        confidence=1.0,
    )
    return upsert_organization(extract, source=source, db_path=db_path)


def save_organizations_from_extraction(
    organizations: list[OrganizationExtract],
    *,
    source: str | None = None,
    customer_name: str | None = None,
    manufacturer_name: str | None = None,
    customer_inn: str | None = None,
    customer_address: str | None = None,
    skip_own_lab: bool = True,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, int | None]:
    """
    Сохраняет организации из заявки; возвращает id заказчика и производителя.

    - Наша ИЛ (lab_profile / Кабель-Тест) **не** пишется в справочник и **не**
      становится customer/manufacturer.
    - customer_name / manufacturer_name из GUI (после правок оператора) —
      приоритетнее списка extract, если роли в extract пустые.
    - Нет fallback manufacturer_id = customer_id (ломало направления: ИЛ = «производитель»).
    """
    from ..generation.lab_profile import is_own_lab_name

    customer_id: int | None = None
    manufacturer_id: int | None = None

    def _skip(org: OrganizationExtract) -> bool:
        if not skip_own_lab:
            return False
        if is_own_lab_name(org.name):
            return True
        if org.org_type == "testing_center" and is_own_lab_name(org.name):
            return True
        return False

    # 1) Явные строки GUI — надёжный источник ролей после human-in-the-loop
    cust = (customer_name or "").strip()
    mfg = (manufacturer_name or "").strip()
    if cust and not (skip_own_lab and is_own_lab_name(cust)):
        cust_org = OrganizationExtract(
            name=cust,
            inn=customer_inn,
            address=customer_address,
            legal_address=customer_address,
            actual_address=customer_address,
            org_type="certification_body"
            if re.search(r"сертификац|фаер|fire\s*lab", cust, re.I)
            else "unknown",
            role="customer",
            confidence=0.95,
        )
        # если в extract уже есть customer с тем же именем — подтянуть тип/телефон
        for o in organizations:
            if o.role == "customer" and normalize_org_name(o.name) == normalize_org_name(cust):
                cust_org = o.model_copy(
                    update={
                        "name": cust,
                        "inn": customer_inn or o.inn,
                        "address": customer_address or o.address,
                        "role": "customer",
                    }
                )
                break
            if o.org_type == "certification_body" and normalize_org_name(
                o.name
            ) == normalize_org_name(cust):
                cust_org = o.model_copy(
                    update={
                        "name": cust,
                        "inn": customer_inn or o.inn,
                        "address": customer_address or o.address,
                        "role": "customer",
                    }
                )
                break
        customer_id = upsert_organization(cust_org, source=source, db_path=db_path)

    if mfg and not (skip_own_lab and is_own_lab_name(mfg)):
        mfg_org = OrganizationExtract(
            name=mfg,
            org_type="manufacturer",
            role="manufacturer",
            confidence=0.95,
        )
        for o in organizations:
            if o.role == "manufacturer" and normalize_org_name(o.name) == normalize_org_name(
                mfg
            ):
                mfg_org = o.model_copy(update={"name": mfg, "role": "manufacturer"})
                break
        manufacturer_id = upsert_organization(mfg_org, source=source, db_path=db_path)

    # 2) Остальные из extract (не ИЛ; роли customer/manufacturer если ещё не заданы)
    for org in organizations:
        if _skip(org):
            continue
        if org.role not in ("customer", "manufacturer", "dealer") and org.org_type not in (
            "manufacturer",
            "certification_body",
            "dealer",
        ):
            # testing_center чужой — можно сохранить справочно, без ролей заказа
            if org.org_type == "testing_center":
                continue
        org_id = upsert_organization(org, source=source, db_path=db_path)
        if org.role == "customer" and customer_id is None and not is_own_lab_name(org.name):
            customer_id = org_id
        if org.role == "manufacturer" and manufacturer_id is None and not is_own_lab_name(
            org.name
        ):
            manufacturer_id = org_id
        if (
            customer_id is None
            and org.org_type == "certification_body"
            and not is_own_lab_name(org.name)
        ):
            customer_id = org_id

    return {"customer_org_id": customer_id, "manufacturer_org_id": manufacturer_id}


def save_document_extraction(
    *,
    source_path: str,
    source_type: str,
    text: str,
    marks_count: int,
    customer_org_id: int | None = None,
    manufacturer_org_id: int | None = None,
    subject: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO document_extractions (
                source_path, source_type, customer_org_id, manufacturer_org_id,
                subject, raw_text_length, marks_count, extracted_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_path,
                source_type,
                customer_org_id,
                manufacturer_org_id,
                subject,
                len(text),
                marks_count,
                datetime.now().isoformat(),
            ),
        )
        return int(cursor.lastrowid or 0)


def list_organizations(
    search: str | None = None,
    org_type: str | None = None,
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    query = "SELECT * FROM organizations"
    params: list[Any] = []
    conditions: list[str] = []
    if search:
        conditions.append("(name LIKE ? OR inn LIKE ? OR address LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
    if org_type:
        conditions.append("org_type = ?")
        params.append(org_type)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY updated_at DESC LIMIT ?"
    params.append(limit)

    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def get_organization_by_id(
    org_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    with get_connection(db_path) as conn:
        row = conn.execute("SELECT * FROM organizations WHERE id = ?", (org_id,)).fetchone()
        return dict(row) if row else None


def update_organization(
    org_id: int,
    *,
    name: str,
    address: str | None = None,
    postal_code: str | None = None,
    phone: str | None = None,
    email: str | None = None,
    inn: str | None = None,
    kpp: str | None = None,
    is_accredited: bool = False,
    fsa_registry_number: str | None = None,
    org_type: str = "unknown",
    db_path: str | Path = DB_PATH_DEFAULT,
) -> bool:
    """Обновляет организацию по id (ручное редактирование в GUI)."""
    now = datetime.now().isoformat()
    name_normalized = normalize_org_name(name)
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            UPDATE organizations SET
                name = ?,
                name_normalized = ?,
                address = ?,
                postal_code = ?,
                phone = ?,
                email = ?,
                inn = ?,
                kpp = ?,
                is_accredited = ?,
                fsa_registry_number = ?,
                org_type = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                name.strip(),
                name_normalized,
                address,
                postal_code,
                phone,
                email,
                inn,
                kpp,
                int(is_accredited),
                fsa_registry_number,
                org_type,
                now,
                org_id,
            ),
        )
        return cursor.rowcount > 0


def get_last_document_extraction(
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    """Последняя обработанная заявка (для панели сводки в GUI)."""
    with get_connection(db_path) as conn:
        row = conn.execute(
            """
            SELECT d.*,
                   c.name AS customer_name,
                   m.name AS manufacturer_name
            FROM document_extractions d
            LEFT JOIN organizations c ON c.id = d.customer_org_id
            LEFT JOIN organizations m ON m.id = d.manufacturer_org_id
            ORDER BY d.extracted_at DESC
            LIMIT 1
            """
        ).fetchone()
        return dict(row) if row else None


def find_organization_id_by_name(
    name: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int | None:
    if not name or not name.strip():
        return None
    normalized = normalize_org_name(name)
    with get_connection(db_path) as conn:
        row = conn.execute(
            """
            SELECT id FROM organizations
            WHERE name_normalized = ? OR name = ?
            ORDER BY updated_at DESC
            LIMIT 1
            """,
            (normalized, name.strip()),
        ).fetchone()
        return int(row["id"]) if row else None


def _find_cable_mark_id(mark: str, db_path: str | Path) -> int | None:
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id FROM cable_marks WHERE full_mark = ? LIMIT 1",
            (mark,),
        ).fetchone()
        return int(row["id"]) if row else None


def create_order_from_kp(
    *,
    customer_name: str,
    manufacturer_name: str | None = None,
    customer_org_id: int | None = None,
    manufacturer_org_id: int | None = None,
    subject: str,
    note: str | None = None,
    calculation_ids: list[int],
    kp_output_path: str,
    document_extraction_id: int | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> int:
    """
    Создаёт заказ после формирования КП.
    Заказчик — на уровне заказа; каждая марка связана с производителем.
    """
    if not calculation_ids:
        raise ValueError("Нет расчётов для заказа")

    if customer_org_id is None and customer_name:
        customer_org_id = find_organization_id_by_name(customer_name, db_path)
    if manufacturer_org_id is None and manufacturer_name:
        manufacturer_org_id = find_organization_id_by_name(manufacturer_name, db_path)
    # HITL: если имена пусты/не совпали — id с подтверждённой заявки
    if document_extraction_id and (
        customer_org_id is None or manufacturer_org_id is None
    ):
        with get_connection(db_path) as conn:
            row = conn.execute(
                """
                SELECT customer_org_id, manufacturer_org_id
                FROM document_extractions WHERE id = ?
                """,
                (document_extraction_id,),
            ).fetchone()
            if row:
                if customer_org_id is None and row["customer_org_id"]:
                    customer_org_id = int(row["customer_org_id"])
                if manufacturer_org_id is None and row["manufacturer_org_id"]:
                    manufacturer_org_id = int(row["manufacturer_org_id"])

    rows = get_calculations_for_kp(calculation_ids, db_path=db_path)
    if not rows:
        raise ValueError("Расчёты не найдены")

    total_without = round(sum(float(r["total_cost_without_vat"]) for r in rows), 2)
    total_with = round(sum(float(r["total_cost_with_vat"]) for r in rows), 2)
    vat_rate = float(rows[0].get("vat_rate") or 0.22)
    now = datetime.now().isoformat()

    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO orders (
                customer_org_id, manufacturer_org_id, subject, note, status,
                total_without_vat, total_with_vat, vat_rate,
                document_extraction_id, kp_output_path, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 'kp_generated', ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                customer_org_id,
                manufacturer_org_id,
                subject,
                note,
                total_without,
                total_with,
                vat_rate,
                document_extraction_id,
                kp_output_path,
                now,
                now,
            ),
        )
        order_id = int(cursor.lastrowid or 0)

        for row in rows:
            calc_id = int(row["id"])
            mark = row["mark"]
            cable_mark_id = _find_cable_mark_id(mark, db_path)
            conn.execute(
                """
                INSERT INTO order_marks (
                    order_id, calculation_id, cable_mark_id, manufacturer_org_id,
                    mark, total_without_vat, total_with_vat
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    order_id,
                    calc_id,
                    cable_mark_id,
                    manufacturer_org_id,
                    mark,
                    float(row["total_cost_without_vat"]),
                    float(row["total_cost_with_vat"]),
                ),
            )

    save_generated_document(
        doc_type="kp",
        file_path=kp_output_path,
        order_id=order_id,
        db_path=db_path,
    )
    return order_id


def list_orders(
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    with get_connection(db_path) as conn:
        rows = conn.execute(
            """
            SELECT o.*,
                   c.name AS customer_name,
                   m.name AS manufacturer_name,
                   (SELECT COUNT(*) FROM order_marks om WHERE om.order_id = o.id) AS marks_count
            FROM orders o
            LEFT JOIN organizations c ON c.id = o.customer_org_id
            LEFT JOIN organizations m ON m.id = o.manufacturer_org_id
            ORDER BY o.created_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def get_order_details(
    order_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    with get_connection(db_path) as conn:
        order_row = conn.execute(
            """
            SELECT o.*,
                   c.name AS customer_name,
                   c.inn AS customer_inn,
                   c.address AS customer_address,
                   c.legal_address AS customer_legal_address,
                   c.actual_address AS customer_actual_address,
                   c.phone AS customer_phone,
                   c.email AS customer_email,
                   m.name AS manufacturer_name,
                   m.inn AS manufacturer_inn,
                   m.address AS manufacturer_address,
                   m.legal_address AS manufacturer_legal_address,
                   m.actual_address AS manufacturer_actual_address,
                   m.phone AS manufacturer_phone,
                   m.email AS manufacturer_email,
                   d.source_path AS source_document
            FROM orders o
            LEFT JOIN organizations c ON c.id = o.customer_org_id
            LEFT JOIN organizations m ON m.id = o.manufacturer_org_id
            LEFT JOIN document_extractions d ON d.id = o.document_extraction_id
            WHERE o.id = ?
            """,
            (order_id,),
        ).fetchone()
        if not order_row:
            return None

        marks = conn.execute(
            """
            SELECT om.*,
                   mo.name AS manufacturer_name
            FROM order_marks om
            LEFT JOIN organizations mo ON mo.id = om.manufacturer_org_id
            WHERE om.order_id = ?
            ORDER BY om.id
            """,
            (order_id,),
        ).fetchall()

        result = dict(order_row)
        result["marks"] = [dict(m) for m in marks]
        return result


def list_cable_marks(
    search: str | None = None,
    limit: int = 200,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    query = "SELECT * FROM cable_marks"
    params: list[Any] = []
    if search:
        query += " WHERE full_mark LIKE ? OR brand LIKE ?"
        params.extend([f"%{search}%", f"%{search}%"])
    query += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)

    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def get_cable_mark_by_id(
    mark_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any] | None:
    """Строка справочника cable_marks по id."""
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT * FROM cable_marks WHERE id = ?", (mark_id,)
        ).fetchone()
        return dict(row) if row else None


def update_cable_mark(
    mark_id: int,
    *,
    full_mark: str,
    brand: str,
    fire_class: str | None = None,
    cores_count: int = 1,
    structural_element_type: str | None = "жила",
    structural_elements_count: int | None = None,
    characteristic_size: float = 1.0,
    size_unit: str = "mm2",
    document: str | None = None,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any]:
    """Обновляет марку по id (редактор справочника).

    Returns:
        ``{"ok": True, ...}`` или ``{"ok": False, "reason": ...}``.
        reason: not_found | empty_mark | duplicate_mark | bad_size
    """
    designation = (full_mark or "").strip()
    if len(designation) < 2:
        return {"ok": False, "reason": "empty_mark"}
    brand_s = (brand or "").strip() or designation
    unit = size_unit if size_unit in ("mm2", "mm") else "mm2"
    try:
        cores = max(1, int(cores_count))
        size = float(characteristic_size)
        if size <= 0:
            return {"ok": False, "reason": "bad_size"}
    except (TypeError, ValueError):
        return {"ok": False, "reason": "bad_size"}
    elem_count = structural_elements_count
    if elem_count is None:
        elem_count = cores
    else:
        try:
            elem_count = max(1, int(elem_count))
        except (TypeError, ValueError):
            elem_count = cores

    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id, full_mark FROM cable_marks WHERE id = ?", (mark_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "reason": "not_found"}
        clash = conn.execute(
            "SELECT id FROM cable_marks WHERE full_mark = ? AND id != ?",
            (designation, mark_id),
        ).fetchone()
        if clash:
            return {
                "ok": False,
                "reason": "duplicate_mark",
                "other_id": int(clash["id"]),
            }
        conn.execute(
            """
            UPDATE cable_marks SET
                full_mark = ?,
                brand = ?,
                fire_class = ?,
                cores_count = ?,
                structural_element_type = ?,
                structural_elements_count = ?,
                characteristic_size = ?,
                size_unit = ?,
                document = ?
            WHERE id = ?
            """,
            (
                designation,
                brand_s,
                (fire_class or "").strip() or None,
                cores,
                (structural_element_type or "").strip() or "жила",
                elem_count,
                size,
                unit,
                (document or "").strip() or None,
                mark_id,
            ),
        )
        return {
            "ok": True,
            "id": mark_id,
            "full_mark": designation,
            "previous_full_mark": row["full_mark"],
        }


def delete_cable_mark(
    mark_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
    *,
    force: bool = False,
) -> dict[str, Any]:
    """Удаляет марку из справочника.

    Если force=False и марка в order_marks — отказ (blocked).
    Если force=True — обнуляет cable_mark_id в order_marks, затем удаляет.
    """
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id, full_mark FROM cable_marks WHERE id = ?", (mark_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "reason": "not_found"}
        refs = conn.execute(
            "SELECT COUNT(*) AS n FROM order_marks WHERE cable_mark_id = ?",
            (mark_id,),
        ).fetchone()["n"]
        if refs and not force:
            return {
                "ok": False,
                "reason": "in_use",
                "refs": int(refs),
                "full_mark": row["full_mark"],
            }
        if refs and force:
            conn.execute(
                "UPDATE order_marks SET cable_mark_id = NULL WHERE cable_mark_id = ?",
                (mark_id,),
            )
        conn.execute("DELETE FROM cable_marks WHERE id = ?", (mark_id,))
        return {"ok": True, "full_mark": row["full_mark"], "unlinked": int(refs)}


def delete_calculation(
    calc_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> dict[str, Any]:
    """Удаляет расчёт и его строки. order_marks.calculation_id → NULL."""
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id, mark FROM calculations WHERE id = ?", (calc_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "reason": "not_found"}
        conn.execute(
            "UPDATE order_marks SET calculation_id = NULL WHERE calculation_id = ?",
            (calc_id,),
        )
        conn.execute("DELETE FROM calculation_lines WHERE calculation_id = ?", (calc_id,))
        conn.execute(
            "UPDATE generated_documents SET calculation_id = NULL WHERE calculation_id = ?",
            (calc_id,),
        )
        conn.execute("DELETE FROM calculations WHERE id = ?", (calc_id,))
        return {"ok": True, "mark": row["mark"]}


def delete_generated_document(
    doc_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
    *,
    delete_file: bool = False,
) -> dict[str, Any]:
    """Удаляет запись КП/документа из generated_documents (опц. файл с диска)."""
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id, doc_type, file_path FROM generated_documents WHERE id = ?",
            (doc_id,),
        ).fetchone()
        if not row:
            return {"ok": False, "reason": "not_found"}
        path = row["file_path"]
        conn.execute("DELETE FROM generated_documents WHERE id = ?", (doc_id,))
    removed_file = False
    if delete_file and path:
        try:
            p = Path(path)
            if p.is_file():
                p.unlink()
                removed_file = True
        except OSError:
            pass
    return {
        "ok": True,
        "doc_type": row["doc_type"],
        "file_path": path,
        "file_removed": removed_file,
    }


def delete_order(
    order_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
    *,
    cascade: bool = False,
) -> dict[str, Any]:
    """Удаляет заказ.

    cascade=False: отказ, если есть order_marks / applications / generated.
    cascade=True: удаляет дочерние записи (не трогает calculations целиком —
    только связи order_marks); файлы на диске не удаляет.
    """
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id, customer_name FROM orders WHERE id = ?", (order_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "reason": "not_found"}
        marks_n = conn.execute(
            "SELECT COUNT(*) AS n FROM order_marks WHERE order_id = ?", (order_id,)
        ).fetchone()["n"]
        apps_n = conn.execute(
            "SELECT COUNT(*) AS n FROM test_applications WHERE order_id = ?",
            (order_id,),
        ).fetchone()["n"]
        docs_n = conn.execute(
            "SELECT COUNT(*) AS n FROM generated_documents WHERE order_id = ?",
            (order_id,),
        ).fetchone()["n"]
        children = int(marks_n) + int(apps_n) + int(docs_n)
        if children and not cascade:
            return {
                "ok": False,
                "reason": "has_children",
                "order_marks": int(marks_n),
                "applications": int(apps_n),
                "generated": int(docs_n),
                "customer_name": row["customer_name"],
            }
        conn.execute("DELETE FROM order_marks WHERE order_id = ?", (order_id,))
        conn.execute("DELETE FROM test_applications WHERE order_id = ?", (order_id,))
        conn.execute(
            "UPDATE generated_documents SET order_id = NULL WHERE order_id = ?",
            (order_id,),
        )
        conn.execute("DELETE FROM orders WHERE id = ?", (order_id,))
        return {
            "ok": True,
            "customer_name": row["customer_name"],
            "removed_marks": int(marks_n),
            "removed_applications": int(apps_n),
            "unlinked_generated": int(docs_n),
        }


def delete_organization(
    org_id: int,
    db_path: str | Path = DB_PATH_DEFAULT,
    *,
    force: bool = False,
) -> dict[str, Any]:
    """Удаляет организацию. force: обнуляет FK в orders/order_marks/extractions."""
    with get_connection(db_path) as conn:
        row = conn.execute(
            "SELECT id, name FROM organizations WHERE id = ?", (org_id,)
        ).fetchone()
        if not row:
            return {"ok": False, "reason": "not_found"}
        o_refs = conn.execute(
            """
            SELECT COUNT(*) AS n FROM orders
            WHERE customer_org_id = ? OR manufacturer_org_id = ?
            """,
            (org_id, org_id),
        ).fetchone()["n"]
        m_refs = conn.execute(
            "SELECT COUNT(*) AS n FROM order_marks WHERE manufacturer_org_id = ?",
            (org_id,),
        ).fetchone()["n"]
        e_refs = conn.execute(
            """
            SELECT COUNT(*) AS n FROM document_extractions
            WHERE customer_org_id = ? OR manufacturer_org_id = ?
            """,
            (org_id, org_id),
        ).fetchone()["n"]
        total = int(o_refs) + int(m_refs) + int(e_refs)
        if total and not force:
            return {
                "ok": False,
                "reason": "in_use",
                "refs": total,
                "name": row["name"],
            }
        if force:
            conn.execute(
                "UPDATE orders SET customer_org_id = NULL WHERE customer_org_id = ?",
                (org_id,),
            )
            conn.execute(
                "UPDATE orders SET manufacturer_org_id = NULL WHERE manufacturer_org_id = ?",
                (org_id,),
            )
            conn.execute(
                "UPDATE order_marks SET manufacturer_org_id = NULL WHERE manufacturer_org_id = ?",
                (org_id,),
            )
            conn.execute(
                "UPDATE document_extractions SET customer_org_id = NULL WHERE customer_org_id = ?",
                (org_id,),
            )
            conn.execute(
                "UPDATE document_extractions SET manufacturer_org_id = NULL WHERE manufacturer_org_id = ?",
                (org_id,),
            )
        conn.execute("DELETE FROM organizations WHERE id = ?", (org_id,))
        return {"ok": True, "name": row["name"], "unlinked": total}


# Минимальный размер «полного» прайса (xlsx/seed ≈ 60). Ниже — считаем каталог урезанным.
_PRICE_CATALOG_MIN_COUNT = 20
_PRICE_SEED_JSON = Path(__file__).resolve().parent / "price_catalog_seed.json"
_MAPPINGS_SEED_JSON = Path(__file__).resolve().parent / "test_mappings_seed.json"


def _find_price_xlsx() -> Path | None:
    """Ищет xlsx прайса рядом с проектом (кириллическое имя допустимо)."""
    data = Path(PROJECT_ROOT) / "data"
    if not data.is_dir():
        return None
    preferred = data / "Обновленная стоимость на 2026 год.xlsx"
    if preferred.is_file():
        return preferred
    for p in sorted(data.glob("*.xlsx")):
        return p
    return None


def load_price_catalog_from_seed(
    db_path: str | Path = DB_PATH_DEFAULT,
    seed_path: str | Path | None = None,
) -> int:
    """Загружает test_items из встроенного JSON (релиз без xlsx / без app.db)."""
    path = Path(seed_path) if seed_path else _PRICE_SEED_JSON
    if not path.is_file():
        return 0
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        return 0
    count = 0
    for row in raw:
        if not isinstance(row, dict) or not row.get("code"):
            continue
        item = TestItem(
            code=str(row["code"]),
            name=str(row.get("name") or row["code"]),
            base_cost=float(row.get("base_cost") or 0),
            category=row.get("category"),
            method=row.get("method"),
            rule_type=row.get("rule_type") or "fixed",  # type: ignore[arg-type]
            rule_params=row.get("rule_params") or {},
        )
        insert_test_item(item, db_path)
        count += 1
    return count


def load_test_mappings_from_seed(
    db_path: str | Path = DB_PATH_DEFAULT,
    seed_path: str | Path | None = None,
) -> int:
    """Дополняет test_mappings из JSON (идемпотентно по requirement_pattern)."""
    path = Path(seed_path) if seed_path else _MAPPINGS_SEED_JSON
    if not path.is_file():
        return 0
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        return 0
    now = datetime.now().isoformat()
    added = 0
    with get_connection(db_path) as conn:
        for row in raw:
            if not isinstance(row, dict):
                continue
            pattern = (row.get("requirement_pattern") or "").strip().lower()
            code = (row.get("test_code") or "").strip()
            if not pattern or not code:
                continue
            cur = conn.execute(
                """
                INSERT INTO test_mappings (requirement_pattern, test_code, note, usage_count, created_at)
                VALUES (?, ?, ?, 0, ?)
                ON CONFLICT(requirement_pattern) DO NOTHING
                """,
                (pattern, code, row.get("note"), now),
            )
            if cur.rowcount:
                added += 1
    return added


def ensure_price_catalog(db_path: str | Path = DB_PATH_DEFAULT) -> dict[str, Any]:
    """
    Гарантирует полный справочник испытаний в test_items.

    На чистой установке без -IncludeAppDb раньше попадали только 3 демо-кода
    (resistance_core …) — на рабочем ПК «пропадал» почти весь Справочник.

    Порядок:
    1) если записей уже >= порога — только sync климатики;
    2) иначе xlsx из data/ (если есть);
    3) иначе встроенный price_catalog_seed.json;
    4) в крайнем случае — 3 демо-позиции.
    """
    existing = get_all_test_items(db_path)
    n = len(existing)
    result: dict[str, Any] = {
        "before": n,
        "loaded": 0,
        "source": None,
        "after": n,
    }
    if n >= _PRICE_CATALOG_MIN_COUNT:
        sync_climatic_tests(db_path)
        result["source"] = "already_full"
        _log.debug(
            "ensure_price_catalog skip (already_full) n=%s path=%s",
            n,
            db_path,
            extra={"tag": "Прайс"},
        )
        return result

    _log.warning(
        "ensure_price_catalog: catalog thin n=%s < %s path=%s — restoring",
        n,
        _PRICE_CATALOG_MIN_COUNT,
        db_path,
        extra={"tag": "Прайс"},
    )

    # Убрать старые EN-демо-коды, если они остались вместо реального прайса
    demo_codes = ("resistance_core", "insulation_resistance", "voltage_test")
    with get_connection(db_path) as conn:
        for code in demo_codes:
            conn.execute("DELETE FROM test_items WHERE code = ?", (code,))

    xlsx = _find_price_xlsx()
    if xlsx is not None:
        try:
            result["loaded"] = load_price_list_from_xlsx(xlsx, db_path)
            result["source"] = f"xlsx:{xlsx.name}"
            _log.info(
                "price loaded from xlsx=%s count=%s",
                xlsx,
                result["loaded"],
                extra={"tag": "Прайс"},
            )
        except Exception as exc:
            _log.exception(
                "price xlsx load failed path=%s: %s",
                xlsx,
                exc,
                extra={"tag": "Прайс"},
            )
            result["loaded"] = 0
            result["source"] = None

    if result["loaded"] < _PRICE_CATALOG_MIN_COUNT:
        seeded = load_price_catalog_from_seed(db_path)
        if seeded:
            result["loaded"] = seeded
            result["source"] = "seed_json"
            _log.info(
                "price loaded from seed_json count=%s",
                seeded,
                extra={"tag": "Прайс"},
            )

    if result["loaded"] == 0 and len(get_all_test_items(db_path)) < 5:
        for item in (
            TestItem(
                code="электрическое_сопротивление_тпж",
                name="Электрическое сопротивление ТПЖ",
                base_cost=400,
                category="Электрические параметры НЧ",
                rule_type="per_core",
            ),
            TestItem(
                code="электрическое_сопротивление_изоляции_тпж",
                name="Электрическое сопротивление изоляции ТПЖ",
                base_cost=600,
                category="Электрические параметры НЧ",
                rule_type="per_core",
            ),
            TestItem(
                code="испытание_напряжением",
                name="Испытание напряжением",
                base_cost=400,
                category="Электрические параметры НЧ",
            ),
        ):
            insert_test_item(item, db_path)
            result["loaded"] += 1
        result["source"] = "demo_fallback"
        _log.warning(
            "price demo_fallback only count=%s",
            result["loaded"],
            extra={"tag": "Прайс"},
        )

    sync_climatic_tests(db_path)
    apply_price_catalog_fixes(db_path)
    # Маппинги: если почти пусто — подтянуть seed
    with get_connection(db_path) as conn:
        map_n = conn.execute("SELECT COUNT(*) AS n FROM test_mappings").fetchone()["n"]
    if int(map_n) < 30:
        added_m = load_test_mappings_from_seed(db_path)
        _log.info(
            "test_mappings seed added=%s (was %s)",
            added_m,
            map_n,
            extra={"tag": "Прайс"},
        )

    result["after"] = len(get_all_test_items(db_path))
    _log.info(
        "ensure_price_catalog done before=%s after=%s source=%s",
        result["before"],
        result["after"],
        result["source"],
        extra={"tag": "Прайс"},
    )
    return result


def _seed_demo_tests(db_path: str | Path) -> None:
    """Обратная совместимость: полный прайс через ensure_price_catalog."""
    ensure_price_catalog(db_path)


def insert_test_item(item: TestItem, db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Вставляет или обновляет элемент прайс-листа по коду."""
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO test_items (code, name, base_cost, category, method, rule_type, rule_params)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                name=excluded.name,
                base_cost=excluded.base_cost,
                category=excluded.category,
                method=excluded.method,
                rule_type=excluded.rule_type,
                rule_params=excluded.rule_params
            """,
            (
                item.code,
                item.name,
                item.base_cost,
                item.category,
                item.method,
                item.rule_type,
                json.dumps(item.rule_params, ensure_ascii=False),
            ),
        )
        return cursor.lastrowid or 0


def get_test_item_by_code(code: str, db_path: str | Path = DB_PATH_DEFAULT) -> TestItem | None:
    """Получает TestItem по коду (с алиасом EN-климатики → slug)."""
    resolved = CLIMATE_ITEM_ALIASES.get(code, code)
    with get_connection(db_path) as conn:
        row = conn.execute("SELECT * FROM test_items WHERE code = ?", (resolved,)).fetchone()
        if not row:
            return None
        data = dict(row)
        data["rule_params"] = json.loads(data.get("rule_params") or "{}")
        return TestItem(**data)


def get_all_test_items(db_path: str | Path = DB_PATH_DEFAULT) -> list[TestItem]:
    """Возвращает все тесты (для отладки и load-data)."""
    with get_connection(db_path) as conn:
        rows = conn.execute("SELECT * FROM test_items ORDER BY id").fetchall()
        return [
            TestItem(**{**dict(row), "rule_params": json.loads(row["rule_params"] or "{}")})
            for row in rows
        ]


def save_calculation(calc: Calculation, db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Сохраняет расчёт + все строки детализации. Возвращает id расчёта."""
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO calculations
                (mark, parsed_mark, total_cost_without_vat, vat_rate,
                 total_cost_with_vat, source, output_path, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                calc.mark,
                calc.parsed_mark.model_dump_json(),
                calc.total_cost_without_vat,
                calc.vat_rate,
                calc.total_cost_with_vat,
                calc.source,
                calc.output_path,
                calc.created_at.isoformat(),
            ),
        )
        calc_id = cursor.lastrowid

        for line in calc.lines:
            conn.execute(
                """
                INSERT INTO calculation_lines
                    (calculation_id, test_item_id, test_name, base_cost,
                     multiplier, quantity, hours, final_cost, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    calc_id,
                    line.test_item_id,
                    line.test_name,
                    line.base_cost,
                    line.multiplier,
                    line.quantity,
                    line.hours,
                    line.final_cost,
                    line.note,
                ),
            )
        return calc_id or 0


def get_recent_calculations(limit: int = 10, db_path: str | Path = DB_PATH_DEFAULT) -> list[dict[str, Any]]:
    """История последних расчётов (для команды history)."""
    with get_connection(db_path) as conn:
        rows = conn.execute(
            """
            SELECT id, mark, total_cost_with_vat, source, created_at
            FROM calculations ORDER BY created_at DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def get_calculations_for_kp(
    calculation_ids: list[int] | None = None,
    limit: int = 100,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict[str, Any]]:
    """Возвращает расчёты для формирования КП (с итогами по марке)."""
    with get_connection(db_path) as conn:
        if calculation_ids:
            placeholders = ",".join("?" * len(calculation_ids))
            rows = conn.execute(
                f"""
                SELECT id, mark, total_cost_without_vat, total_cost_with_vat,
                       vat_rate, source, created_at, output_path
                FROM calculations
                WHERE id IN ({placeholders})
                ORDER BY id
                """,
                calculation_ids,
            ).fetchall()
            return [dict(row) for row in rows]
        rows = conn.execute(
            """
            SELECT id, mark, total_cost_without_vat, total_cost_with_vat,
                   vat_rate, source, created_at, output_path
            FROM calculations ORDER BY created_at DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [dict(row) for row in rows]


def update_calculation_output_path(
    calculation_id: int,
    output_path: str,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> None:
    with get_connection(db_path) as conn:
        conn.execute(
            "UPDATE calculations SET output_path = ? WHERE id = ?",
            (output_path, calculation_id),
        )



'''---Загрузка прайс-листа из Excel (минимальная, версия)---'''


try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def _slugify(text: str) -> str:
    """Делает короткий код из русского названия теста."""
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "_", text)
    return text[:55]


def load_price_list_from_xlsx(
    xlsx_path: str | Path,
    db_path: str | Path = DB_PATH_DEFAULT,
    sheet_name: str = "Стоимость",
) -> int:
    """
    Загружает прайс-лист в таблицу test_items.
    
    Использует openpyxl. Берёт:
    - B (Наименование) → name + code (slug)
    - D (Стоимость) → base_cost
    - F (Категория)
    - G (Метод)
    
    rule_type определяется автоматически (per_core, per_group, time_based, fixed).
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl не установлен")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    count = 0
    with get_connection(db_path) as conn:
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[1]:
                continue
            name = str(row[1]).strip()
            if not name or name.lower() == "наименование":
                continue

            try:
                base_cost = float(row[3]) if row[3] is not None else 0.0
            except (TypeError, ValueError):
                base_cost = 0.0

            category = str(row[5]).strip() if row[5] else None
            method = str(row[6]).strip() if row[6] else None

            code = _slugify(name) or f"test_{idx}"
            rule_type, rule_params = infer_rule_type(name, category, code)

            item = TestItem(
                code=code,
                name=name,
                base_cost=base_cost,
                category=category,
                method=method,
                rule_type=rule_type,  # type: ignore[arg-type]
                rule_params=rule_params,
            )
            insert_test_item(item, db_path)
            count += 1

    print(f"Загружено {count} позиций прайс-листа")
    return count

'''---Новые функции управления справочником испытаний (Итерация 2)---'''

def add_test_item(item: TestItemCreate, db_path: str | Path = DB_PATH_DEFAULT) -> int:
    """Добавляет новое испытание в справочник test_items.
    
    Использует Pydantic-модель TestItemCreate для валидации входных данных.
    Если code уже существует — SQLite выбросит IntegrityError.
    
    Args:
        item: Валидированная модель с данными испытания
        db_path: Путь к базе данных
        
    Returns:
        id созданной записи в БД
    """
    with get_connection(db_path) as conn:
        cursor = conn.execute(
            """
            INSERT INTO test_items 
            (code, name, base_cost, category, method, rule_type, rule_params)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item.code,
                item.name,
                item.base_cost,
                item.category,
                item.method,
                item.rule_type,
                json.dumps(item.rule_params, ensure_ascii=False),
            ),
        )
        return cursor.lastrowid or 0


def update_test_item(
    code: str, 
    updates: TestItemUpdate, 
    db_path: str | Path = DB_PATH_DEFAULT
) -> bool:
    """Частично обновляет испытание по коду.
    
    Обновляет только те поля, которые реально были переданы в модели updates.
    Использует exclude_unset=True внутри Pydantic.
    
    Args:
        code: Уникальный код испытания (например, 'temp_low')
        updates: Модель TestItemUpdate с полями для обновления
        db_path: Путь к БД
        
    Returns:
        True, если была обновлена хотя бы одна запись
    """
    # Получаем только те поля, которые пользователь реально передал
    update_data = updates.model_dump(exclude_unset=True)
    if not update_data:
        return False

    # Если обновляем rule_params — сериализуем в JSON
    if "rule_params" in update_data and update_data["rule_params"] is not None:
        update_data["rule_params"] = json.dumps(update_data["rule_params"], ensure_ascii=False)

    set_clause = ", ".join(f"{k} = ?" for k in update_data.keys())
    values = list(update_data.values()) + [code]

    with get_connection(db_path) as conn:
        cursor = conn.execute(
            f"UPDATE test_items SET {set_clause} WHERE code = ?", 
            values
        )
        return cursor.rowcount > 0


def list_test_items(
    category: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 200,
    db_path: str | Path = DB_PATH_DEFAULT,
) -> list[dict]:
    """Возвращает список испытаний с фильтрацией (для команды list-tests в CLI).
    
    Поддерживает фильтр по категории и поиск по названию/коду.
    """
    query = "SELECT * FROM test_items"
    params: list[Any] = []

    conditions: list[str] = []
    if category:
        conditions.append("category = ?")
        params.append(category)
    if search:
        conditions.append("(name LIKE ? OR code LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY category, name LIMIT ?"
    params.append(limit)

    with get_connection(db_path) as conn:
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]


def bulk_upsert_test_items(
    items: list[TestItemCreate], 
    db_path: str | Path = DB_PATH_DEFAULT
) -> dict[str, int]:
    """Пакетное добавление/обновление испытаний (для команды import-tests).
    
    Использует INSERT OR REPLACE — если code уже есть, запись обновляется.
    Это удобно при повторной загрузке прайс-листа или его части.
    
    Returns:
        Статистика: сколько обработано и сколько было ошибок
    """
    stats = {"processed": 0, "errors": 0}

    with get_connection(db_path) as conn:
        for item in items:
            try:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO test_items 
                    (code, name, base_cost, category, method, rule_type, rule_params)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item.code,
                        item.name,
                        item.base_cost,
                        item.category,
                        item.method,
                        item.rule_type,
                        json.dumps(item.rule_params, ensure_ascii=False),
                    ),
                )
                stats["processed"] += 1
            except Exception:
                stats["errors"] += 1

        conn.commit()

    return stats